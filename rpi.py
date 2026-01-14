#!/usr/bin/env python3
"""
GDS - Vessel Management System (Black/Red Theme)
Fixes/Features added (without breaking previous professional UI):
- Camera buttons functional even without camera (demo-safe)
- Demo camera stream returns a real JPEG frame (img tag works)
- Show captured images in VJR & VDR pages (gallery)
- Embed captured images in VJR PDF and VDR PDF
- Add PDF export button to BOTH VJR and VDR
- Keep live map in NAV, VJR, VDR using dummy ST6100-like data
- User roles: Captain / Operator / Viewer
- PTZ sync to vessel COG (toggle)
"""
OFFLINE_MODE = True
CONNECTIVITY_UI = True
SYNC_ENABLED = True
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from flask import (
    Flask, render_template_string, jsonify, request, Response,
    send_file, redirect, url_for, session
)
from flask_cors import CORS
import threading, time, random, math, json, requests, base64, os, uuid
import sqlite3

from datetime import datetime
from io import BytesIO, StringIO
import csv
from functools import wraps
# ---------- Firebase (Realtime DB + Storage) ----------
FIREBASE_ENABLED = False

if FIREBASE_ENABLED:
    import firebase_admin
    from firebase_admin import credentials, db, storage


# Image generation for demo camera frames
from PIL import Image, ImageDraw, ImageFont

try:
    import openpyxl
    from openpyxl.styles import Font as XLFont, PatternFill
except ImportError:
    openpyxl = None

# Optional PDF - ReportLab (Windows-friendly, pure Python)
try:
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

# WeasyPrint (fallback)
try:
    from weasyprint import HTML
    WEASYPRINT_AVAILABLE = True
except Exception:
    WEASYPRINT_AVAILABLE = False

# ------------------------------------------------------------------------------
# CONFIGURATION
# ------------------------------------------------------------------------------

IP_CAMERA_URL = "http://192.168.0.164:8080/video"
CAMERA_USERNAME = ""
CAMERA_PASSWORD = ""
AUTO_FALLBACK_TO_DEMO = True
CAMERA_CONNECTION_TIMEOUT = 5
# ---------- Firebase Config ----------
FIREBASE_KEY_PATH = os.path.join(os.path.dirname(__file__), "firebase_key.json")
FIREBASE_DB_URL = "https://gds-vessel-simulator-default-rtdb.asia-southeast1.firebasedatabase.app"
FIREBASE_STORAGE_BUCKET = "gds-vessel-simulator.appspot.com"  # usually <project-id>.appspot.com
VESSEL_ID = "demo_vessel"
# Local offline database
LOCAL_DB_ENABLED = True
LOCAL_DB_PATH = "/home/rpi2/ship_system/db/ship_data.db"

# Logo file in same folder
LOGO_FILE = os.path.join(os.path.dirname(__file__), "GDS Logo.jpg")

# Embedded fallback logo PNG (small placeholder)
FALLBACK_LOGO_PNG_BASE64 = (
    "iVBORw0KGgoAAAANSUhEUgAAAJYAAAAqCAYAAABkGxjYAAAACXBIWXMAAAsSAAALEgHS3X78AAAB"
    "x0lEQVR4nO3bMW7TQBiG4d8gBqQ0aUqkY0iQpDkA0mZQwB0hUQqg5cI3o0FQj9i5hQ9vZtqGQp7"
    "3m8qfZc9m1uW2yQAAAAAAAAAAAAAAAAAAAAAAAPg3y0xg1uC7h5cJtZ0fZp7x3b0c3g8x2oR6B"
    "mZq2Qb6i7m0d0m0c0m9sQyq7K3+fZb7cE2nqQWmP8f1g5s9qG2d2Q8j3l9pQqYwq0WmZc6m8o2"
    "o8w1b3q8v2y+Yh6b0w0m0fQk8x5lqz9xw0m0VQm8xqf3m1b0m0VQm8xqf3m1b0m0VQm8xqf3m1"
    "b0m0VQm8xqf3m1b0m0VQm8xqf3m1b0m0VQm8xqf3m1b0m0VQm8xqf3m1b0m0VQm8xqf3m1b0m0"
    "VQm8xqf3m1b0m0VQm8xqf3m1b0m0VQm8xqf3m1b0m0VQm8xqf3m1b0m0VQm8xqf3m0AAAAAAAA"
    "AAAAAAAAAAAAAAADwp/0B7c7f0w3zVnQAAAAASUVORK5CYII="
)

# Simple local users (change before production)
USERS = {
    "captain":  {"password": "captain123",  "role": "Captain"},
    "operator": {"password": "operator123", "role": "Operator"},
    "viewer":   {"password": "viewer123",   "role": "Viewer"},
}
ROLE_ORDER = {"Viewer": 1, "Operator": 2, "Captain": 3}

app = Flask(__name__)
app.secret_key = "CHANGE_THIS_TO_A_LONG_RANDOM_SECRET"
CORS(app)

# ------------------------------------------------------------------------------
# GLOBAL STATE
# ------------------------------------------------------------------------------

camera_status = {
    "connected": False, "mode": "demo", "url": IP_CAMERA_URL,
    "error": None, "last_check": None
}
camera_controls = {
    "pan": 0, "tilt": 0, "zoom": 1.0,
    "led_brightness": 50, "led_enabled": False,
    "night_vision": False, "autofocus": True, "white_balance": "auto"
}
captured_images = {"vjr_images": [], "vdr_images": []}

nav_lock = threading.Lock()
weather_lock = threading.Lock()
vdr_lock = threading.Lock()
camera_lock = threading.Lock()
image_lock = threading.Lock()
sync_lock = threading.Lock()

nav_history = []
nav_current = {
    "latitude": None,
    "longitude": None,
    "heading": None,
    "timestamp": None
}

weather_current = {
    "wind_speed": None,
    "wind_direction": None,
    "temperature": None,
    "humidity": None,
    "pressure": None,
    "noise": None,
    "pm25": None,
    "pm10": None,
    "rainfall": None,
    "timestamp": None
}

vdr_records = []

sim_lat, sim_lon, sim_heading, sim_speed = 3.006633, 101.380133, 45.0, 8.0

# PTZ sync to COG
ptz_sync_enabled = False

MAX_NAV_HISTORY = 800
MAX_CAPTURE_IMAGES = 50
# ---------- Firebase init (run once) ----------
_firebase_ready = False
_firebase_lock = threading.Lock()

def init_firebase():
    """Initialize firebase_admin once (DB + Storage). Safe to call multiple times."""
    global _firebase_ready
    if not FIREBASE_ENABLED:
        return False

    with _firebase_lock:
        if _firebase_ready:
            return True

        try:
            if not os.path.exists(FIREBASE_KEY_PATH):
                print("Firebase key not found:", FIREBASE_KEY_PATH)
                return False

            cred = credentials.Certificate(FIREBASE_KEY_PATH)

            # ?? IMPORTANT FIX
            if not firebase_admin._apps:
                firebase_admin.initialize_app(cred, {
                    "databaseURL": FIREBASE_DB_URL,
                    "storageBucket": FIREBASE_STORAGE_BUCKET,
                })

            _firebase_ready = True
            print("Firebase initialized OK.")
            return True

        except Exception as e:
            print("Firebase init error:", e)
            return False


def firebase_weather_listener():
    global weather_current

    if not init_firebase():
        return

    ref = db.reference("gds_vessel_sensor_data/weather/current")

    def _on_change(event):
        if not event.data:
            return

        with weather_lock:
            weather_current.update({
                "wind_speed": event.data.get("wind_speed"),
                "wind_direction": event.data.get("wind_dir"),
                "temperature": event.data.get("temperature"),
                "humidity": event.data.get("humidity"),
                "pressure": event.data.get("pressure"),
                "noise": event.data.get("noise"),
                "pm25": event.data.get("pm25"),
                "pm10": event.data.get("pm10"),
                "rainfall": event.data.get("rainfall"),
                "timestamp": event.data.get("timestamp"),
            })

        print("?? Weather updated:", weather_current)

    ref.listen(_on_change)

def firebase_marinelite_listener():
    global nav_current

    print("?? Starting MarineLite Firebase listener...")

    if not init_firebase():
        print("? Firebase init failed (MarineLite)")
        return

    try:
        ref = db.reference("gds_vessel_sensor_data/marinelite/current")
    except Exception as e:
        print("? Firebase reference error:", e)
        return

    def _on_change(event):
        global nav_current

        if not event.data:
            print("?? MarineLite event with no data")
            return

        print("?? Raw MarineLite Firebase data:", event.data)

        with nav_lock:
            nav_current["latitude"]  = event.data.get("latitude")
            nav_current["longitude"] = event.data.get("longitude")
            nav_current["heading"]   = event.data.get("heading")
            nav_current["timestamp"] = event.data.get("timestamp")

        print("? NAV updated:", nav_current)

    ref.listen(_on_change)


def upload_jpeg_to_firebase_storage(jpg_bytes: bytes, report_type: str):
    """
    Upload bytes to Firebase Storage and return (public_url, object_path).
    Note: For production, prefer signed URLs or authenticated access; public is fine for testing.
    """
    if not init_firebase():
        return None, None

    ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    uid = uuid.uuid4().hex[:10]
    report_type = (report_type or "vdr").lower()
    object_path = f"vessels/{VESSEL_ID}/captures/{report_type}/{ts}_{uid}.jpg"

    bucket = storage.bucket()
    blob = bucket.blob(object_path)
    blob.upload_from_string(jpg_bytes, content_type="image/jpeg")

    # For testing convenience (public). You can harden later.
    blob.make_public()
    return blob.public_url, object_path

def _db_connect():
    con = sqlite3.connect(LOCAL_DB_PATH, timeout=30, check_same_thread=False)
    con.execute("PRAGMA journal_mode=WAL;")
    return con


def db_get_latest_marinelite():
    con = sqlite3.connect(LOCAL_DB_PATH)
    row = con.execute("""
        SELECT ts, latitude, longitude, heading
        FROM nav_data
        ORDER BY id DESC LIMIT 1
    """).fetchone()
    con.close()

    if not row:
        return {}

    return {
        "timestamp": row[0],
        "latitude": row[1],
        "longitude": row[2],
        "heading": row[3]
    }




def db_get_latest_weather():
    con = sqlite3.connect(LOCAL_DB_PATH)
    row = con.execute("""
        SELECT ts, wind_speed, wind_dir, humidity, temperature,
               pressure, pm25, pm10, rainfall, noise
        FROM weather_data
        ORDER BY id DESC LIMIT 1
    """).fetchone()
    con.close()

    if not row:
        return {}

    return {
        "timestamp": row[0],
        "wind_speed": row[1],
        "wind_dir": row[2],
        "humidity": row[3],
        "temperature": row[4],
        "pressure": row[5],
        "pm25": row[6],
        "pm10": row[7],
        "rainfall": row[8],
        "noise": row[9]
    }


def push_capture_event_to_firebase(report_type: str, image_url: str, object_path: str, enc_cfg: dict):
    """Store capture metadata + NAV snapshot in Realtime DB."""
    if not init_firebase():
        return False

    # snapshot nav at capture time
    with nav_lock:
        nav_snapshot = nav_current.copy()

    payload = {
        "timestamp": datetime.utcnow().isoformat(),
        "vessel_id": VESSEL_ID,
        "report_type": (report_type or "vdr").lower(),
        "image_url": image_url,
        "storage_path": object_path,
        "encoding": enc_cfg,
        "nav": {
            "latitude": nav_snapshot.get("latitude"),
            "longitude": nav_snapshot.get("longitude"),
            "speed": nav_snapshot.get("speed"),
            "heading": nav_snapshot.get("heading"),
            "cog": nav_snapshot.get("cog"),
            "voltage": nav_snapshot.get("voltage"),
            "time": nav_snapshot.get("time"),
            "date": nav_snapshot.get("date"),
        },
        "source": "IP_CAMERA",
    }

    db.reference(f"vessels/{VESSEL_ID}/captures").push(payload)
    return True

# Navigation logging (persistent)
NAV_LOG_FILE = os.path.join(os.path.dirname(__file__), "nav_log.csv")
nav_log_lock = threading.Lock()

def _ensure_nav_log_header():
    """Create nav log file with header if it does not exist."""
    try:
        if not os.path.exists(NAV_LOG_FILE) or os.path.getsize(NAV_LOG_FILE) == 0:
            with nav_log_lock:
                with open(NAV_LOG_FILE, "a", newline="", encoding="utf-8") as f:
                    w = csv.writer(f)
                    w.writerow(["date","time","latitude","longitude","speed","cog","heading","voltage","panic","ext_heading","raw_string"])
    except Exception:
        pass

def append_nav_log(sample: dict):
    """Append one nav sample (thread-safe)."""
    try:
        _ensure_nav_log_header()
        with nav_log_lock:
            with open(NAV_LOG_FILE, "a", newline="", encoding="utf-8") as f:
                w = csv.writer(f)
                w.writerow([
                    sample.get("date",""),
                    sample.get("time",""),
                    sample.get("latitude",""),
                    sample.get("longitude",""),
                    sample.get("speed",""),
                    sample.get("cog",""),
                    sample.get("heading",""),
                    sample.get("voltage",""),
                    sample.get("panic",""),
                    sample.get("ext_heading",""),
                    sample.get("raw_string",""),
                ])
    except Exception:
        pass


# ------------------------------------------------------------------------------
# AUTH / ROLES
# ------------------------------------------------------------------------------

def current_user():
    return session.get("user")

def current_role():
    return session.get("role", "Viewer")

def require_role(min_role: str):
    def decorator(fn):
        @wraps(fn)
        def wrapper(*args, **kwargs):
            role = current_role()
            if ROLE_ORDER.get(role, 0) < ROLE_ORDER.get(min_role, 0):
                return jsonify({"error": "forbidden", "required": min_role, "role": role}), 403
            return fn(*args, **kwargs)
        return wrapper
    return decorator

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "GET":
        return render_template_string(LOGIN_HTML, error=None)
    username = (request.form.get("username") or "").strip()
    password = request.form.get("password") or ""
    u = USERS.get(username)
    if not u or u["password"] != password:
        return render_template_string(LOGIN_HTML, error="Invalid username or password.")
    session["user"] = username
    session["role"] = u["role"]
    return redirect(url_for("index"))

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

# ------------------------------------------------------------------------------
# LOGO / ASSETS
# ------------------------------------------------------------------------------

@app.route("/assets/logo")
def gds_logo():
    if os.path.exists(LOGO_FILE):
        mime = "image/jpeg" if LOGO_FILE.lower().endswith((".jpg", ".jpeg")) else "image/png"
        return send_file(LOGO_FILE, mimetype=mime)
    png = base64.b64decode(FALLBACK_LOGO_PNG_BASE64)
    return Response(png, mimetype="image/png")

@app.route("/assets/no-signal.png")
def no_signal():
    png = base64.b64decode(
        "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mP8/x8AAwMB/6Xw2m8AAAAASUVORK5CYII="
    )
    return Response(png, mimetype="image/png")

# ------------------------------------------------------------------------------
# CAMERA HELPERS (DEMO-SAFE)
# ------------------------------------------------------------------------------

def _demo_camera_frame(text="DEMO CAMERA (NO SIGNAL)"):
    """
    Returns JPEG bytes. This ensures <img src="/camera/stream"> always shows something.
    """
    img = Image.new("RGB", (960, 540), (10, 10, 12))
    draw = ImageDraw.Draw(img)

    # Border + header bar
    draw.rectangle((0, 0, 959, 539), outline=(204, 0, 0), width=6)
    draw.rectangle((0, 0, 959, 70), fill=(30, 0, 0))
    draw.text((18, 22), "GDS PTZ FEED", fill=(255, 255, 255))

    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    draw.text((18, 110), text, fill=(255, 200, 200))
    draw.text((18, 150), f"Timestamp: {ts}", fill=(220, 220, 220))

    with camera_lock:
        pan = camera_controls.get("pan", 0)
        tilt = camera_controls.get("tilt", 0)
        zoom = camera_controls.get("zoom", 1.0)
    draw.text((18, 200), f"PAN: {pan}   TILT: {tilt}   ZOOM: {zoom}", fill=(220, 220, 220))

    # Simple crosshair
    cx, cy = 480, 320
    draw.line((cx - 60, cy, cx + 60, cy), fill=(255, 31, 31), width=3)
    draw.line((cx, cy - 60, cx, cy + 60), fill=(255, 31, 31), width=3)

    out = BytesIO()
    img.save(out, format="JPEG", quality=85)
    return out.getvalue()

def check_camera_connection():
    global camera_status
    try:
        auth = (CAMERA_USERNAME, CAMERA_PASSWORD) if CAMERA_USERNAME else None

        # IMPORTANT: Android IP Webcam does NOT support HEAD
        response = requests.get(
            IP_CAMERA_URL,
            auth=auth,
            stream=True,
            timeout=CAMERA_CONNECTION_TIMEOUT
        )

        if response.status_code < 400:
            with camera_lock:
                camera_status["connected"] = True
                camera_status["mode"] = "ip"
                camera_status["error"] = None
                camera_status["last_check"] = datetime.now().isoformat()
            return True

        raise Exception(f"Status {response.status_code}")

    except Exception as e:
        if AUTO_FALLBACK_TO_DEMO:
            with camera_lock:
                camera_status["connected"] = False
                camera_status["mode"] = "demo"
                camera_status["error"] = str(e)
                camera_status["last_check"] = datetime.now().isoformat()
        return False


def init_camera():
    check_camera_connection()

def clamp(v, lo, hi):
    return max(lo, min(hi, v))

def map_heading_to_pan(heading_deg: float) -> int:
    h = float(heading_deg) % 360.0
    pan = int(round(h if h <= 180 else h - 360))
    return clamp(pan, -180, 180)

def ptz_sync_worker():
    global ptz_sync_enabled
    while True:
        try:
            with sync_lock:
                enabled = ptz_sync_enabled
            if enabled:
                with nav_lock:
                    cog = nav_current.get("cog", 0.0)
                desired_pan = map_heading_to_pan(cog)
                with camera_lock:
                    camera_controls["pan"] = desired_pan
            time.sleep(1)
        except Exception:
            time.sleep(1)

threading.Thread(target=ptz_sync_worker, daemon=True).start()

# ------------------------------------------------------------------------------
# SIMULATION THREADS (dummy ST6100-like)
# ------------------------------------------------------------------------------

def simulate_navigation():
    global sim_lat, sim_lon, sim_heading, sim_speed, nav_current
    while True:
        try:
            sim_speed += random.uniform(-0.3, 0.3)
            sim_speed = max(5, min(12, sim_speed))
            sim_heading = (sim_heading + random.uniform(-2, 2)) % 360

            distance = (sim_speed * 1.852 / 3600) / 111  # approx degrees per sec
            sim_lat += distance * math.cos(math.radians(sim_heading))
            sim_lon += distance * math.sin(math.radians(sim_heading))
            sim_lat = max(1.5, min(6.0, sim_lat))
            sim_lon = max(100.0, min(104.5, sim_lon))

            cog = (sim_heading + random.uniform(-2, 2)) % 360
            voltage = random.randint(1190, 1240)

            now = datetime.now()
            date_str = now.strftime("%d/%m/%Y")
            time_str = now.strftime("%H:%M:%S")

            raw_string = f"{date_str},{time_str},{sim_lat:.6f},{sim_lon:.6f},{sim_speed:.1f},{cog:.0f},{voltage},0,{sim_heading:.0f}"

            with nav_lock:
                if nav_current.get("date"):
                    nav_history.insert(0, nav_current.copy())
                    if len(nav_history) > MAX_NAV_HISTORY:
                        nav_history.pop()

                nav_current = {
                    "date": date_str,
                    "time": time_str,
                    "latitude": round(sim_lat, 6),
                    "longitude": round(sim_lon, 6),
                    "speed": round(sim_speed, 1),
                    "heading": round(sim_heading, 0),
                    "cog": round(cog, 0),
                    "voltage": voltage,
                    "panic": 0,
                    "ext_heading": round(sim_heading, 0),
                    "raw_string": raw_string
                }


            # Persist NAV log
            append_nav_log(nav_current)

            time.sleep(1)
        except Exception:
            time.sleep(1)

def simulate_weather():
    global weather_current
    base_temp, base_humid, base_press, base_wind, base_dir = 28.0, 75.0, 1013.25, 5.0, 180.0
    while True:
        try:
            base_wind = max(0, base_wind + random.uniform(-1, 1))
            base_dir = (base_dir + random.uniform(-5, 5)) % 360
            base_temp = max(20, min(40, base_temp + random.uniform(-0.2, 0.2)))
            base_humid = max(40, min(100, base_humid + random.uniform(-0.5, 0.5)))
            base_press = max(990, min(1030, base_press + random.uniform(-0.1, 0.1)))

            with weather_lock:
                weather_current = {
                    "wind_speed": round(base_wind, 1),
                    "wind_direction": round(base_dir, 0),
                    "temperature": round(base_temp, 1),
                    "humidity": round(base_humid, 1),
                    "pressure": round(base_press, 2),
                    "illumination": 50000,
                    "timestamp": datetime.now().strftime("%H:%M:%S")
                }
            time.sleep(2)
        except Exception:
            time.sleep(2)

#threading.Thread(target=simulate_navigation, daemon=True).start()
#threading.Thread(target=simulate_weather, daemon=True).start()
#threading.Thread(target=firebase_weather_listener, daemon=True).start()
#threading.Thread(target=firebase_marinelite_listener, daemon=True).start()



# ------------------------------------------------------------------------------
# ROUTES
# ------------------------------------------------------------------------------
@app.route("/api/connectivity/status")
def api_connectivity_status():
    """
    Stub API.
    Later this will read Wi-Fi + SIM + Ethernet status from RPi.
    """
    return jsonify({
        "wifi": {
            "available": True,
            "connected": False,
            "ssid": None,
            "signal": None
        },
        "sim": {
            "available": True,
            "connected": False,
            "operator": None,
            "country": None,
            "signal": None
        },
        "internet": False
    })
@app.route("/api/sync/status")
def api_sync_status():
    """
    Stub API.
    Later this will reflect SQLite sync state.
    """
    return jsonify({
        "pending_records": 0,
        "last_sync": None,
        "sync_state": "idle"
    })
@app.route("/api/sync/start", methods=["POST"])
def api_sync_start():
    """
    Stub API.
    Later this will trigger background sync.
    """
    return jsonify({
        "status": "accepted",
        "message": "Sync will start when connectivity is available"
    })

@app.route("/")
def index():
    if not current_user():
        return redirect(url_for("login"))
    with camera_lock:
        mode = camera_status["mode"]
    role = current_role()
    user = current_user()
    return render_template_string(HTML_DASHBOARD, camera_mode=mode, role=role, user=user, weasy=WEASYPRINT_AVAILABLE)

@app.route("/api/me")
def api_me():
    return jsonify({"user": current_user(), "role": current_role(), "weasyprint": WEASYPRINT_AVAILABLE})

@app.route("/camera/capture", methods=["POST"])
@require_role("Operator")
def capture_image():
    data = request.json or {}
    report_type = data.get("report_type", "vdr")

    # OFFLINE MODE: Firebase disabled here
    upload_firebase = False

    try:
        jpg_quality = int(data.get("jpg_quality", 75))
    except Exception:
        jpg_quality = 75

    try:
        max_width = int(data.get("max_width", 1280))
    except Exception:
        max_width = 1280

    jpg_quality = max(35, min(95, jpg_quality))
    max_width = max(320, min(1920, max_width))

    # ---- Capture image ----
    content = _fetch_camera_snapshot_bytes()
    if not content:
        content = _demo_camera_frame("CAPTURED (FALLBACK)")

    content = _compress_jpeg(content, max_width=max_width, quality=jpg_quality)
    base64_data = base64.b64encode(content).decode()

    # ---- SIZE GUARD (INSIDE FUNCTION) ----
    size_kb = len(base64_data) * 0.75 / 1024
    if size_kb > 80:
        return jsonify({
            "status": "error",
            "message": f"Image too large for DB-only mode ({size_kb:.1f} KB)"
        }), 400

    item = {
        "id": uuid.uuid4().hex,
        "data": base64_data,
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "type": "jpeg"
    }

    # ---- Store locally for UI ----
    with image_lock:
        key = "vjr_images" if report_type == "vjr" else "vdr_images"
        captured_images[key].insert(0, item)
        if len(captured_images[key]) > MAX_CAPTURE_IMAGES:
            captured_images[key].pop()

    return jsonify({
        "status": "success",
        "report_type": report_type,
        "size_kb": round(size_kb, 2)
    })


@app.route("/camera/control", methods=["POST"])
@require_role("Operator")
def control_camera():
    """
    Always returns OK so buttons feel functional even without hardware.
    """
    data = request.json or {}
    action = data.get("action", "")
    value = data.get("value", 0)

    with camera_lock:
        if action == "pan_left":
            camera_controls["pan"] = max(-180, camera_controls["pan"] - 15)
        elif action == "pan_right":
            camera_controls["pan"] = min(180, camera_controls["pan"] + 15)
        elif action == "tilt_up":
            camera_controls["tilt"] = min(90, camera_controls["tilt"] + 15)
        elif action == "tilt_down":
            camera_controls["tilt"] = max(-90, camera_controls["tilt"] - 15)
        elif action == "zoom_in":
            camera_controls["zoom"] = min(10, camera_controls["zoom"] + 0.5)
        elif action == "zoom_out":
            camera_controls["zoom"] = max(1, camera_controls["zoom"] - 0.5)
        elif action == "zoom_set":
            try:
                camera_controls["zoom"] = max(1, min(10, float(value)))
            except Exception:
                pass
        elif action == "led_toggle":
            camera_controls["led_enabled"] = not camera_controls["led_enabled"]
        elif action == "led_brightness":
            try:
                camera_controls["led_brightness"] = max(0, min(100, int(value)))
            except Exception:
                pass
        elif action == "night_vision_toggle":
            camera_controls["night_vision"] = not camera_controls["night_vision"]

        control_copy = camera_controls.copy()

    return jsonify({"status": "ok", "message": f"Action executed: {action}", "controls": control_copy})

@app.route("/camera/status")
def get_camera_status():
    with camera_lock:
        st = camera_status.copy()
        st["controls"] = camera_controls.copy()
    with sync_lock:
        st["ptz_sync_enabled"] = ptz_sync_enabled
    return jsonify(st)

@app.route("/camera/reconnect", methods=["POST"])
@require_role("Operator")
def reconnect_camera():
    result = check_camera_connection()
    with camera_lock:
        status_copy = camera_status.copy()
    return jsonify({"reconnected": result, "status": status_copy})

@app.route("/camera/sync_cog", methods=["POST"])
@require_role("Operator")
def camera_sync_cog():
    global ptz_sync_enabled
    enabled = bool((request.json or {}).get("enabled", False))
    with sync_lock:
        ptz_sync_enabled = enabled
    return jsonify({"status": "ok", "enabled": ptz_sync_enabled})


def _extract_first_jpeg_from_mjpeg(resp, max_bytes=3_000_000, max_seconds=1.5):
    """Extract the first JPEG frame from an MJPEG stream response.

    Hard-stops after max_seconds to avoid UI feeling 'stuck' when the IP camera
    stalls or buffers.
    """
    buf = bytearray()
    start = -1
    t0 = time.time()
    for chunk in resp.iter_content(chunk_size=4096):
        if (time.time() - t0) > max_seconds:
            break
        if not chunk:
            continue
        buf.extend(chunk)
        if start < 0:
            s = buf.find(b"\xff\xd8")
            if s >= 0:
                start = s
        if start >= 0:
            e = buf.find(b"\xff\xd9", start)
            if e >= 0:
                return bytes(buf[start:e+2])
        if len(buf) > max_bytes:
            break
    return None


def _compress_jpeg(jpg_bytes: bytes, max_width=1280, quality=75):
    """Downscale and recompress JPEG for faster UI + smaller base64."""
    try:
        im = Image.open(BytesIO(jpg_bytes))
        im = im.convert("RGB")
        w, h = im.size
        if w > max_width:
            nh = int(h * (max_width / float(w)))
            im = im.resize((max_width, nh))
        out = BytesIO()
        im.save(out, format="JPEG", quality=quality, optimize=True)
        return out.getvalue()
    except Exception:
        return jpg_bytes

def _fetch_camera_snapshot_bytes():
    """Fast snapshot: get a single JPEG frame (works for MJPEG and JPEG endpoints)."""
    with camera_lock:
        mode = camera_status.get("mode", "demo")

    if mode == "demo":
        return _demo_camera_frame("CAPTURED (DEMO MODE)")

    try:
        auth = (CAMERA_USERNAME, CAMERA_PASSWORD) if CAMERA_USERNAME else None

        # Use streaming so we can cut after the first JPEG instead of downloading the full MJPEG feed
        resp = requests.get(
            IP_CAMERA_URL,
            auth=auth,
            stream=True,
            timeout=(CAMERA_CONNECTION_TIMEOUT, 2),
            headers={"User-Agent": "GDS-VMS/1.0"},
        )
        if resp.status_code != 200:
            return None

        ctype = (resp.headers.get("Content-Type") or "").lower()
        if "multipart" in ctype or "mjpeg" in ctype or "x-mixed-replace" in ctype:
            frame = _extract_first_jpeg_from_mjpeg(resp)
            if frame:
                return _compress_jpeg(frame)
            return None

        # Non-multipart: assume direct JPEG snapshot endpoint
        data = resp.content
        if data:
            return _compress_jpeg(data)
    except Exception:
        return None

    return None







@app.route("/camera/captured_images")
def get_captured_images_count():
    with image_lock:
        return jsonify({"vjr_images": len(captured_images["vjr_images"]), "vdr_images": len(captured_images["vdr_images"])})

@app.route("/camera/captured_images_full")
def get_captured_images_full():
    """
    Used by UI galleries (VJR & VDR) to display thumbnails.
    """
    with image_lock:
        return jsonify({
            "vjr": captured_images["vjr_images"],
            "vdr": captured_images["vdr_images"]
        })

@app.route("/camera/clear_captures", methods=["POST"])
@require_role("Operator")
def clear_captures():
    data = request.json or {}
    report_type = (data.get("report_type") or "").lower()
    if report_type not in ("vjr", "vdr"):
        return jsonify({"error": "Invalid report_type"}), 400

    key = "vjr_images" if report_type == "vjr" else "vdr_images"
    with image_lock:
        captured_images[key].clear()

    return jsonify({"status": "success", "cleared": report_type})

@app.route("/camera/delete_capture", methods=["POST"])
@require_role("Operator")
def delete_capture():
    """Delete one captured image.

    Old behavior used an array index (fragile when UI re-renders).
    New behavior supports stable deletion by capture_id.
    """
    data = request.json or {}
    report_type = (data.get("report_type") or "").lower()

    capture_id = data.get("capture_id") or data.get("id")
    idx = data.get("index", None)

    if report_type not in ("vjr", "vdr"):
        return jsonify({"error": "Invalid report_type"}), 400

    key = "vjr_images" if report_type == "vjr" else "vdr_images"

    with image_lock:
        arr = captured_images.get(key, [])

        # Preferred: delete by capture_id (stable)
        if capture_id:
            for i, it in enumerate(arr):
                if str(it.get("id")) == str(capture_id):
                    arr.pop(i)
                    return jsonify({"status": "success", "deleted_id": capture_id, "report_type": report_type})
            return jsonify({"error": "capture_id not found"}), 404

        # Backward-compatible: delete by index
        if idx is None:
            return jsonify({"error": "capture_id or index required"}), 400
        try:
            idx = int(idx)
        except Exception:
            return jsonify({"error": "index must be int"}), 400

        if idx < 0 or idx >= len(arr):
            return jsonify({"error": "index out of range"}), 400
        arr.pop(idx)

    return jsonify({"status": "success", "deleted_index": idx, "report_type": report_type})



@app.route("/nav_data")
def get_navigation_data():
    if LOCAL_DB_ENABLED:
        return jsonify(db_get_latest_marinelite())

    # fallback (old behavior)
    with nav_lock:
        return jsonify({
            "latitude": nav_current.get("latitude"),
            "longitude": nav_current.get("longitude"),
            "heading": nav_current.get("heading"),
            "timestamp": nav_current.get("timestamp")
        })



@app.route("/weather_data")
def get_weather_data():
    if LOCAL_DB_ENABLED:
        return jsonify(db_get_latest_weather())

    with weather_lock:
        return jsonify(weather_current)




@app.route("/export_nav_csv")
def export_nav_csv():
    """Browser download of the persistent navigation log (CSV)."""
    if not current_user():
        return redirect(url_for("login"))
    _ensure_nav_log_header()
    try:
        return send_file(
            NAV_LOG_FILE,
            as_attachment=True,
            download_name=f"NAV_Log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            mimetype="text/csv",
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ------------------------------------------------------------------------------
# VDR RECORDS + EXPORTS
# ------------------------------------------------------------------------------

@app.route("/vdr_records")
def get_vdr_records():
    with vdr_lock:
        return jsonify(vdr_records)

@app.route("/save_vdr", methods=["POST"])
@require_role("Operator")
def save_vdr_record():
    data = request.json or {}
    with vdr_lock:
        record = {k: data.get(k, "") for k in [
            "date","vessel","imo","mmsi","callsign","client","location","country",
            "activity","weather","sea_state","crew_count","deck_crew","engine_crew",
            "officers","fuel_consumption","main_engine_hours","generator_hours",
            "distance_traveled","incidents","maintenance_work","remarks"
        ]}
        record["id"] = (vdr_records[0]["id"] + 1) if vdr_records else 1
        record["timestamp"] = datetime.now().isoformat()
        vdr_records.insert(0, record)
    return jsonify({"status": "ok", "total": len(vdr_records)})

@app.route("/clear_vdr", methods=["POST"])
@require_role("Captain")
def clear_vdr_records():
    with vdr_lock:
        vdr_records.clear()
    return jsonify({"status": "ok"})

@app.route("/export_vdr_excel", methods=["POST"])
@require_role("Operator")
def export_vdr_excel():
    if not openpyxl:
        return jsonify({"error": "openpyxl not installed"}), 400

    records = (request.json or {}).get("records", [])
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Daily Report"

    headers = [
        "ID","Date","Vessel","IMO","MMSI","Call Sign","Client","Location","Country",
        "Activity","Weather","Sea State","Crew","Deck Crew","Engine Crew","Officers",
        "Fuel","M/E Hours","Gen Hours","Distance","Incidents","Maintenance","Remarks"
    ]
    ws.append(headers)

    for cell in ws[1]:
        cell.fill = PatternFill(start_color="CC0000", end_color="CC0000", fill_type="solid")
        cell.font = XLFont(color="FFFFFF", bold=True)

    for r in records:
        ws.append([r.get(k, "") for k in [
            "id","date","vessel","imo","mmsi","callsign","client","location","country",
            "activity","weather","sea_state","crew_count","deck_crew","engine_crew","officers",
            "fuel_consumption","main_engine_hours","generator_hours","distance_traveled",
            "incidents","maintenance_work","remarks"
        ]])

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return Response(
        out.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=VDR_Report.xlsx"}
    )

@app.route("/export_vdr_csv", methods=["POST"])
@require_role("Operator")
def export_vdr_csv():
    records = (request.json or {}).get("records", [])
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "ID","Date","Vessel","IMO","MMSI","Call Sign","Client","Location","Country",
        "Activity","Weather","Sea State","Crew","Deck Crew","Engine Crew","Officers",
        "Fuel","M/E Hours","Gen Hours","Distance","Incidents","Maintenance","Remarks"
    ])
    for r in records:
        writer.writerow([r.get(k, "") for k in [
            "id","date","vessel","imo","mmsi","callsign","client","location","country",
            "activity","weather","sea_state","crew_count","deck_crew","engine_crew","officers",
            "fuel_consumption","main_engine_hours","generator_hours","distance_traveled",
            "incidents","maintenance_work","remarks"
        ]])
    return Response(output.getvalue(), mimetype="text/csv",
                    headers={"Content-Disposition": "attachment; filename=VDR_Report.csv"})

def _logo_b64():
    if os.path.exists(LOGO_FILE):
        return base64.b64encode(open(LOGO_FILE, "rb").read()).decode()
    return FALLBACK_LOGO_PNG_BASE64

@app.route("/export_vdr_pdf", methods=["POST"])
@require_role("Operator")
def export_vdr_pdf():
    records = (request.json or {}).get("records", [])
    with image_lock:
        images = captured_images["vdr_images"][:18]

    # TRY REPORTLAB FIRST (Windows-friendly, pure Python)
    if REPORTLAB_AVAILABLE:
        try:
            buffer = BytesIO()
            doc = SimpleDocTemplate(
                buffer,
                pagesize=letter,
                rightMargin=0.5*inch,
                leftMargin=0.5*inch,
                topMargin=0.75*inch,
                bottomMargin=0.75*inch
            )
            
            story = []
            styles = getSampleStyleSheet()
            
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                textColor=colors.HexColor('#cc0000'),
                spaceAfter=10
            )
            
            story.append(Paragraph(
                "<b><font color='#cc0000'>Vessel Daily Report (VDR)</font></b><br/>"
                f"<font size=9>Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</font>",
                title_style
            ))
            story.append(Spacer(1, 0.2*inch))
            
            # Records Table
            if records:
                table_data = [['Date', 'Vessel', 'Activity', 'Location', 'Weather', 'Remarks']]
                for r in records[:50]:
                    table_data.append([
                        str(r.get('date', '')),
                        str(r.get('vessel', '')),
                        str(r.get('activity', '')),
                        str(r.get('location', '')),
                        str(r.get('weather', '')),
                        str(r.get('remarks', ''))[:40]
                    ])
                
                table = Table(table_data, colWidths=[0.8*inch, 0.9*inch, 0.8*inch, 0.9*inch, 0.9*inch, 1.2*inch])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a1a1a')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 9),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('FONTSIZE', (0, 1), (-1, -1), 7),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
                ]))
                story.append(table)
            
            story.append(Spacer(1, 0.2*inch))
            
            # Images
            if images:
                img_data = []
                for img in images[:18]:
                    try:
                        img_b64 = img.get('data', '')
                        if img_b64:
                            img_bytes = base64.b64decode(img_b64)
                            pil_img = Image.open(BytesIO(img_bytes))
                            pil_img.thumbnail((1.2*inch, 1.2*inch))
                            img_path = f'/tmp/vdr_img_{int(time.time()*1000)}.png'
                            pil_img.save(img_path)
                            img_data.append(RLImage(img_path, width=1.1*inch, height=1.1*inch))
                    except Exception:
                        pass
                
                if img_data:
                    img_grid = []
                    for i in range(0, len(img_data), 2):
                        if i+1 < len(img_data):
                            img_grid.append([img_data[i], img_data[i+1]])
                        else:
                            img_grid.append([img_data[i], ''])
                    
                    img_table = Table(img_grid)
                    img_table.setStyle(TableStyle([
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('PADDING', (5, 5), (-1, -1), 5),
                    ]))
                    story.append(img_table)
            
            story.append(Spacer(1, 0.1*inch))
            story.append(Paragraph(f"<font size=7>GDS Maritime System  {datetime.now().year}</font>", styles['Normal']))
            
            doc.build(story)
            pdf_bytes = buffer.getvalue()
            buffer.close()
            
            return Response(
                pdf_bytes,
                mimetype="application/pdf",
                headers={"Content-Disposition": "attachment; filename=VDR_Report.pdf"}
            )
        
        except Exception as e:
            print(f"ReportLab PDF generation failed: {e}, falling back to WeasyPrint")
    
    # FALLBACK TO WEASYPRINT
    if not WEASYPRINT_AVAILABLE:
        return jsonify({"error": "PDF generation failed. Install: pip install reportlab pillow"}), 400
    
    rows = ""
    for r in records[:50]:
        rows += f"""
        <tr>
          <td>{r.get('id','')}</td>
          <td>{r.get('date','')}</td>
          <td>{r.get('vessel','')}</td>
          <td>{r.get('activity','')}</td>
          <td>{r.get('location','')}</td>
          <td>{r.get('weather','')}</td>
          <td>{r.get('remarks','')}</td>
        </tr>
        """

    gallery = ""
    if images:
        gallery += "<h2>Captured Images (VDR)</h2><div class='grid'>"
        for img in images:
            gallery += f"""
            <div class="card">
              <img src="data:image/jpeg;base64,{img['data']}" />
              <div class="cap">{img['timestamp']}</div>
            </div>
            """
        gallery += "</div>"

    html = f"""
    <!doctype html>
    <html>
    <head>
      <meta charset="utf-8">
      <style>
        body {{ font-family: Arial, sans-serif; padding: 24px; color: #111; }}
        .header {{ display:flex; align-items:center; gap:14px; border-bottom: 4px solid #cc0000; padding-bottom: 14px; margin-bottom: 14px; }}
        .header img {{ height: 46px; }}
        h1 {{ margin:0; color:#cc0000; font-size: 22px; }}
        .sub {{ color:#555; font-size: 12px; margin-top: 4px; }}
        table {{ width:100%; border-collapse:collapse; font-size: 11px; margin-top: 12px; }}
        th, td {{ border:1px solid #ddd; padding: 8px; vertical-align: top; }}
        th {{ background:#111; color:#fff; }}
        tr:nth-child(even) {{ background:#fafafa; }}
        h2 {{ color:#cc0000; margin: 18px 0 10px; font-size: 14px; border-bottom: 2px solid #cc0000; padding-bottom: 6px; }}
        .grid {{ display:grid; grid-template-columns: repeat(2, 1fr); gap: 12px; }}
        .card {{ border:1px solid #ddd; border-radius: 8px; overflow:hidden; }}
        .card img {{ width:100%; height:auto; display:block; }}
        .cap {{ padding: 8px; font-size: 10px; color:#555; }}
        .foot {{ margin-top: 16px; border-top: 1px solid #ddd; padding-top: 10px; font-size: 10px; color:#666; text-align:center; }}
      </style>
    </head>
    <body>
      <div class="header">
        <img src="data:image/png;base64,{_logo_b64()}" />
        <div>
          <h1>Vessel Daily Report (VDR)</h1>
          <div class="sub">Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</div>
        </div>
      </div>

      <h2>Records Summary</h2>
      <table>
        <thead>
          <tr>
            <th>ID</th><th>Date</th><th>Vessel</th><th>Activity</th><th>Location</th><th>Weather</th><th>Remarks</th>
          </tr>
        </thead>
        <tbody>{rows if rows else '<tr><td colspan="7">No records</td></tr>'}</tbody>
      </table>

      {gallery}

      <div class="foot">Generated by GDS - Maritime Management System | &copy; {datetime.now().year}</div>
    </body>
    </html>
    """

    pdf = HTML(string=html).write_pdf()
    return Response(pdf, mimetype="application/pdf",
                    headers={"Content-Disposition": "attachment; filename=VDR_Report.pdf"})

# ------------------------------------------------------------------------------
# VJR PDF EXPORT (NEW)
# ------------------------------------------------------------------------------

@app.route("/export_vjr_pdf", methods=["POST"])
@require_role("Operator")
def export_vjr_pdf():
    vjr = (request.json or {}).get("vjr", {})
    navlog = (request.json or {}).get("nav", [])
    with image_lock:
        images = captured_images["vjr_images"][:18]

    # TRY REPORTLAB FIRST (Windows-friendly, pure Python)
    if REPORTLAB_AVAILABLE:
        try:
            buffer = BytesIO()
            doc = SimpleDocTemplate(
                buffer,
                pagesize=letter,
                rightMargin=0.5*inch,
                leftMargin=0.5*inch,
                topMargin=0.75*inch,
                bottomMargin=0.75*inch
            )
            
            story = []
            styles = getSampleStyleSheet()
            
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                textColor=colors.HexColor('#cc0000'),
                spaceAfter=10
            )
            
            story.append(Paragraph(
                "<b><font color='#cc0000'>Vessel Journey Report (VJR)</font></b><br/>"
                f"<font size=9>Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</font>",
                title_style
            ))
            story.append(Spacer(1, 0.2*inch))
            
            # Journey Details
            details_data = [
                ['Report Date:', vjr.get('date', '')],
                ['Vessel:', vjr.get('vessel', '')],
                ['IMO:', vjr.get('imo', '')],
                ['Master:', vjr.get('master', '')],
                ['Departure:', vjr.get('departure', '')],
                ['Arrival:', vjr.get('arrival', '')],
            ]
            
            details_table = Table(details_data, colWidths=[1.2*inch, 4*inch])
            details_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f0f0f0')),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('PADDING', (8, 6), (-1, -1), 6),
            ]))
            story.append(details_table)
            story.append(Spacer(1, 0.2*inch))
            
            # Images
            if images:
                img_data = []
                for img in images[:16]:
                    try:
                        img_b64 = img.get('data', '')
                        if img_b64:
                            img_bytes = base64.b64decode(img_b64)
                            pil_img = Image.open(BytesIO(img_bytes))
                            pil_img.thumbnail((1.2*inch, 1.2*inch))
                            img_path = f'/tmp/vjr_img_{int(time.time()*1000)}.png'
                            pil_img.save(img_path)
                            img_data.append(RLImage(img_path, width=1.1*inch, height=1.1*inch))
                    except Exception:
                        pass
                
                if img_data:
                    img_grid = []
                    for i in range(0, len(img_data), 2):
                        if i+1 < len(img_data):
                            img_grid.append([img_data[i], img_data[i+1]])
                        else:
                            img_grid.append([img_data[i], ''])
                    
                    img_table = Table(img_grid)
                    img_table.setStyle(TableStyle([
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('PADDING', (5, 5), (-1, -1), 5),
                    ]))
                    story.append(img_table)
            
            story.append(Spacer(1, 0.1*inch))
            story.append(Paragraph(f"<font size=7>GDS Maritime System {datetime.now().year}</font>", styles['Normal']))
            
            doc.build(story)
            pdf_bytes = buffer.getvalue()
            buffer.close()
            
            return Response(
                pdf_bytes,
                mimetype="application/pdf",
                headers={"Content-Disposition": "attachment; filename=VJR_Report.pdf"}
            )
        
        except Exception as e:
            print(f"ReportLab PDF generation failed: {e}, falling back to WeasyPrint")
    
    # FALLBACK TO WEASYPRINT
    if not WEASYPRINT_AVAILABLE:
        return jsonify({"error": "PDF generation failed. Install: pip install reportlab pillow"}), 400

    nav_rows = ""
    for i, n in enumerate(navlog[:40], start=1):
        nav_rows += f"""
        <tr>
          <td>{i}</td>
          <td>{n.get('date','')}</td>
          <td>{n.get('time','')}</td>
          <td>{n.get('latitude','')}</td>
          <td>{n.get('longitude','')}</td>
          <td>{n.get('speed','')}</td>
          <td>{n.get('cog','')}</td>
        </tr>
        """

    gallery = ""
    if images:
        gallery += "<h2>Captured Images (VJR)</h2><div class='grid'>"
        for img in images:
            gallery += f"""
            <div class="card">
              <img src="data:image/jpeg;base64,{img['data']}" />
              <div class="cap">{img['timestamp']}</div>
            </div>
            """
        gallery += "</div>"

    html = f"""
    <!doctype html>
    <html>
    <head>
      <meta charset="utf-8">
      <style>
        body {{ font-family: Arial, sans-serif; padding: 24px; color: #111; }}
        .header {{ display:flex; align-items:center; gap:14px; border-bottom: 4px solid #cc0000; padding-bottom: 14px; margin-bottom: 14px; }}
        .header img {{ height: 46px; }}
        h1 {{ margin:0; color:#cc0000; font-size: 22px; }}
        .sub {{ color:#555; font-size: 12px; margin-top: 4px; }}
        .meta {{ display:grid; grid-template-columns: 1fr 1fr; gap: 10px; margin: 12px 0 6px; }}
        .box {{ border:1px solid #ddd; border-radius: 8px; padding: 10px; font-size: 12px; }}
        .k {{ color:#666; font-size:11px; }}
        .v {{ font-weight:700; }}
        table {{ width:100%; border-collapse:collapse; font-size: 11px; margin-top: 12px; }}
        th, td {{ border:1px solid #ddd; padding: 8px; vertical-align: top; }}
        th {{ background:#111; color:#fff; }}
        tr:nth-child(even) {{ background:#fafafa; }}
        h2 {{ color:#cc0000; margin: 18px 0 10px; font-size: 14px; border-bottom: 2px solid #cc0000; padding-bottom: 6px; }}
        .grid {{ display:grid; grid-template-columns: repeat(2, 1fr); gap: 12px; }}
        .card {{ border:1px solid #ddd; border-radius: 8px; overflow:hidden; }}
        .card img {{ width:100%; height:auto; display:block; }}
        .cap {{ padding: 8px; font-size: 10px; color:#555; }}
        .foot {{ margin-top: 16px; border-top: 1px solid #ddd; padding-top: 10px; font-size: 10px; color:#666; text-align:center; }}
      </style>
    </head>
    <body>
      <div class="header">
        <img src="data:image/png;base64,{_logo_b64()}" />
        <div>
          <h1>Vessel Journey Report (VJR)</h1>
          <div class="sub">Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</div>
        </div>
      </div>

      <div class="meta">
        <div class="box"><div class="k">Report Date</div><div class="v">{vjr.get('date','')}</div></div>
        <div class="box"><div class="k">Vessel</div><div class="v">{vjr.get('vessel','')}</div></div>
        <div class="box"><div class="k">IMO</div><div class="v">{vjr.get('imo','')}</div></div>
        <div class="box"><div class="k">Master</div><div class="v">{vjr.get('master','')}</div></div>
        <div class="box"><div class="k">Departure</div><div class="v">{vjr.get('departure','')}</div></div>
        <div class="box"><div class="k">Arrival</div><div class="v">{vjr.get('arrival','')}</div></div>
      </div>

      <h2>Navigation Log</h2>
      <table>
        <thead>
          <tr>
            <th>#</th><th>Date</th><th>Time</th><th>Lat</th><th>Lon</th><th>Speed</th><th>COG</th>
          </tr>
        </thead>
        <tbody>{nav_rows if nav_rows else '<tr><td colspan="7">No navigation samples</td></tr>'}</tbody>
      </table>

      {gallery}

      <div class="foot">Generated by GDS - Maritime Management System | &copy; {datetime.now().year}</div>
    </body>
    </html>
    """

    pdf = HTML(string=html).write_pdf()
    return Response(pdf, mimetype="application/pdf",
                    headers={"Content-Disposition": "attachment; filename=VJR_Report.pdf"})

# ------------------------------------------------------------------------------
# UI
# ------------------------------------------------------------------------------

LOGIN_HTML = r"""
<!doctype html>
<html>
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>GDS Login</title>
  <style>
    :root{
      --bg:#070709; --panel:#0f0f13; --border:#262631;
      --red:#cc0000; --red2:#ff1f1f; --text:#ffffff; --muted:#a7a7b3;
    }
    body{
      background:
        radial-gradient(1200px 600px at 10% 10%, rgba(204,0,0,.14), transparent 60%),
        radial-gradient(900px 500px at 90% 20%, rgba(255,31,31,.10), transparent 55%),
        var(--bg);
      color:var(--text);font-family:Segoe UI,Arial,sans-serif;
      display:flex;min-height:100vh;align-items:center;justify-content:center;margin:0;
    }
    .card{
      width:380px;background:linear-gradient(180deg, rgba(255,255,255,.04), transparent 60%), var(--panel);
      border:1px solid var(--border);border-radius:14px;box-shadow:0 18px 50px rgba(0,0,0,.55);padding:22px;
    }
    .head{display:flex;align-items:center;gap:12px;margin-bottom:14px;}
    .head img{height:40px;filter: drop-shadow(0 10px 24px rgba(0,0,0,.5));}
    h1{font-size:15px;margin:0;letter-spacing:.3px;}
    .sub{font-size:12px;color:var(--muted);margin-top:3px;}
    label{font-size:10px;color:var(--muted);text-transform:uppercase;letter-spacing:.8px;font-weight:800;display:block;margin-top:12px;margin-bottom:6px;}
    input{width:100%;padding:11px 10px;background:rgba(0,0,0,.45);border:1px solid var(--border);border-radius:10px;color:var(--text);}
    input:focus{outline:none;border-color:rgba(255,31,31,.65);box-shadow:0 0 0 3px rgba(255,31,31,.12);}
    button{width:100%;margin-top:14px;padding:11px;border:0;border-radius:10px;background:linear-gradient(180deg, var(--red2), var(--red));
           color:#fff;font-weight:900;text-transform:uppercase;letter-spacing:.8px;cursor:pointer;}
    .err{margin-top:10px;background:rgba(255,31,31,.10);border:1px solid rgba(255,31,31,.6);color:#ff8a8a;padding:10px;border-radius:10px;font-size:12px;}
    .hint{margin-top:10px;font-size:11px;color:var(--muted);line-height:1.5;}
    code{background:rgba(255,255,255,.06);padding:2px 6px;border-radius:8px;border:1px solid rgba(255,255,255,.08);}
  </style>
</head>
<body>
  <div class="card">
    <div class="head">
      <img src="/assets/logo" alt="GDS">
      <div>
        <h1>GDS Vessel Management System</h1>
        <div class="sub">Black/Red Operational Console</div>
      </div>
    </div>
    <form method="POST">
      <label>Username</label>
      <input name="username" placeholder="captain / operator / viewer" autocomplete="username" required>
      <label>Password</label>
      <input type="password" name="password" autocomplete="current-password" required>
      <button type="submit">Login</button>
      {% if error %}<div class="err">{{ error }}</div>{% endif %}
      <div class="hint">
        Demo users:<br>
        <code>captain / captain123</code><br>
        <code>operator / operator123</code><br>
        <code>viewer / viewer123</code>
      </div>
    </form>
  </div>
</body>
</html>
"""

# NOTE: This HTML is your previous professional UI with added:
# - Gallery panels in VJR & VDR
# - VJR PDF button
# - Live gallery loader in JS

HTML_DASHBOARD = r'''<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>GDS - Vessel Management</title>
  <link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css" />
  <script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
  <style>
  /* ===== NAVIGATION FEED ===== */

.nav-grid{
  display:grid;
  grid-template-columns:1fr 1fr;
  gap:14px;
}

.nav-card{
  background:rgba(255,255,255,0.03);
  border:1px solid rgba(255,255,255,0.08);
  border-radius:10px;
  padding:12px;
  text-align:center;
}

.nav-card.wide{
  grid-column:span 2;
}

.nav-label{
  font-size:12px;
  color:#aaa;
  letter-spacing:0.5px;
}

.nav-value{
  font-size:22px;
  font-weight:600;
  color:#00ff99;
  margin-top:6px;
}

.nav-value.big{
  font-size:38px;
  font-weight:800;
  letter-spacing:.6px;

}

.nav-unit{
  font-size:11px;
  color:#888;
}

.nav-status{
  grid-column:span 2;
  display:flex;
  justify-content:space-between;
  align-items:center;
  padding:10px 14px;
  border-radius:10px;
  background:rgba(255,0,0,0.08);
  border:1px solid rgba(255,0,0,0.3);
}
.sys-health{
  display:flex;
  gap:8px;
}

.health-pill{
  padding:4px 10px;
  border-radius:20px;
  font-size:11px;
  font-weight:700;
  letter-spacing:.4px;
}

.health-pill.ok{
  background:rgba(0,208,132,.15);
  color:#00d084;
  border:1px solid rgba(0,208,132,.4);
}

.health-pill.warn{
  background:rgba(255,204,0,.15);
  color:#ffcc00;
  border:1px solid rgba(255,204,0,.4);
}

.health-pill.bad{
  background:rgba(255,68,102,.15);
  color:#ff4466;
  border:1px solid rgba(255,68,102,.4);
}

.status-bad{
  color:#ff4d4d;
  font-weight:700;
}

.status-good{
  color:#00ff99;
  font-weight:700;
}

.nav-time{
  grid-column:span 2;
  text-align:right;
  font-size:12px;
  color:#aaa;
}

    :root{
      --bg:#070709;
      --panel:#0f0f13;
      --panel2:#13131a;
      --border:#262631;
      --text:#ffffff;
      --muted:#a7a7b3;
      --red:#cc0000;
      --red2:#ff1f1f;
      --good:#00d084;
      --warn:#ffcc00;
      --bad:#ff4466;
      --shadow: 0 18px 55px rgba(0,0,0,.55);
    }
    *{ margin:0; padding:0; box-sizing:border-box; }
    html, body{ height:100%; background: var(--bg); color:var(--text); font-family: Segoe UI, Arial, sans-serif; }
    body{
      display:flex; flex-direction:column;
      background:
        radial-gradient(1200px 700px at 10% 10%, rgba(204,0,0,.14), transparent 60%),
        radial-gradient(900px 600px at 90% 20%, rgba(255,31,31,.10), transparent 55%),
        var(--bg);
    }
    .header{
      display:flex; align-items:center; justify-content:space-between; gap:12px;
      padding: 14px 16px;
      border-bottom: 1px solid rgba(255,255,255,.06);
      background: linear-gradient(180deg, rgba(255,255,255,.04), transparent 70%);
      backdrop-filter: blur(8px);
      flex-shrink:0;
    }
    .brand{ display:flex; align-items:center; gap:12px; min-width: 280px; }
    .brand img{
      height: 44px; width:auto; display:block;
      filter: drop-shadow(0 10px 26px rgba(0,0,0,.55));
    }
    .brand h1{ font-size: 15px; letter-spacing:.3px; margin:0; }
    .brand p{ font-size: 11px; color: var(--muted); margin-top:2px; }
    .hdr-right{ display:flex; align-items:center; gap:10px; flex-wrap:wrap; justify-content:flex-end; }
    .pill{
      display:flex; align-items:center; gap:8px;
      padding: 8px 10px;
      border: 1px solid rgba(255,255,255,.08);
      border-radius: 999px;
      background: rgba(0,0,0,.35);
      font-size: 11px;
      color: var(--text);
    }
    .pill b{ color: var(--red2); }
    .clock{
      font-family: ui-monospace, SFMono-Regular, Menlo, Monaco, Consolas, "Courier New", monospace;
      font-size: 12px;
      padding: 8px 10px;
      border-radius: 10px;
      border: 1px solid rgba(255,31,31,.30);
      background: rgba(255,31,31,.08);
    }
    .pill{
  padding:4px 10px;
  border-radius:999px;
  font-size:12px;
  font-weight:600;
  border:1px solid rgba(255,255,255,.35);
  color:#fff;
  background:rgba(0,0,0,.25);
  white-space:nowrap;
}

    .link{ color: #ffd1d1; text-decoration: none; font-weight: 800; font-size: 12px; }
    .link:hover{ text-decoration: underline; }

    .tabs{
      display:flex; gap:8px; padding: 10px 14px;
      border-bottom: 1px solid rgba(255,255,255,.06);
      background: rgba(0,0,0,.18);
      flex-shrink:0;
      overflow-x:auto;
    }
    .tab-btn{
      border: 1px solid rgba(255,255,255,.08);
      background: rgba(0,0,0,.25);
      color: var(--muted);
      padding: 10px 14px;
      border-radius: 999px;
      cursor:pointer;
      font-weight: 900;
      font-size: 11px;
      text-transform: uppercase;
      letter-spacing: .8px;
      white-space: nowrap;
      transition: all .18s ease;
    }
    .tab-btn:hover{ border-color: rgba(255,31,31,.30); color: #fff; }
    .tab-btn.active{
      background: linear-gradient(180deg, rgba(255,31,31,.25), rgba(204,0,0,.18));
      border-color: rgba(255,31,31,.55);
      color:#fff;
    }

    .content-wrapper{ flex:1; overflow:hidden; display:flex; }
    .tab-content{ display:none; width:100%; height:100%; }
    .tab-content.active{ display:block; }

    .grid2{
      display:grid;
      grid-template-columns: 1fr 1fr;
      height:100%;
      gap: 12px;
      padding: 12px;
      overflow:hidden;
    }
    .grid3{
      display:grid;
      grid-template-columns: 1fr 1fr 1fr;
      height:100%;
      gap: 12px;
      padding: 12px;
      overflow:hidden;
    }

    .panel{
      background: linear-gradient(180deg, rgba(255,255,255,.04), transparent 60%), var(--panel);
      border: 1px solid rgba(255,255,255,.08);
      border-radius: 14px;
      box-shadow: var(--shadow);
      overflow:hidden;
      display:flex;
      flex-direction:column;
      min-height: 0;
    }
/* ===============================
   CONNECTIVITY STATUS COLORS
   =============================== */

.status-ok {
  color: #2ecc71;   /* green */
  font-weight: 600;
}

.status-warn {
  color: #f39c12;   /* orange */
  font-weight: 600;
}

.status-bad {
  color: #e74c3c;   /* red */
  font-weight: 600;
}

.status-muted {
  color: #7f8c8d;   /* gray */
}

    .panel-header{
      padding: 12px 14px;
      display:flex;
      justify-content:space-between;
      align-items:center;
      gap:10px;
      border-bottom: 1px solid rgba(255,255,255,.06);
      background: rgba(0,0,0,.22);
      font-weight: 950;
      font-size: 11px;
      text-transform: uppercase;
      letter-spacing: .9px;
    }
.panel-head .sub{
  font-size:11px;
  color:#aaa;
  margin-left:8px;
}

    .panel-header .hint{ font-size: 11px; color: var(--muted); font-weight: 800; text-transform:none; letter-spacing:0; }
    .panel-content{ padding: 14px; overflow:auto; min-height:0; }

    .raw{
      background: rgba(0,0,0,.45);
      border: 1px solid rgba(255,255,255,.08);
      border-radius: 12px;
      padding: 12px;
      color: #b9ffdb;
      font-family: ui-monospace, SFMono-Regular, Menlo, Monaco, Consolas, "Courier New", monospace;
      font-size: 11px;
      word-break: break-all;
      margin-bottom: 12px;
    }

    .data-grid{
      display:grid;
      grid-template-columns: 1fr 1fr;
      gap: 10px;
      margin-bottom: 10px;
    }
    .card{
      background: rgba(0,0,0,.30);
      border: 1px solid rgba(255,255,255,.08);
      border-radius: 12px;
      padding: 12px;
    }
    .label{
      color: var(--muted);
      font-size: 10px;
      text-transform: uppercase;
      letter-spacing: .9px;
      font-weight: 900;
      margin-bottom: 6px;
    }
    .value{
      font-family: ui-monospace, SFMono-Regular, Menlo, Monaco, Consolas, "Courier New", monospace;
      font-weight: 950;
      color: var(--good);
      font-size: 16px;
      letter-spacing: .2px;
    }
    .value.red{ color: #ff8a8a; }

    .form-group{ margin-bottom: 12px; }
    .form-label{
      display:block;
      margin-bottom: 6px;
      font-size: 10px;
      font-weight: 900;
      text-transform: uppercase;
      letter-spacing: .9px;
      color: var(--muted);
    }
    input, select, textarea{
      width:100%;
      padding: 10px 10px;
      border-radius: 12px;
      border: 1px solid rgba(255,255,255,.10);
      background: rgba(0,0,0,.45);
      color: #fff;
      font-size: 12px;
      outline:none;
      transition: box-shadow .15s ease, border-color .15s ease;
    }
    input:focus, select:focus, textarea:focus{
      border-color: rgba(255,31,31,.55);
      box-shadow: 0 0 0 3px rgba(255,31,31,.14);
    }
    textarea{ min-height: 64px; resize: vertical; }

    .btn{
      width:100%;
      padding: 11px 12px;
      border: 0;
      border-radius: 12px;
      cursor:pointer;
      font-weight: 950;
      font-size: 11px;
      letter-spacing: .9px;
      text-transform: uppercase;
      background: linear-gradient(180deg, var(--red2), var(--red));
      color: #fff;
      transition: transform .12s ease, box-shadow .12s ease, filter .12s ease;
      box-shadow: 0 12px 28px rgba(204,0,0,.22);
    }
    .btn:hover{ transform: translateY(-1px); filter: brightness(1.02); }
    .btn:active{ transform: translateY(0); }
    .btn.secondary{
      background: rgba(255,255,255,.06);
      border: 1px solid rgba(255,255,255,.10);
      box-shadow:none;
    }
    .btn.secondary:hover{ border-color: rgba(255,31,31,.35); }
    .btn.danger{
      background: linear-gradient(180deg, #ff4466, #c2183a);
      box-shadow: 0 12px 28px rgba(255,68,102,.18);
    }
    .btn.disabled{ opacity:.45; pointer-events:none; }

    .row2{ display:grid; grid-template-columns: 1fr 1fr; gap: 10px; }
    .mapbox{
      width:100%;
      height: 360px;
      border-radius: 14px;
      border: 1px solid rgba(255,255,255,.10);
      overflow:hidden;
      background:#000;
    }

    .cam-wrap{
      position:relative;
      width:100%;
      height: 260px;
      border-radius: 14px;
      border: 1px solid rgba(255,255,255,.10);
      overflow:hidden;
      background:#000;
    }
    .cam-wrap img{
      width:100%;
      height:100%;
      object-fit: cover;
      display:block;
    }
    .badge{
      position:absolute;
      left:12px;
      top:12px;
      padding: 7px 10px;
      border-radius: 999px;
      background: rgba(0,0,0,.55);
      border: 1px solid rgba(255,255,255,.12);
      font-size: 11px;
      font-weight: 900;
      color: #fff;
    }
    .badge b{ color: var(--red2); }

    .joy{
      width: 160px; height:160px; border-radius:50%;
      margin: 0 auto;
      background: rgba(0,0,0,.45);
      border: 1px solid rgba(255,255,255,.12);
      position:relative;
      touch-action:none;
      box-shadow: inset 0 0 0 1px rgba(255,31,31,.14);
    }
    .stick{
      width: 48px; height:48px; border-radius: 50%;
      position:absolute; left:50%; top:50%;
      transform: translate(-50%,-50%);
      background: radial-gradient(circle at 30% 25%, rgba(255,255,255,.22), rgba(255,255,255,.06)),
                  linear-gradient(180deg, rgba(255,31,31,.95), rgba(204,0,0,.85));
      box-shadow: 0 16px 28px rgba(0,0,0,.55);
    }
    /* ===== PTZ PANEL WRAPPER (NEW) ===== */

.ptz-panel{
  margin-top: 14px;
  padding-top: 12px;
  border-top: 1px solid rgba(255,255,255,.08);
  text-align: center;
}

.ptz-title{
  font-size: 11px;
  font-weight: 900;
  letter-spacing: .9px;
  text-transform: uppercase;
  color: var(--muted);
  margin-bottom: 10px;
}

.ptz-zoom{
  margin-top: 10px;
  display: grid;
  grid-template-columns: 1fr 1fr;
  gap: 10px;
}

    table{ width:100%; border-collapse:collapse; font-size: 11px; }
    th{
      text-align:left;
      padding: 10px;
      background: rgba(0,0,0,.35);
      border-bottom: 2px solid rgba(255,31,31,.35);
      color: #ffd1d1;
      font-weight: 950;
    }
    td{
      padding: 9px 10px;
      border-bottom: 1px solid rgba(255,255,255,.06);
      color: #e9e9ee;
    }
    tr:hover td{ background: rgba(255,31,31,.06); }

    .note{
      padding: 12px;
      background: rgba(255,31,31,.07);
      border: 1px solid rgba(255,31,31,.25);
      border-radius: 14px;
      color: #ffd1d1;
      font-size: 12px;
      line-height: 1.55;
    }

    .notification{
      position: fixed; right: 16px; bottom: 16px;
      padding: 12px 14px;
      border-radius: 14px;
      background: rgba(0,0,0,.65);
      border: 1px solid rgba(255,255,255,.10);
      box-shadow: var(--shadow);
      z-index: 9999;
      font-weight: 900;
      letter-spacing: .2px;
      max-width: 360px;
    }
    .notification.ok{ border-color: rgba(0,208,132,.35); color: #b9ffdb; }
    .notification.err{ border-color: rgba(255,68,102,.45); color: #ffd1d1; }
    .notification.warn{ border-color: rgba(255,204,0,.45); color: #fff2b3; }

    /* NEW: image galleries */
    .image-grid{
      display:grid;
      grid-template-columns: repeat(auto-fill, minmax(140px, 1fr));
      gap: 10px;
    }
    .thumb{
      position:relative;
      border: 1px solid rgba(255,255,255,.10);
      border-radius: 12px;
      overflow:hidden;
      background: rgba(0,0,0,.35);
    }
    .thumb img{ width:100%; display:block; }

    .xbtn{
      position:absolute;
      top:6px; right:6px;
      width:28px; height:28px;
      border-radius:10px;
      border:1px solid rgba(255,255,255,.14);
      background: rgba(0,0,0,.55);
      color:#fff;
      font-weight:900;
      cursor:pointer;
      line-height:26px;
      text-align:center;
      padding:0;
    }
    .xbtn:hover{ border-color: rgba(255,31,31,.55); }
    .thumb .cap{
      padding: 8px;
      font-size: 10px;
      color: var(--muted);
      border-top: 1px solid rgba(255,255,255,.06);
    }

    @media (max-width: 1100px){
      .grid3{ grid-template-columns: 1fr 1fr; }
      .grid2{ grid-template-columns: 1fr; }
      .brand{ min-width: auto; }
      .mapbox{ height: 300px; }
      .cam-wrap{ height: 240px; }
    }
    .dot{
  width:10px;
  height:10px;
  border-radius:50%;
  display:inline-block;
  margin-right:6px;
}

.dot.red{ background:#ff4466; }
.dot.green{ background:#00d084; }
.dot.yellow{ background:#ffcc00; }

  </style>
</head>
<body>
  <div class="header">
    <div class="brand">
      <img src="/assets/logo" alt="GDS">
      <div>
        <h1>GDS Vessel Management System</h1>
        <p>Black/Red Operational Console</p>
      </div>
    </div>
    <div class="hdr-right">
    <div class="pill" id="netStatus">MODE: --</div>
      <div class="pill">User: <b>{{ user }}</b></div>
      <div class="pill">Role: <b>{{ role }}</b></div>
      <div class="clock" id="clock">--:--:--</div>
      <a class="link" href="/logout">Logout</a>
    </div>
  </div>
<div class="sys-health">
  <span class="health-pill ok" id="gnss_health">GNSS: NO FIX</span>
  <span class="health-pill warn" id="net_health">NET: OFFLINE</span>
  <span class="health-pill ok" id="db_health">DB: LOCAL</span>
</div>

  <div class="tabs">
    
    <button class="tab-btn active" onclick="openTab('navigation')">Navigation</button>

    <button class="tab-btn" onclick="openTab('connectivity')">Connectivity & Sync</button>

    <button class="tab-btn" onclick="openTab('weather')">Weather</button>
    <button class="tab-btn" onclick="openTab('camera')">Camera/PTZ</button>
    <button class="tab-btn" onclick="openTab('vjr')">VJR</button>
    <button class="tab-btn" onclick="openTab('vdr')">VDR</button>
    <button class="tab-btn" onclick="openTab('logs')">Logs</button>
    
  </div>
  <div class="tab-content active" id="tab-navigation">

  <div class="grid2">

    <!-- LEFT: NAVIGATION FEED -->
    <div class="panel">
  <div class="panel-header">
    NAVIGATION FEED
    <span class="hint">MarineLite GNSS (Realtime)</span>
  </div>

  <div class="panel-content nav-grid">

    <!-- POSITION -->
    <div class="nav-card">
      <div class="nav-label">Latitude</div>
      <div class="nav-value" id="nav_lat">--</div>
      <div class="nav-unit"> N</div>
    </div>

    <div class="nav-card">
      <div class="nav-label">Longitude</div>
      <div class="nav-value" id="nav_lon">--</div>
      <div class="nav-unit"> E</div>
    </div>

    <!-- HEADING -->
    <div class="nav-card wide">
      <div class="nav-label">Heading</div>
      <div class="nav-value big" id="nav_heading">--</div>
      <div class="nav-unit">deg</div>
    </div>

    <!-- STATUS -->
    <div class="nav-status">
  <span>GNSS STATUS</span>
  <span class="status-bad" id="gnss_fix">NO FIX</span>
</div>


    <!-- TIME -->
    <div class="nav-time">
      Last Update:
      <span id="nav_time">--</span>
    </div>

  </div>
</div>


    <!-- RIGHT: LIVE MAP -->
    <div class="panel">
      <div class="panel-header">
        LIVE MAP
        <span class="hint">Realtime Vessel Position</span>
      </div>

      <div class="panel-content">
        <div class="mapbox" id="mapNav"></div>

        <div class="note">
          Vessel position is updated from MarineLite GNSS via local database (offline-capable).
        </div>
      </div>
    </div>

  </div>
</div>

</div>

<div class="tab-content" id="tab-connectivity">

  <div class="grid3">

    <!-- ================= NETWORK STATUS ================= -->
    <div class="panel">
      <div class="panel-header">
        Network Status
        <span class="hint">Current connectivity</span>
      </div>
      <div class="panel-content">

        <table>
          <tr><th>Interface</th><th>Status</th><th>Details</th></tr>

          <tr>
  <td>Wi-Fi</td>
  <td>
    <span class="dot red" id="wifi_dot"></span>
    <span id="wifi_status">--</span>
  </td>
  <td id="wifi_info">--</td>
</tr>


          <tr>
  <td>SIM / LTE</td>
  <td>
    <span class="dot red" id="sim_dot"></span>
    <span id="sim_status">--</span>
  </td>
  <td id="sim_info">--</td>
</tr>


          <tr>
  <td>Internet</td>
  <td>
    <span class="dot red" id="net_dot"></span>
    <span id="net_status">--</span>
  </td>
  <td id="net_info">--</td>
</tr>


          <tr>
            <td>Active Path</td>
            <td colspan="2" id="active_path">--</td>
          </tr>

        </table>

      </div>
    </div>

    <!-- ================= CONNECTION MANAGEMENT ================= -->
    <div class="panel">
      <div class="panel-header">
        Connection Management
        <span class="hint">Policy & override</span>
      </div>
      <div class="panel-content">

        <div class="radio-group">
          <label>
            <input type="radio" name="conn_mode" value="auto" checked
                   onchange="setConnMode('auto')">
            Automatic (System decides)
          </label>

          <label>
            <input type="radio" name="conn_mode" value="manual"
                   onchange="setConnMode('manual')">
            Manual (Operator selects)
          </label>
        </div>

        <div class="row2" style="margin-top:12px;">
          <button class="btn secondary" id="btn_wifi"
                  onclick="requestManualPath('wifi')" disabled>
            Use Wi-Fi
          </button>

          <button class="btn secondary" id="btn_sim"
                  onclick="requestManualPath('sim')" disabled>
            Use SIM / LTE
          </button>
        </div>

        <div class="note" style="margin-top:10px;">
          Manual selection is only allowed when policy is set to Manual.
        </div>

      </div>
    </div>

    <!-- ================= OFFLINE SYNC ================= -->
    <div class="panel">
      <div class="panel-header">
        Offline Data Sync
      </div>
      <div class="panel-content">

        <div class="data-grid">
          <div class="card">
            <div class="label">Pending Records</div>
            <div class="value" id="sync_pending">--</div>
          </div>

          <div class="card">
            <div class="label">Last Sync</div>
            <div class="value" id="sync_last">--</div>
          </div>
        </div>

        <div style="margin-top:12px;">
          <button class="btn primary" onclick="requestSync()">
            Sync Offline Records
          </button>

          <button class="btn secondary" onclick="refreshConnectivity()">
            Refresh Status
          </button>
        </div>

      </div>
    </div>

  </div>
</div>

    </div>

    

  </div>
</div>


    <!-- ================= WEATHER TAB ================= -->
<div class="tab-content" id="tab-weather">
  <div class="grid2">

    <!-- -------- LEFT: WEATHER DATA -------- -->
    <div class="panel">
      <div class="panel-header">Weather Feed</div>
      <div class="panel-content">

        <div class="data-grid">

          <!-- Core Weather -->
          <div class="card">
            <div class="label">Wind Speed (m/s)</div>
            <div class="value" id="wspd">--</div>
          </div>

          <div class="card">
            <div class="label">Wind Direction </div>
            <div class="value" id="wdir">--</div>
          </div>

          <div class="card">
            <div class="label">Temperature (C)</div>
            <div class="value" id="temp">--</div>
          </div>

          <div class="card">
            <div class="label">Humidity (%RH)</div>
            <div class="value" id="hum">--</div>
          </div>

          <div class="card">
            <div class="label">Pressure (hPa)</div>
            <div class="value" id="pres">--</div>
          </div>

          <div class="card">
            <div class="label">Last Update</div>
            <div class="value red" id="wts">--</div>
          </div>

          <!-- Environmental (from Firebase) -->
          <div class="card">
            <div class="label">Noise (dB)</div>
            <div class="value" id="noise">--</div>
          </div>

          <div class="card">
            <div class="label">PM2.5 </div>
            <div class="value" id="pm25">--</div>
          </div>

          <div class="card">
            <div class="label">PM10 </div>
            <div class="value" id="pm10">--</div>
          </div>

          <div class="card">
            <div class="label">Rainfall (mm)</div>
            <div class="value" id="rain">--</div>
          </div>

        </div>

        <div style="margin-top:10px">
          <button class="btn secondary"
                  onclick="notify('Weather feed is Active (Local DB)', 'ok')">
            Status
          </button>
        </div>

      </div>
    </div>

    <!-- -------- RIGHT: WEATHER MAP -------- -->
    <div class="panel">
      <div class="panel-header">Map (Weather Context)</div>
      <div class="panel-content">

        <div class="mapbox" id="mapWeather"></div>

        <div style="height:10px"></div>

        <div class="note">
          Weather data is linked to vessel position for contextual awareness.
        </div>

      </div>
    </div>

  </div>
</div>
<!-- =============== END WEATHER TAB =============== -->


    <!-- CAMERA TAB -->
<div class="tab-content" id="tab-camera">
  <div class="grid3">

    <!-- ================= PTZ FEED ================= -->
    <div class="panel">
      <div class="panel-header">
        PTZ Feed <span class="hint">Mode: {{ camera_mode }}</span>
      </div>

      <div class="panel-content">

        <!-- CAMERA VIEW -->
        <div class="cam-wrap">
          <img id="camImg" src="/camera/stream" alt="Camera feed">
          <div class="badge">
            CAMERA: <b id="camMode">{{ camera_mode }}</b>
          </div>
        </div>
<!-- IMAGE ENCODING + FIREBASE OPTIONS -->
<div class="row2" style="margin-top:10px; align-items:center;">
  <label class="hint" style="display:flex; gap:8px; align-items:center;">
    <input type="checkbox" id="fbUpload">
    Upload to Firebase (disabled offline)

  </label>

  <div style="display:flex; gap:10px; justify-content:flex-end; flex-wrap:wrap;">
    <label class="hint">
      JPEG Quality
      <select id="jpgQuality">
        <option value="60">60</option>
        <option value="70">70</option>
        <option value="75" selected>75</option>
        <option value="85">85</option>
      </select>
    </label>

    <label class="hint">
      Max Width
      <select id="jpgWidth">
        <option value="640">640</option>
        <option value="960">960</option>
        <option value="1280" selected>1280</option>
      </select>
    </label>
  </div>
</div>

        <!-- CAPTURE BUTTONS -->
<div class="row2" style="margin-top:12px;">
  <button class="btn secondary" onclick="captureFromPTZ('vjr')">
    CAPTURE ? VJR
  </button>

  <button class="btn secondary" onclick="captureFromPTZ('vdr')">
    CAPTURE ? VDR
  </button>

  <!-- NEW: FIREBASE CAPTURE BUTTON -->
  <button class="btn primary" onclick="captureToFirebase()">
    CAPTURE ? FIREBASE
  </button>
</div>


        <!-- DELETE BUTTONS -->
        <div class="row2" style="margin-top:10px;">
          <button class="btn danger" onclick="deleteLastCapture('vjr')">
            DELETE LAST (VJR)
          </button>
          <button class="btn danger" onclick="deleteLastCapture('vdr')">
            DELETE LAST (VDR)
          </button>
        </div>

        <!-- CONNECTION CONTROLS -->
        <div style="height:10px"></div>
        <div class="row2">
          <button class="btn secondary" onclick="reconnectCamera()">
            Reconnect
          </button>
          <button class="btn secondary" onclick="toggleCamRefresh()">
            Refresh On/Off
          </button>
        </div>

        <div style="height:10px"></div>
        <div class="note" id="camNote">
          PTZ controls are demo-safe (will respond even without hardware).
        </div>

      </div>
    </div>

    <!-- ================= PTZ CONTROLS ================= -->
    <div class="panel">
      <div class="panel-header">PTZ Controls</div>

      <div class="panel-content">

        <!-- JOYSTICK -->
        <div class="joy" id="joy">
          <div class="stick" id="stick"></div>
        </div>

        <div style="height:10px"></div>

        <!-- ZOOM -->
        <div class="row2">
          <button class="btn secondary" onclick="controlCamera('zoom_in')">
            Zoom +
          </button>
          <button class="btn secondary" onclick="controlCamera('zoom_out')">
            Zoom -
          </button>
        </div>

        <div style="height:10px"></div>

        <!-- CAMERA MODES -->
        <div class="row2">
          <button class="btn secondary" onclick="controlCamera('led_toggle')">
            LED
          </button>
          <button class="btn secondary" onclick="controlCamera('night_vision_toggle')">
            Night
          </button>
        </div>

        <div style="height:10px"></div>

        <!-- PTZ SYNC -->
        <div class="row2">
          <button class="btn secondary" onclick="setPTZSync(true)">
            Sync COG
          </button>
          <button class="btn secondary" onclick="setPTZSync(false)">
            Unsync
          </button>
        </div>

      </div>
    </div>

    <!-- ================= CAMERA STATUS ================= -->
    <div class="panel">
      <div class="panel-header">Camera Status</div>

      <div class="panel-content">
        <div class="raw" id="camStatus">Loading...</div>
        <button class="btn secondary" onclick="updateCameraStatusUI()">
          Refresh Status
        </button>
      </div>
    </div>

  </div>
</div>


    <!-- VJR TAB -->
    <div class="tab-content" id="tab-vjr">
      <div class="grid2">
        <div class="panel">
          <div class="panel-header">VJR - Journey Details</div>
          <div class="panel-content">
            <div class="form-group">
              <label class="form-label">Report Date</label>
              <input id="vjr_date" placeholder="YYYY-MM-DD">
            </div>
            <div class="form-group">
              <label class="form-label">Vessel</label>
              <input id="vjr_vessel" placeholder="Vessel name">
            </div>
            <div class="row2">
              <div class="form-group">
                <label class="form-label">IMO</label>
                <input id="vjr_imo" placeholder="IMO">
              </div>
              <div class="form-group">
                <label class="form-label">Master</label>
                <input id="vjr_master" placeholder="Master name">
              </div>
            </div>
            <div class="row2">
              <div class="form-group">
                <label class="form-label">Departure</label>
                <input id="vjr_departure" placeholder="Port/Location">
              </div>
              <div class="form-group">
                <label class="form-label">Arrival</label>
                <input id="vjr_arrival" placeholder="Port/Location">
              </div>
            </div>

            <div class="row2">
              <button class="btn secondary" onclick="captureImage('vjr')">Capture Image (VJR)</button>
              <button class="btn secondary" onclick="exportVJRPDF()">Export VJR PDF</button>
            </div>
            <div style="height:10px"></div>
            <div class="row2">
              <button class="btn danger" onclick="clearCaptures('vjr')">Clear Captures (VJR)</button>
              <button class="btn secondary" onclick="notify('VJR will include navigation log + images.', 'ok')">Info</button>
            </div>
          </div>
        </div>

        <div class="panel">
          <div class="panel-header">VJR - Captured Images <span class="hint" id="vjrCount">0</span></div>
          <div class="panel-content">
            <div class="image-grid" id="vjrGallery"></div>
            <div style="height:10px"></div>
            <div class="note">Captured images will be embedded into the VJR PDF report.</div>
          </div>
        </div>
      </div>
    </div>

    <!-- VDR TAB -->
    <div class="tab-content" id="tab-vdr">
      <div class="grid2">
        <div class="panel">
          <div class="panel-header">VDR - Daily Report Entry</div>
          <div class="panel-content">
            <div class="form-group"><label class="form-label">Date</label><input id="vdr_date" placeholder="YYYY-MM-DD"></div>
            <div class="row2">
              <div class="form-group"><label class="form-label">Vessel</label><input id="vdr_vessel" placeholder="Vessel"></div>
              <div class="form-group"><label class="form-label">IMO</label><input id="vdr_imo" placeholder="IMO"></div>
            </div>
            <div class="row2">
              <div class="form-group"><label class="form-label">MMSI</label><input id="vdr_mmsi" placeholder="MMSI"></div>
              <div class="form-group"><label class="form-label">Call Sign</label><input id="vdr_callsign" placeholder="Call Sign"></div>
            </div>
            <div class="row2">
              <div class="form-group"><label class="form-label">Client</label><input id="vdr_client" placeholder="Client"></div>
              <div class="form-group"><label class="form-label">Location</label><input id="vdr_location" placeholder="Location"></div>
            </div>
            <div class="row2">
              <div class="form-group"><label class="form-label">Country</label><input id="vdr_country" placeholder="Country"></div>
              <div class="form-group"><label class="form-label">Activity</label><input id="vdr_activity" placeholder="Activity"></div>
            </div>
            <div class="row2">
              <div class="form-group"><label class="form-label">Weather</label><input id="vdr_weather" placeholder="Weather"></div>
              <div class="form-group"><label class="form-label">Sea State</label><input id="vdr_sea_state" placeholder="Sea State"></div>
            </div>
            <div class="row2">
              <div class="form-group"><label class="form-label">Crew Count</label><input id="vdr_crew_count" placeholder="Crew"></div>
              <div class="form-group"><label class="form-label">Fuel Consumption</label><input id="vdr_fuel_consumption" placeholder="Fuel"></div>
            </div>
            <div class="form-group"><label class="form-label">Remarks</label><textarea id="vdr_remarks" placeholder="Remarks"></textarea></div>

            <div class="row2">
              <button class="btn secondary" onclick="captureImage('vdr')">Capture Image (VDR)</button>
              <button class="btn secondary" onclick="saveVDRRecord()">Save VDR Record</button>
            </div>
            <div style="height:10px"></div>
            <div class="row2">
              <button class="btn secondary" onclick="exportVDRPDF()">Export VDR PDF</button>
              <button class="btn danger" onclick="clearCaptures('vdr')">Clear Captures (VDR)</button>
            </div>
          </div>
        </div>

        <div class="panel">
          <div class="panel-header">VDR - Records & Images <span class="hint" id="vdrCount">0</span></div>
          <div class="panel-content">
            <div class="row2">
              <button class="btn secondary" onclick="exportVDRCSV()">Export CSV</button>
              <button class="btn secondary" onclick="exportVDRExcel()">Export Excel</button>
            </div>
            <div style="height:10px"></div>
            <div class="row2">
              <button class="btn danger" onclick="deleteAllVDRRecords()">Delete All Records</button>
              <button class="btn secondary" onclick="loadVDRList()">Refresh List</button>
            </div>
            <div style="height:12px"></div>
            <div class="image-grid" id="vdrGallery"></div>
            <div style="height:12px"></div>
            <table>
              <thead><tr><th>ID</th><th>Date</th><th>Vessel</th><th>Activity</th><th>Location</th><th>Weather</th></tr></thead>
              <tbody id="vdrTable"><tr><td colspan="6">No records</td></tr></tbody>
            </table>
          </div>
        </div>
      </div>
    </div>

    <!-- LOGS TAB -->
    <div class="tab-content" id="tab-logs">
      <div class="grid2">
        <div class="panel">
          <div class="panel-header">System Logs</div>
          <div class="panel-content">
            <div class="note">
              Navigation data is auto-logged into <b>nav_log.csv</b>. You can download it from the NAV tab.
              VDR records are stored in memory (demo). Images are stored in memory (demo).
            </div>
          </div>
        </div>

        <div class="panel">
          <div class="panel-header">Developer Info</div>
          <div class="panel-content">
            <div class="raw">
              WeasyPrint available: {{ 'YES' if weasy else 'NO' }}
              \nReportLab available: {{ 'YES' if True else 'NO' }}
            </div>
          </div>
        </div>
      </div>
    </div>
  </div>

  <div id="toast" class="notification" style="display:none"></div>


  <script>
    // Role helpers
    const ROLE_ORDER = {"Viewer":1,"Operator":2,"Captain":3};
    let CURRENT_ROLE = "{{ role }}";

    function can(minRole){
      return (ROLE_ORDER[CURRENT_ROLE] || 0) >= (ROLE_ORDER[minRole] || 0);
    }
function updateNetStatus(){
  const el = document.getElementById("netStatus");
  if(!el) return;

  if(navigator.onLine){
    el.innerHTML = "MODE: <b>ONLINE</b>";
    el.style.borderColor = "rgba(0,208,132,.45)";
  }else{
    el.innerHTML = "MODE: <b>OFFLINE</b>";
    el.style.borderColor = "rgba(255,204,0,.45)";
  }
}

window.addEventListener("online", updateNetStatus);
window.addEventListener("offline", updateNetStatus);
updateNetStatus();

    function notify(msg, kind="ok"){
      const t = document.getElementById("toast");
      t.className = "notification " + kind;
      t.textContent = msg;
      t.style.display = "block";
      clearTimeout(window.__toastTimer);
      window.__toastTimer = setTimeout(()=>{ t.style.display = "none"; }, 2600);
    }

    // Clock
    function tickClock(){
      const now = new Date();
      const pad = (n)=> String(n).padStart(2,"0");
      document.getElementById("clock").textContent =
        pad(now.getHours()) + ":" + pad(now.getMinutes()) + ":" + pad(now.getSeconds());
    }
    setInterval(tickClock, 1000); tickClock();

    // Tabs
    function openTab(tab){
  document.querySelectorAll(".tab-btn").forEach(b=>b.classList.remove("active"));
  document.querySelectorAll(".tab-content").forEach(c=>c.classList.remove("active"));
  document.querySelector(`.tab-btn[onclick="openTab('${tab}')"]`).classList.add("active");
  document.getElementById("tab-"+tab).classList.add("active");

// ---- NEW: tab-specific actions ----
  if(tab === "connectivity"){
    refreshConnectivity();
    refreshSyncStatus();
  }
  
  // ?? FORCE WEATHER REFRESH WHEN WEATHER TAB OPENS
  if(tab === "weather"){
    fetchWeatherData();
  }

  // map fix
  setTimeout(()=>{
    try{ mapNav && mapNav.invalidateSize(); }catch(e){}
    try{ mapWeather && mapWeather.invalidateSize(); }catch(e){}
  }, 200);
}


    // Map - NAV (MarineLite Realtime Position Only)
let mapNav, navMarker;

mapNav = L.map('mapNav').setView([3.006633, 101.380133], 12);

L.tileLayer('https://{s}.tile.openstreetmap.org/{z}/{x}/{y}.png', {
  maxZoom: 19
}).addTo(mapNav);

navMarker = L.marker([3.006633, 101.380133]).addTo(mapNav);


    // Map - Weather
    let mapWeather, weatherMarker;
    mapWeather = L.map('mapWeather').setView([3.006633, 101.380133], 12);
    L.tileLayer('https://{s}.tile.openstreetmap.org/{z}/{x}/{y}.png', { maxZoom: 19 }).addTo(mapWeather);
    weatherMarker = L.marker([3.006633, 101.380133]).addTo(mapWeather);

    // NAV CSV download
    function downloadNavCSV(){
      window.location = "/export_nav_csv";
    }

    // Camera stream refresh toggle
    let camRefresh = true;
    function toggleCamRefresh(){
      camRefresh = !camRefresh;
      notify("Camera refresh " + (camRefresh ? "ON" : "OFF"), camRefresh ? "ok" : "warn");
    }
    

    // Ensure all UI actions work (async fixes)
    async function controlCamera(action, value=0){
      if (!can("Operator")) return notify("Permission denied (Operator required).", "err");
      try{
        const res = await fetch("/camera/control", {
          method:"POST",
          headers:{ "Content-Type":"application/json" },
          body: JSON.stringify({action, value})
        });
        const j = await res.json().catch(()=>({}));
        if (res.ok){
          notify(j.message || "Control OK", "ok");
          updateCameraStatusUI();
        } else {
          notify(j.error || "Control failed", "err");
        }
      } catch(e){
        notify("Control error: " + e, "err");
      }
    }

    async function reconnectCamera(){
      if (!can("Operator")) return notify("Permission denied (Operator required).", "err");
      try{
        const res = await fetch("/camera/reconnect", { method:"POST" });
        const j = await res.json().catch(()=>({}));
        if (res.ok){
          notify("Reconnect: " + (j.reconnected ? "OK" : "FALLBACK"), j.reconnected ? "ok" : "warn");
          updateCameraStatusUI();
        } else {
          notify("Reconnect failed", "err");
        }
      } catch(e){
        notify("Reconnect error: " + e, "err");
      }
    }

    async function setPTZSync(enabled){
      if (!can("Operator")) return notify("Permission denied (Operator required).", "err");
      try{
        const res = await fetch("/camera/sync_cog", {
          method:"POST",
          headers:{ "Content-Type":"application/json" },
          body: JSON.stringify({enabled})
        });
        const j = await res.json().catch(()=>({}));
        if (res.ok){
          notify("PTZ Sync: " + (j.enabled ? "ON" : "OFF"), "ok");
          updateCameraStatusUI();
        } else {
          notify("Sync failed", "err");
        }
      } catch(e){
        notify("Sync error: " + e, "err");
      }
    }

    async function captureImage(type){
      if (!can("Operator")) return notify("Permission denied (Operator required).", "err");
      try{
        const res = await fetch("/camera/capture", {
          method:"POST",
          headers:{ "Content-Type":"application/json" },
          body: JSON.stringify({report_type:type})
        });
        const j = await res.json().catch(()=>({}));
        if (res.ok && j.status === "success"){
          notify("Captured image for " + type.toUpperCase(), "ok");
          updateCaptureCounts();
          loadCapturedImages();
        } else {
          notify("Capture failed: " + (j.error || j.message || res.status), "err");
        }
      } catch(e){
        notify("Capture error: " + e, "err");
      }
    }
function captureFromPTZ(reportType) {
// Firebase upload is disabled in offline-first architecture
const fbUpload = false;
  const jpgQuality = Number(document.getElementById("jpgQuality")?.value || 75);
  const jpgWidth = Number(document.getElementById("jpgWidth")?.value || 1280);

  fetch("/camera/capture", {
    method: "POST",
    headers: { "Content-Type": "application/json" },
    body: JSON.stringify({
      report_type: reportType,
      upload_firebase: fbUpload,
      jpg_quality: jpgQuality,
      max_width: jpgWidth
    })
  })
  .then(r => r.json())
  .then(d => {
    console.log("Capture result:", d);

    if (d.firebase_uploaded) {
      console.log("Uploaded to Firebase:", d.firebase_url);
    }
  })
  .catch(err => console.error("Capture error:", err));

  // force live feed refresh (avoid frozen frame)
  const img = document.getElementById("camImg");
  if (img) {
    img.src = "/camera/stream?t=" + Date.now();
  }
}
/* ===== ADD THIS RIGHT BELOW ===== */

function captureToFirebase() {
  fetch("/camera/capture", {
    method: "POST",
    headers: { "Content-Type": "application/json" },
    body: JSON.stringify({
      report_type: "firebase",
      upload_firebase: true,
      include_pixel_data: true,
      jpg_quality: 70,
      max_width: 1280,
      pixel_quality: 40
    })
  })
  .then(r => r.json())
  .then(d => {
    console.log("Firebase capture:", d);
    if (d.firebase_uploaded) {
      alert("? Image sent to Firebase successfully");
      console.log("Firebase URL:", d.firebase_url);
    } else {
      alert("? Capture done, but Firebase upload failed");
    }
  })
  .catch(err => {
    console.error("Firebase capture error:", err);
    alert("? Firebase capture error (see console)");
  });

  const img = document.getElementById("camImg");
  if (img) {
    img.src = "/camera/stream?t=" + Date.now();
  }
}


async function deleteLastCapture(reportType) {
  if (!can("Operator")) return notify("Permission denied (Operator required).", "err");

  try {
    const res = await fetch("/camera/captured_images_full");
    const data = await res.json().catch(() => ({}));

    const list = (reportType === "vjr") ? (data.vjr || []) : (data.vdr || []);

    if (!list.length) {
      return notify("No images to delete for " + reportType.toUpperCase(), "warn");
    }

    const payload = { report_type: reportType };

    if (list[0] && list[0].id){
      payload.capture_id = list[0].id;
    } else {
      payload.index = 0;
    }

    const del = await fetch("/camera/delete_capture", {
      method: "POST",
      headers: { "Content-Type": "application/json" },
      body: JSON.stringify(payload)
    });

    const j = await del.json().catch(() => ({}));

    if (del.ok && j.status === "success") {
      notify("Deleted last image (" + reportType.toUpperCase() + ")", "ok");
      updateCaptureCounts();
      loadCapturedImages();
    } else {
      notify("Delete failed", "err");
    }
  } catch (e) {
    notify("Delete error: " + e, "err");
  }
}


    async function updateCaptureCounts(){
      try{
        const res = await fetch("/camera/captured_images");
        const j = await res.json().catch(()=>({}));
        document.getElementById("vjrCount").textContent = j.vjr_images || 0;
        document.getElementById("vdrCount").textContent = j.vdr_images || 0;
      }catch(e){}
    }

    async function deleteCapture(type, captureId, index){
      if (!can("Operator")) return notify("Permission denied (Operator required).", "err");
      try{
        const payload = { report_type: type };

        // Prefer stable delete by capture_id. Fall back to index for old items.
        if (captureId && String(captureId).trim().length > 0){
          payload.capture_id = captureId;
          console.log("Deleting", type, "capture_id", captureId);
        } else {
          payload.index = index;
          console.log("Deleting", type, "index", index);
        }

        const res = await fetch("/camera/delete_capture", {
          method:"POST",
          headers:{ "Content-Type":"application/json" },
          body: JSON.stringify(payload)
        });

        const j = await res.json().catch(()=>({}));
        console.log("Delete response:", res.ok, j);

        if (res.ok && j.status === "success"){
          notify("Deleted image", "ok");
          updateCaptureCounts();
          loadCapturedImages();
        } else {
          notify("Delete failed: " + (j.error || j.message || res.status), "err");
        }
      } catch(e){
        notify("Delete error: " + e, "err");
      }
    }

    async function clearCaptures(type){
      if (!can("Operator")) return notify("Permission denied (Operator required).", "err");
      if (!confirm("Clear all captured images for " + type.toUpperCase() + "?")) return;
      try{
        const res = await fetch("/camera/clear_captures", {
          method:"POST",
          headers:{ "Content-Type":"application/json" },
          body: JSON.stringify({report_type:type})
        });
        const j = await res.json().catch(()=>({}));
        console.log("Clear response:", res.ok, j);
        if (res.ok && j.status === "success"){
          notify("Cleared captures for " + type.toUpperCase(), "ok");
          updateCaptureCounts();
          loadCapturedImages();
        } else {
          notify("Clear failed: " + (j.error || j.message || res.status), "err");
        }
      } catch(e){
        notify("Clear error: " + e, "err");
      }
    }
    async function loadCapturedImages(){
      try{
        const res = await fetch("/camera/captured_images_full");
        const j = await res.json().catch(()=>({}));
        const vjr = j.vjr || [];
        const vdr = j.vdr || [];

        const vjrG = document.getElementById("vjrGallery");
        const vdrG = document.getElementById("vdrGallery");
        vjrG.innerHTML = "";
        vdrG.innerHTML = "";

        vjr.forEach((img, idx)=>{
          const d = document.createElement("div");
          d.className = "thumb";
          d.innerHTML = `
            <img src="data:image/jpeg;base64,${img.data}" />
<button class="xbtn" title="Delete" onclick="deleteCapture('vjr', '${img.id || ''}', ${idx})">&times;</button>
            <div class="cap">${img.timestamp || ""}</div>
          `;
          vjrG.appendChild(d);
        });

        vdr.forEach((img, idx)=>{
          const d = document.createElement("div");
          d.className = "thumb";
          d.innerHTML = `
            <img src="data:image/jpeg;base64,${img.data}" />
<button class="xbtn" title="Delete" onclick="deleteCapture('vdr', '${img.id || ''}', ${idx})">&times;</button>
            <div class="cap">${img.timestamp || ""}</div>
          `;
          vdrG.appendChild(d);
        });

      }catch(e){
        console.warn("loadCapturedImages error", e);
      }
    }

    async function updateCameraStatusUI(){
      try{
        const res = await fetch("/camera/status");
        const j = await res.json().catch(()=>({}));
        document.getElementById("camMode").textContent = (j.mode || "demo").toUpperCase();
        document.getElementById("camStatus").textContent = JSON.stringify(j, null, 2);
        const note = document.getElementById("camNote");
        if (j.mode === "ip"){
          note.textContent = "Connected to IP camera. Snapshot capture uses first JPEG frame extraction.";
        } else {
          note.textContent = "Demo mode (no live camera). Controls remain functional.";
        }
      }catch(e){
        document.getElementById("camStatus").textContent = "Status error: " + e;
      }
    }

    // NAV + WEATHER polling
    async function fetchNavigationData(){
  try{
    const res = await fetch("/nav_data");
    const n = await res.json().catch(()=>({}));

    if(!n.latitude || !n.longitude) return;

    // Update NAV text fields
    document.getElementById("nav_lat").textContent =
  Math.abs(Number(n.latitude)).toFixed(6) + (Number(n.latitude) >= 0 ? " N" : " S");

document.getElementById("nav_lon").textContent =
  Math.abs(Number(n.longitude)).toFixed(6) + (Number(n.longitude) >= 0 ? " E" : " W");

    const hdg = Number(n.heading);
document.getElementById("nav_heading").textContent =
  Number.isFinite(hdg) ? `${hdg.toFixed(0)}` : "N/A";


    if(n.timestamp){
      const t = new Date(n.timestamp * 1000);
      document.getElementById("nav_time").textContent = t.toLocaleTimeString();
    }
// GNSS status indicator (no UI change)
const fix = document.getElementById("nav_fix");
if(!n.timestamp){
  fix.textContent = "NO DATA";
}else{
  const age = Math.floor(Date.now()/1000 - Number(n.timestamp));
  if(age <= 5) fix.textContent = "LIVE";
  else if(age <= 30) fix.textContent = "STALE";
  else fix.textContent = "NO FIX";
}

    // Update map marker
    navMarker.setLatLng([n.latitude, n.longitude]);
    mapNav.panTo([n.latitude, n.longitude], { animate: true });
// Sync Weather map position with NAV (GNSS)
weatherMarker.setLatLng([n.latitude, n.longitude]);
mapWeather.panTo([n.latitude, n.longitude], { animate: true });


  }catch(e){
    console.error("NAV fetch failed", e);
  }
}


    async function fetchWeatherData(){
  try{
    const res = await fetch("/weather_data");
    const j = await res.json();

    document.getElementById("wspd").innerText  = j.wind_speed ?? "--";
    document.getElementById("wdir").innerText  = j.wind_direction ?? "--";
    document.getElementById("temp").innerText  = j.temperature ?? "--";
    document.getElementById("hum").innerText   = j.humidity ?? "--";
    document.getElementById("pres").innerText  = j.pressure ?? "--";
    document.getElementById("noise").innerText = j.noise ?? "--";
    document.getElementById("pm25").innerText  = j.pm25 ?? "--";
    document.getElementById("pm10").innerText  = j.pm10 ?? "--";
    document.getElementById("rain").innerText  = j.rainfall ?? "--";

    if(j.timestamp){
      document.getElementById("wts").innerText =
        new Date(j.timestamp * 1000).toLocaleString();
    }else{
      document.getElementById("wts").innerText = "--";
    }

  }catch(e){
    console.error("Weather fetch failed", e);
  }
}


    // Export VJR PDF
    async function exportVJRPDF(){
      if (!can("Operator")) return notify("Permission denied (Operator required).", "err");

      const vjr = {
        date: document.getElementById("vjr_date").value,
        vessel: document.getElementById("vjr_vessel").value,
        imo: document.getElementById("vjr_imo").value,
        master: document.getElementById("vjr_master").value,
        departure: document.getElementById("vjr_departure").value,
        arrival: document.getElementById("vjr_arrival").value
      };

      // take nav samples from the live endpoint (most recent)
      let nav = [];
      try{
        const resNav = await fetch("/nav_data");
        const jNav = await resNav.json().catch(()=>({}));
        nav = (jNav.history || []).slice(0, 40);
      }catch(e){}

      try{
        const res = await fetch("/export_vjr_pdf", {
          method:"POST",
          headers:{ "Content-Type":"application/json" },
          body: JSON.stringify({vjr, nav})
        });
        if (!res.ok){
          const j = await res.json().catch(()=>({}));
          return notify("PDF failed: " + (j.error || res.status), "err");
        }
        const blob = await res.blob();
        const url = URL.createObjectURL(blob);
        const a = document.createElement("a");
        a.href = url;
        a.download = "VJR_Report.pdf";
        document.body.appendChild(a);
        a.click();
        a.remove();
        URL.revokeObjectURL(url);
        notify("VJR PDF downloaded.", "ok");
      }catch(e){
        notify("PDF error: " + e, "err");
      }
    }

    // VDR record operations
    async function saveVDRRecord(){
      if (!can("Operator")) return notify("Permission denied (Operator required).", "err");

      const record = {
        date: document.getElementById("vdr_date").value,
        vessel: document.getElementById("vdr_vessel").value,
        imo: document.getElementById("vdr_imo").value,
        mmsi: document.getElementById("vdr_mmsi").value,
        callsign: document.getElementById("vdr_callsign").value,
        client: document.getElementById("vdr_client").value,
        location: document.getElementById("vdr_location").value,
        country: document.getElementById("vdr_country").value,
        activity: document.getElementById("vdr_activity").value,
        weather: document.getElementById("vdr_weather").value,
        sea_state: document.getElementById("vdr_sea_state").value,
        crew_count: document.getElementById("vdr_crew_count").value,
        fuel_consumption: document.getElementById("vdr_fuel_consumption").value,
        remarks: document.getElementById("vdr_remarks").value
      };

      try{
        const res = await fetch("/save_vdr", {
          method:"POST",
          headers:{ "Content-Type":"application/json" },
          body: JSON.stringify(record)
        });
        const j = await res.json().catch(()=>({}));
        if (res.ok){
          notify("VDR record saved. Total: " + (j.total || 0), "ok");
          loadVDRList();
        }else{
          notify("Save failed: " + (j.error || res.status), "err");
        }
      }catch(e){
        notify("Save error: " + e, "err");
      }
    }

    async function loadVDRList(){
      try{
        const res = await fetch("/vdr_records");
        const rows = await res.json().catch(()=>[]);
        const tbody = document.getElementById("vdrTable");
        if (!rows.length){
          tbody.innerHTML = `<tr><td colspan="6">No records</td></tr>`;
          return;
        }
        tbody.innerHTML = rows.slice(0, 30).map(r => `
          <tr>
            <td>${r.id || ""}</td>
            <td>${r.date || ""}</td>
            <td>${r.vessel || ""}</td>
            <td>${r.activity || ""}</td>
            <td>${r.location || ""}</td>
            <td>${r.weather || ""}</td>
          </tr>
        `).join("");
      }catch(e){}
    }

    async function exportVDRCSV(){
      if (!can("Operator")) return notify("Permission denied (Operator required).", "err");
      try{
        const resList = await fetch("/vdr_records");
        const records = await resList.json().catch(()=>[]);
        const res = await fetch("/export_vdr_csv", {
          method:"POST",
          headers:{ "Content-Type":"application/json" },
          body: JSON.stringify({records})
        });
        const blob = await res.blob();
        const url = URL.createObjectURL(blob);
        const a = document.createElement("a");
        a.href = url;
        a.download = "VDR_Report.csv";
        document.body.appendChild(a);
        a.click();
        a.remove();
        URL.revokeObjectURL(url);
        notify("CSV downloaded.", "ok");
      }catch(e){
        notify("CSV error: " + e, "err");
      }
    }

    async function exportVDRExcel(){
      if (!can("Operator")) return notify("Permission denied (Operator required).", "err");
      try{
        const resList = await fetch("/vdr_records");
        const records = await resList.json().catch(()=>[]);
        const res = await fetch("/export_vdr_excel", {
          method:"POST",
          headers:{ "Content-Type":"application/json" },
          body: JSON.stringify({records})
        });
        if (!res.ok){
          const j = await res.json().catch(()=>({}));
          return notify("Excel failed: " + (j.error || res.status), "err");
        }
        const blob = await res.blob();
        const url = URL.createObjectURL(blob);
        const a = document.createElement("a");
        a.href = url;
        a.download = "VDR_Report.xlsx";
        document.body.appendChild(a);
        a.click();
        a.remove();
        URL.revokeObjectURL(url);
        notify("Excel downloaded.", "ok");
      }catch(e){
        notify("Excel error: " + e, "err");
      }
    }

    async function exportVDRPDF(){
      if (!can("Operator")) return notify("Permission denied (Operator required).", "err");
      try{
        const resList = await fetch("/vdr_records");
        const records = await resList.json().catch(()=>[]);
        const res = await fetch("/export_vdr_pdf", {
          method:"POST",
          headers:{ "Content-Type":"application/json" },
          body: JSON.stringify({records})
        });
        if (!res.ok){
          const j = await res.json().catch(()=>({}));
          return notify("PDF failed: " + (j.error || res.status), "err");
        }
        const blob = await res.blob();
        const url = URL.createObjectURL(blob);
        const a = document.createElement("a");
        a.href = url;
        a.download = "VDR_Report.pdf";
        document.body.appendChild(a);
        a.click();
        a.remove();
        URL.revokeObjectURL(url);
        notify("VDR PDF downloaded.", "ok");
      }catch(e){
        notify("PDF error: " + e, "err");
      }
    }

    async function deleteAllVDRRecords(){
      if (!can("Captain")) return notify("Permission denied (Captain required).", "err");
      if (!confirm("Delete ALL VDR records?")) return;
      try{
        const res = await fetch("/clear_vdr", { method:"POST" });
        const j = await res.json().catch(()=>({}));
        if (res.ok){
          notify("All VDR records deleted.", "ok");
          loadVDRList();
        } else {
          notify("Delete failed: " + (j.error || res.status), "err");
        }
      }catch(e){
        notify("Delete error: " + e, "err");
      }
    }

    // Joystick
    const joy = document.getElementById("joy");
    const stick = document.getElementById("stick");
    let dragging = false;

    function setStick(dx, dy){
      const r = 60;
      const len = Math.min(r, Math.hypot(dx, dy));
      const ang = Math.atan2(dy, dx);
      const x = len * Math.cos(ang);
      const y = len * Math.sin(ang);
      stick.style.transform = `translate(${x}px, ${y}px)`;
      return {x, y};
    }

    async function up(dx, dy){
      // map stick to PTZ actions
      const ax = Math.abs(dx), ay = Math.abs(dy);
      if (ax < 10 && ay < 10) return;
      if (ax > ay){
        await controlCamera(dx > 0 ? "pan_right" : "pan_left");
      }else{
        await controlCamera(dy > 0 ? "tilt_down" : "tilt_up");
      }
    }

    joy.addEventListener("pointerdown", (e)=>{
      dragging = true;
      joy.setPointerCapture(e.pointerId);
    });
    joy.addEventListener("pointerup", ()=>{
      dragging = false;
      stick.style.transform = "translate(-50%,-50%)";
    });
    joy.addEventListener("pointermove", (e)=>{
      if (!dragging) return;
      const rect = joy.getBoundingClientRect();
      const dx = e.clientX - (rect.left + rect.width/2);
      const dy = e.clientY - (rect.top + rect.height/2);
      setStick(dx, dy);
      up(dx, dy);
    });

    // Init
    updateCameraStatusUI();
    updateCaptureCounts();
    loadCapturedImages();
    loadVDRList();

    setInterval(fetchNavigationData, 1000);
    setInterval(fetchWeatherData, 2000);
    fetchNavigationData(); fetchWeatherData();
  /* ================= PTZ JOYSTICK LOGIC ================= */

let joyActive = false;
let joyCenter = { x: 0, y: 0 };

document.addEventListener("DOMContentLoaded", () => {
  const joy = document.getElementById("joy");
  const stick = document.getElementById("stick");

  if (!joy || !stick) return;

  // Start dragging
  joy.addEventListener("pointerdown", (e) => {
    joyActive = true;
    const r = joy.getBoundingClientRect();
    joyCenter = {
      x: r.left + r.width / 2,
      y: r.top + r.height / 2
    };
    joy.setPointerCapture(e.pointerId);
  });

  // Drag move
  document.addEventListener("pointermove", (e) => {
    if (!joyActive) return;

    let dx = e.clientX - joyCenter.x;
    let dy = e.clientY - joyCenter.y;

    const max = 45; // joystick radius
    dx = Math.max(-max, Math.min(max, dx));
    dy = Math.max(-max, Math.min(max, dy));

    stick.style.transform = `translate(${dx}px, ${dy}px)`;

    // Send PTZ commands
    if (Math.abs(dx) > 12) {
      controlCamera(dx > 0 ? "pan_right" : "pan_left");
    }
    if (Math.abs(dy) > 12) {
      controlCamera(dy > 0 ? "tilt_down" : "tilt_up");
    }
  });

  // Release joystick
  document.addEventListener("pointerup", () => {
    if (!joyActive) return;
    joyActive = false;
    stick.style.transform = "translate(-50%, -50%)";
  });
});
document.addEventListener("DOMContentLoaded", () => {
    fetchWeatherData();
    setInterval(fetchWeatherData, 2000);
  });
  let currentConnMode = "auto";

/* ---- Refresh connectivity status ---- */
function refreshConnectivity(){
  fetch("/api/connectivity/status")
    .then(r => r.json())
    .then(d => {
      // Wi-Fi
      document.getElementById("wifi_status").innerText =
        d.wifi.connected ? "Connected" : "Disconnected";
        document.getElementById("wifi_dot").className =
  "dot " + (d.wifi.connected ? "green" : "red");

      document.getElementById("wifi_info").innerText =
        d.wifi.ssid || "--";

      // SIM
      document.getElementById("sim_status").innerText =
        d.sim.connected ? "Connected" : "Disconnected";
        document.getElementById("sim_dot").className =
  "dot " + (d.sim.connected ? "green" : "red");

      document.getElementById("sim_info").innerText =
        d.sim.operator || "--";

      // Internet
      document.getElementById("net_status").innerText =
        d.internet ? "Available" : "Offline";
        document.getElementById("net_dot").className =
  "dot " + (d.internet ? "green" : "red");

      document.getElementById("net_info").innerText =
        d.internet ? "Online" : "No route";

      // Active path
      document.getElementById("active_path").innerText =
        d.active_path || "None";

      // Policy
      currentConnMode = d.policy?.mode || "auto";
      document.querySelector(
        `input[name="conn_mode"][value="${currentConnMode}"]`
      ).checked = true;

      updateManualButtons();
    })
    .catch(()=>{
      notify("Failed to read connectivity status", "error");
    });
}


function refreshSyncStatus(){
  fetch("/api/sync/status")
    .then(r => r.json())
    .then(d => {
      document.getElementById("sync_pending").innerText =
        d.pending_records;
      document.getElementById("sync_last").innerText =
        d.last_sync || "--";
    })
    .catch(()=>{});
}

function requestSync(){
  fetch("/api/sync/start", {method:"POST"})
    .then(r => r.json())
    .then(d => notify(d.message, "info"))
    .catch(()=>notify("Sync request failed","error"));
}


/* ---- Auto / Manual toggle ---- */
function setConnMode(mode){
  currentConnMode = mode;
  updateManualButtons();

  fetch("/api/connectivity/policy", {
    method: "POST",
    headers: {"Content-Type":"application/json"},
    body: JSON.stringify({ mode: mode })
  }).catch(()=>{});
}

/* ---- Enable / disable buttons ---- */
function updateManualButtons(){
  const manual = (currentConnMode === "manual");
  document.getElementById("btn_wifi").disabled = !manual;
  document.getElementById("btn_sim").disabled = !manual;
}

/* ---- Manual Wi-Fi / SIM request ---- */
function requestManualPath(path){
  fetch("/api/connectivity/policy", {
    method: "POST",
    headers: {"Content-Type":"application/json"},
    body: JSON.stringify({
      mode: "manual",
      preferred: path
    })
  })
  .then(()=>{
    notify(`Manual path requested: ${path.toUpperCase()}`, "info");
  })
  .catch(()=>{
    notify("Request failed", "error");
  });
}
  </script>
</body>
</html>
'''
# ======================================================
# Firebase initialization & background listeners
# ======================================================

# Initialize Firebase ONCE before starting any listeners
""" if not init_firebase():
    print("? Firebase failed to initialize at startup")
else:
    print("? Firebase initialized at startup")

# Start listeners AFTER Firebase init
threading.Thread(
    target=firebase_weather_listener,
    daemon=True
).start()

threading.Thread(
    target=firebase_marinelite_listener,
    daemon=True
).start() """
""" def firebase_poll_loop():
    global nav_current, weather_current

    print("?? Firebase polling loop started")

    while True:
        try:
            nav = db.reference(
                "gsd_vessel_sensor_data/marinelite/current"
            ).get()

            if isinstance(nav, dict):
                with nav_lock:
                    nav_current["latitude"]  = nav.get("latitude")
                    nav_current["longitude"] = nav.get("longitude")
                    nav_current["heading"]   = nav.get("heading")
                    nav_current["timestamp"] = nav.get("timestamp")

            weather = db.reference(
                "gsd_vessel_sensor_data/weather/current"
            ).get()

            if isinstance(weather, dict):
                with weather_lock:
                    weather_current.update({
                        "wind_speed": weather.get("wind_speed"),
                        "wind_direction": weather.get("wind_dir"),
                        "temperature": weather.get("temperature"),
                        "humidity": weather.get("humidity"),
                        "pressure": weather.get("pressure"),
                        "noise": weather.get("noise"),
                        "pm25": weather.get("pm25"),
                        "pm10": weather.get("pm10"),
                        "rainfall": weather.get("rainfall"),
                        "timestamp": weather.get("timestamp"),
                    })

            print("? Firebase polled")

        except Exception as e:
            print("? Firebase poll error:", e)

        time.sleep(1)  # 1 Hz is perfect """

# ------------------------------------------------------------------------------
# MAIN
# ------------------------------------------------------------------------------

if __name__ == "__main__":
    init_camera()
    app.run(host="0.0.0.0", port=5000, debug=False)



#!/usr/bin/env python3
"""
GDS - Vessel Management System (Black/Red Theme)
Fixes/Features added (without breaking previous professional UI):
- Camera buttons functional even without camera (demo-safe)
- Demo camera stream returns a real JPEG frame (img tag works)
- Show captured images in VJR & VDR pages (gallery)
- Embed captured images in VJR PDF and VDR PDF
- Add PDF export button to BOTH VJR and VDR
- Keep live map in NAV, VJR, VDR using dummy ST6100-like data
- User roles: Captain / Operator / Viewer
- PTZ sync to vessel COG (toggle)
"""
OFFLINE_MODE = True
CONNECTIVITY_UI = True
SYNC_ENABLED = True
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from flask import (
    Flask, render_template_string, jsonify, request, Response,
    send_file, redirect, url_for, session
)
from flask_cors import CORS
import threading, time, random, math, json, requests, base64, os, uuid
import sqlite3

from datetime import datetime
from io import BytesIO, StringIO
import csv
from functools import wraps
# ---------- Firebase (Realtime DB + Storage) ----------
FIREBASE_ENABLED = False

if FIREBASE_ENABLED:
    import firebase_admin
    from firebase_admin import credentials, db, storage


# Image generation for demo camera frames
from PIL import Image, ImageDraw, ImageFont

try:
    import openpyxl
    from openpyxl.styles import Font as XLFont, PatternFill
except ImportError:
    openpyxl = None

# Optional PDF - ReportLab (Windows-friendly, pure Python)
try:
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

# WeasyPrint (fallback)
try:
    from weasyprint import HTML
    WEASYPRINT_AVAILABLE = True
except Exception:
    WEASYPRINT_AVAILABLE = False

# ------------------------------------------------------------------------------
# CONFIGURATION
# ------------------------------------------------------------------------------

IP_CAMERA_URL = "http://192.168.0.164:8080/video"
CAMERA_USERNAME = ""
CAMERA_PASSWORD = ""
AUTO_FALLBACK_TO_DEMO = True
CAMERA_CONNECTION_TIMEOUT = 5
# ---------- Firebase Config ----------
FIREBASE_KEY_PATH = os.path.join(os.path.dirname(__file__), "firebase_key.json")
FIREBASE_DB_URL = "https://gds-vessel-simulator-default-rtdb.asia-southeast1.firebasedatabase.app"
FIREBASE_STORAGE_BUCKET = "gds-vessel-simulator.appspot.com"  # usually <project-id>.appspot.com
VESSEL_ID = "demo_vessel"
# Local offline database
LOCAL_DB_ENABLED = True
LOCAL_DB_PATH = "/home/rpi2/ship_system/db/ship_data.db"

# Logo file in same folder
LOGO_FILE = os.path.join(os.path.dirname(__file__), "GDS Logo.jpg")

# Embedded fallback logo PNG (small placeholder)
FALLBACK_LOGO_PNG_BASE64 = (
    "iVBORw0KGgoAAAANSUhEUgAAAJYAAAAqCAYAAABkGxjYAAAACXBIWXMAAAsSAAALEgHS3X78AAAB"
    "x0lEQVR4nO3bMW7TQBiG4d8gBqQ0aUqkY0iQpDkA0mZQwB0hUQqg5cI3o0FQj9i5hQ9vZtqGQp7"
    "3m8qfZc9m1uW2yQAAAAAAAAAAAAAAAAAAAAAAAPg3y0xg1uC7h5cJtZ0fZp7x3b0c3g8x2oR6B"
    "mZq2Qb6i7m0d0m0c0m9sQyq7K3+fZb7cE2nqQWmP8f1g5s9qG2d2Q8j3l9pQqYwq0WmZc6m8o2"
    "o8w1b3q8v2y+Yh6b0w0m0fQk8x5lqz9xw0m0VQm8xqf3m1b0m0VQm8xqf3m1b0m0VQm8xqf3m1"
    "b0m0VQm8xqf3m1b0m0VQm8xqf3m1b0m0VQm8xqf3m1b0m0VQm8xqf3m1b0m0VQm8xqf3m1b0m0"
    "VQm8xqf3m1b0m0VQm8xqf3m1b0m0VQm8xqf3m1b0m0VQm8xqf3m1b0m0VQm8xqf3m0AAAAAAAA"
    "AAAAAAAAAAAAAAADwp/0B7c7f0w3zVnQAAAAASUVORK5CYII="
)

# Simple local users (change before production)
USERS = {
    "captain":  {"password": "captain123",  "role": "Captain"},
    "operator": {"password": "operator123", "role": "Operator"},
    "viewer":   {"password": "viewer123",   "role": "Viewer"},
}
ROLE_ORDER = {"Viewer": 1, "Operator": 2, "Captain": 3}

app = Flask(__name__)
app.secret_key = "CHANGE_THIS_TO_A_LONG_RANDOM_SECRET"
CORS(app)

# ------------------------------------------------------------------------------
# GLOBAL STATE
# ------------------------------------------------------------------------------

camera_status = {
    "connected": False, "mode": "demo", "url": IP_CAMERA_URL,
    "error": None, "last_check": None
}
camera_controls = {
    "pan": 0, "tilt": 0, "zoom": 1.0,
    "led_brightness": 50, "led_enabled": False,
    "night_vision": False, "autofocus": True, "white_balance": "auto"
}
captured_images = {"vjr_images": [], "vdr_images": []}

nav_lock = threading.Lock()
weather_lock = threading.Lock()
vdr_lock = threading.Lock()
camera_lock = threading.Lock()
image_lock = threading.Lock()
sync_lock = threading.Lock()

nav_history = []
nav_current = {
    "latitude": None,
    "longitude": None,
    "heading": None,
    "timestamp": None
}

weather_current = {
    "wind_speed": None,
    "wind_direction": None,
    "temperature": None,
    "humidity": None,
    "pressure": None,
    "noise": None,
    "pm25": None,
    "pm10": None,
    "rainfall": None,
    "timestamp": None
}

vdr_records = []

sim_lat, sim_lon, sim_heading, sim_speed = 3.006633, 101.380133, 45.0, 8.0

# PTZ sync to COG
ptz_sync_enabled = False

MAX_NAV_HISTORY = 800
MAX_CAPTURE_IMAGES = 50
# ---------- Firebase init (run once) ----------
_firebase_ready = False
_firebase_lock = threading.Lock()

def init_firebase():
    """Initialize firebase_admin once (DB + Storage). Safe to call multiple times."""
    global _firebase_ready
    if not FIREBASE_ENABLED:
        return False

    with _firebase_lock:
        if _firebase_ready:
            return True

        try:
            if not os.path.exists(FIREBASE_KEY_PATH):
                print("Firebase key not found:", FIREBASE_KEY_PATH)
                return False

            cred = credentials.Certificate(FIREBASE_KEY_PATH)

            # ?? IMPORTANT FIX
            if not firebase_admin._apps:
                firebase_admin.initialize_app(cred, {
                    "databaseURL": FIREBASE_DB_URL,
                    "storageBucket": FIREBASE_STORAGE_BUCKET,
                })

            _firebase_ready = True
            print("Firebase initialized OK.")
            return True

        except Exception as e:
            print("Firebase init error:", e)
            return False


def firebase_weather_listener():
    global weather_current

    if not init_firebase():
        return

    ref = db.reference("gds_vessel_sensor_data/weather/current")

    def _on_change(event):
        if not event.data:
            return

        with weather_lock:
            weather_current.update({
                "wind_speed": event.data.get("wind_speed"),
                "wind_direction": event.data.get("wind_dir"),
                "temperature": event.data.get("temperature"),
                "humidity": event.data.get("humidity"),
                "pressure": event.data.get("pressure"),
                "noise": event.data.get("noise"),
                "pm25": event.data.get("pm25"),
                "pm10": event.data.get("pm10"),
                "rainfall": event.data.get("rainfall"),
                "timestamp": event.data.get("timestamp"),
            })

        print("?? Weather updated:", weather_current)

    ref.listen(_on_change)

def firebase_marinelite_listener():
    global nav_current

    print("?? Starting MarineLite Firebase listener...")

    if not init_firebase():
        print("? Firebase init failed (MarineLite)")
        return

    try:
        ref = db.reference("gds_vessel_sensor_data/marinelite/current")
    except Exception as e:
        print("? Firebase reference error:", e)
        return

    def _on_change(event):
        global nav_current

        if not event.data:
            print("?? MarineLite event with no data")
            return

        print("?? Raw MarineLite Firebase data:", event.data)

        with nav_lock:
            nav_current["latitude"]  = event.data.get("latitude")
            nav_current["longitude"] = event.data.get("longitude")
            nav_current["heading"]   = event.data.get("heading")
            nav_current["timestamp"] = event.data.get("timestamp")

        print("? NAV updated:", nav_current)

    ref.listen(_on_change)


def upload_jpeg_to_firebase_storage(jpg_bytes: bytes, report_type: str):
    """
    Upload bytes to Firebase Storage and return (public_url, object_path).
    Note: For production, prefer signed URLs or authenticated access; public is fine for testing.
    """
    if not init_firebase():
        return None, None

    ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    uid = uuid.uuid4().hex[:10]
    report_type = (report_type or "vdr").lower()
    object_path = f"vessels/{VESSEL_ID}/captures/{report_type}/{ts}_{uid}.jpg"

    bucket = storage.bucket()
    blob = bucket.blob(object_path)
    blob.upload_from_string(jpg_bytes, content_type="image/jpeg")

    # For testing convenience (public). You can harden later.
    blob.make_public()
    return blob.public_url, object_path

def _db_connect():
    con = sqlite3.connect(LOCAL_DB_PATH, timeout=30, check_same_thread=False)
    con.execute("PRAGMA journal_mode=WAL;")
    return con


def db_get_latest_marinelite():
    con = sqlite3.connect(LOCAL_DB_PATH)
    row = con.execute("""
        SELECT ts, latitude, longitude, heading
        FROM nav_data
        ORDER BY id DESC LIMIT 1
    """).fetchone()
    con.close()

    if not row:
        return {}

    return {
        "timestamp": row[0],
        "latitude": row[1],
        "longitude": row[2],
        "heading": row[3]
    }




def db_get_latest_weather():
    con = sqlite3.connect(LOCAL_DB_PATH)
    row = con.execute("""
        SELECT ts, wind_speed, wind_dir, humidity, temperature,
               pressure, pm25, pm10, rainfall, noise
        FROM weather_data
        ORDER BY id DESC LIMIT 1
    """).fetchone()
    con.close()

    if not row:
        return {}

    return {
        "timestamp": row[0],
        "wind_speed": row[1],
        "wind_dir": row[2],
        "humidity": row[3],
        "temperature": row[4],
        "pressure": row[5],
        "pm25": row[6],
        "pm10": row[7],
        "rainfall": row[8],
        "noise": row[9]
    }


def push_capture_event_to_firebase(report_type: str, image_url: str, object_path: str, enc_cfg: dict):
    """Store capture metadata + NAV snapshot in Realtime DB."""
    if not init_firebase():
        return False

    # snapshot nav at capture time
    with nav_lock:
        nav_snapshot = nav_current.copy()

    payload = {
        "timestamp": datetime.utcnow().isoformat(),
        "vessel_id": VESSEL_ID,
        "report_type": (report_type or "vdr").lower(),
        "image_url": image_url,
        "storage_path": object_path,
        "encoding": enc_cfg,
        "nav": {
            "latitude": nav_snapshot.get("latitude"),
            "longitude": nav_snapshot.get("longitude"),
            "speed": nav_snapshot.get("speed"),
            "heading": nav_snapshot.get("heading"),
            "cog": nav_snapshot.get("cog"),
            "voltage": nav_snapshot.get("voltage"),
            "time": nav_snapshot.get("time"),
            "date": nav_snapshot.get("date"),
        },
        "source": "IP_CAMERA",
    }

    db.reference(f"vessels/{VESSEL_ID}/captures").push(payload)
    return True

# Navigation logging (persistent)
NAV_LOG_FILE = os.path.join(os.path.dirname(__file__), "nav_log.csv")
nav_log_lock = threading.Lock()

def _ensure_nav_log_header():
    """Create nav log file with header if it does not exist."""
    try:
        if not os.path.exists(NAV_LOG_FILE) or os.path.getsize(NAV_LOG_FILE) == 0:
            with nav_log_lock:
                with open(NAV_LOG_FILE, "a", newline="", encoding="utf-8") as f:
                    w = csv.writer(f)
                    w.writerow(["date","time","latitude","longitude","speed","cog","heading","voltage","panic","ext_heading","raw_string"])
    except Exception:
        pass

def append_nav_log(sample: dict):
    """Append one nav sample (thread-safe)."""
    try:
        _ensure_nav_log_header()
        with nav_log_lock:
            with open(NAV_LOG_FILE, "a", newline="", encoding="utf-8") as f:
                w = csv.writer(f)
                w.writerow([
                    sample.get("date",""),
                    sample.get("time",""),
                    sample.get("latitude",""),
                    sample.get("longitude",""),
                    sample.get("speed",""),
                    sample.get("cog",""),
                    sample.get("heading",""),
                    sample.get("voltage",""),
                    sample.get("panic",""),
                    sample.get("ext_heading",""),
                    sample.get("raw_string",""),
                ])
    except Exception:
        pass


# ------------------------------------------------------------------------------
# AUTH / ROLES
# ------------------------------------------------------------------------------

def current_user():
    return session.get("user")

def current_role():
    return session.get("role", "Viewer")

def require_role(min_role: str):
    def decorator(fn):
        @wraps(fn)
        def wrapper(*args, **kwargs):
            role = current_role()
            if ROLE_ORDER.get(role, 0) < ROLE_ORDER.get(min_role, 0):
                return jsonify({"error": "forbidden", "required": min_role, "role": role}), 403
            return fn(*args, **kwargs)
        return wrapper
    return decorator

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "GET":
        return render_template_string(LOGIN_HTML, error=None)
    username = (request.form.get("username") or "").strip()
    password = request.form.get("password") or ""
    u = USERS.get(username)
    if not u or u["password"] != password:
        return render_template_string(LOGIN_HTML, error="Invalid username or password.")
    session["user"] = username
    session["role"] = u["role"]
    return redirect(url_for("index"))

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

# ------------------------------------------------------------------------------
# LOGO / ASSETS
# ------------------------------------------------------------------------------

@app.route("/assets/logo")
def gds_logo():
    if os.path.exists(LOGO_FILE):
        mime = "image/jpeg" if LOGO_FILE.lower().endswith((".jpg", ".jpeg")) else "image/png"
        return send_file(LOGO_FILE, mimetype=mime)
    png = base64.b64decode(FALLBACK_LOGO_PNG_BASE64)
    return Response(png, mimetype="image/png")

@app.route("/assets/no-signal.png")
def no_signal():
    png = base64.b64decode(
        "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mP8/x8AAwMB/6Xw2m8AAAAASUVORK5CYII="
    )
    return Response(png, mimetype="image/png")

# ------------------------------------------------------------------------------
# CAMERA HELPERS (DEMO-SAFE)
# ------------------------------------------------------------------------------

def _demo_camera_frame(text="DEMO CAMERA (NO SIGNAL)"):
    """
    Returns JPEG bytes. This ensures <img src="/camera/stream"> always shows something.
    """
    img = Image.new("RGB", (960, 540), (10, 10, 12))
    draw = ImageDraw.Draw(img)

    # Border + header bar
    draw.rectangle((0, 0, 959, 539), outline=(204, 0, 0), width=6)
    draw.rectangle((0, 0, 959, 70), fill=(30, 0, 0))
    draw.text((18, 22), "GDS PTZ FEED", fill=(255, 255, 255))

    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    draw.text((18, 110), text, fill=(255, 200, 200))
    draw.text((18, 150), f"Timestamp: {ts}", fill=(220, 220, 220))

    with camera_lock:
        pan = camera_controls.get("pan", 0)
        tilt = camera_controls.get("tilt", 0)
        zoom = camera_controls.get("zoom", 1.0)
    draw.text((18, 200), f"PAN: {pan}   TILT: {tilt}   ZOOM: {zoom}", fill=(220, 220, 220))

    # Simple crosshair
    cx, cy = 480, 320
    draw.line((cx - 60, cy, cx + 60, cy), fill=(255, 31, 31), width=3)
    draw.line((cx, cy - 60, cx, cy + 60), fill=(255, 31, 31), width=3)

    out = BytesIO()
    img.save(out, format="JPEG", quality=85)
    return out.getvalue()

def check_camera_connection():
    global camera_status
    try:
        auth = (CAMERA_USERNAME, CAMERA_PASSWORD) if CAMERA_USERNAME else None

        # IMPORTANT: Android IP Webcam does NOT support HEAD
        response = requests.get(
            IP_CAMERA_URL,
            auth=auth,
            stream=True,
            timeout=CAMERA_CONNECTION_TIMEOUT
        )

        if response.status_code < 400:
            with camera_lock:
                camera_status["connected"] = True
                camera_status["mode"] = "ip"
                camera_status["error"] = None
                camera_status["last_check"] = datetime.now().isoformat()
            return True

        raise Exception(f"Status {response.status_code}")

    except Exception as e:
        if AUTO_FALLBACK_TO_DEMO:
            with camera_lock:
                camera_status["connected"] = False
                camera_status["mode"] = "demo"
                camera_status["error"] = str(e)
                camera_status["last_check"] = datetime.now().isoformat()
        return False


def init_camera():
    check_camera_connection()

def clamp(v, lo, hi):
    return max(lo, min(hi, v))

def map_heading_to_pan(heading_deg: float) -> int:
    h = float(heading_deg) % 360.0
    pan = int(round(h if h <= 180 else h - 360))
    return clamp(pan, -180, 180)

def ptz_sync_worker():
    global ptz_sync_enabled
    while True:
        try:
            with sync_lock:
                enabled = ptz_sync_enabled
            if enabled:
                with nav_lock:
                    cog = nav_current.get("cog", 0.0)
                desired_pan = map_heading_to_pan(cog)
                with camera_lock:
                    camera_controls["pan"] = desired_pan
            time.sleep(1)
        except Exception:
            time.sleep(1)

threading.Thread(target=ptz_sync_worker, daemon=True).start()

# ------------------------------------------------------------------------------
# SIMULATION THREADS (dummy ST6100-like)
# ------------------------------------------------------------------------------

def simulate_navigation():
    global sim_lat, sim_lon, sim_heading, sim_speed, nav_current
    while True:
        try:
            sim_speed += random.uniform(-0.3, 0.3)
            sim_speed = max(5, min(12, sim_speed))
            sim_heading = (sim_heading + random.uniform(-2, 2)) % 360

            distance = (sim_speed * 1.852 / 3600) / 111  # approx degrees per sec
            sim_lat += distance * math.cos(math.radians(sim_heading))
            sim_lon += distance * math.sin(math.radians(sim_heading))
            sim_lat = max(1.5, min(6.0, sim_lat))
            sim_lon = max(100.0, min(104.5, sim_lon))

            cog = (sim_heading + random.uniform(-2, 2)) % 360
            voltage = random.randint(1190, 1240)

            now = datetime.now()
            date_str = now.strftime("%d/%m/%Y")
            time_str = now.strftime("%H:%M:%S")

            raw_string = f"{date_str},{time_str},{sim_lat:.6f},{sim_lon:.6f},{sim_speed:.1f},{cog:.0f},{voltage},0,{sim_heading:.0f}"

            with nav_lock:
                if nav_current.get("date"):
                    nav_history.insert(0, nav_current.copy())
                    if len(nav_history) > MAX_NAV_HISTORY:
                        nav_history.pop()

                nav_current = {
                    "date": date_str,
                    "time": time_str,
                    "latitude": round(sim_lat, 6),
                    "longitude": round(sim_lon, 6),
                    "speed": round(sim_speed, 1),
                    "heading": round(sim_heading, 0),
                    "cog": round(cog, 0),
                    "voltage": voltage,
                    "panic": 0,
                    "ext_heading": round(sim_heading, 0),
                    "raw_string": raw_string
                }


            # Persist NAV log
            append_nav_log(nav_current)

            time.sleep(1)
        except Exception:
            time.sleep(1)

def simulate_weather():
    global weather_current
    base_temp, base_humid, base_press, base_wind, base_dir = 28.0, 75.0, 1013.25, 5.0, 180.0
    while True:
        try:
            base_wind = max(0, base_wind + random.uniform(-1, 1))
            base_dir = (base_dir + random.uniform(-5, 5)) % 360
            base_temp = max(20, min(40, base_temp + random.uniform(-0.2, 0.2)))
            base_humid = max(40, min(100, base_humid + random.uniform(-0.5, 0.5)))
            base_press = max(990, min(1030, base_press + random.uniform(-0.1, 0.1)))

            with weather_lock:
                weather_current = {
                    "wind_speed": round(base_wind, 1),
                    "wind_direction": round(base_dir, 0),
                    "temperature": round(base_temp, 1),
                    "humidity": round(base_humid, 1),
                    "pressure": round(base_press, 2),
                    "illumination": 50000,
                    "timestamp": datetime.now().strftime("%H:%M:%S")
                }
            time.sleep(2)
        except Exception:
            time.sleep(2)

#threading.Thread(target=simulate_navigation, daemon=True).start()
#threading.Thread(target=simulate_weather, daemon=True).start()
#threading.Thread(target=firebase_weather_listener, daemon=True).start()
#threading.Thread(target=firebase_marinelite_listener, daemon=True).start()



# ------------------------------------------------------------------------------
# ROUTES
# ------------------------------------------------------------------------------
@app.route("/api/connectivity/status")
def api_connectivity_status():
    """
    Stub API.
    Later this will read Wi-Fi + SIM + Ethernet status from RPi.
    """
    return jsonify({
        "wifi": {
            "available": True,
            "connected": False,
            "ssid": None,
            "signal": None
        },
        "sim": {
            "available": True,
            "connected": False,
            "operator": None,
            "country": None,
            "signal": None
        },
        "internet": False
    })
@app.route("/api/sync/status")
def api_sync_status():
    """
    Stub API.
    Later this will reflect SQLite sync state.
    """
    return jsonify({
        "pending_records": 0,
        "last_sync": None,
        "sync_state": "idle"
    })
@app.route("/api/sync/start", methods=["POST"])
def api_sync_start():
    """
    Stub API.
    Later this will trigger background sync.
    """
    return jsonify({
        "status": "accepted",
        "message": "Sync will start when connectivity is available"
    })

@app.route("/")
def index():
    if not current_user():
        return redirect(url_for("login"))
    with camera_lock:
        mode = camera_status["mode"]
    role = current_role()
    user = current_user()
    return render_template_string(HTML_DASHBOARD, camera_mode=mode, role=role, user=user, weasy=WEASYPRINT_AVAILABLE)

@app.route("/api/me")
def api_me():
    return jsonify({"user": current_user(), "role": current_role(), "weasyprint": WEASYPRINT_AVAILABLE})

@app.route("/camera/capture", methods=["POST"])
@require_role("Operator")
def capture_image():
    data = request.json or {}
    report_type = data.get("report_type", "vdr")

    # OFFLINE MODE: Firebase disabled here
    upload_firebase = False

    try:
        jpg_quality = int(data.get("jpg_quality", 75))
    except Exception:
        jpg_quality = 75

    try:
        max_width = int(data.get("max_width", 1280))
    except Exception:
        max_width = 1280

    jpg_quality = max(35, min(95, jpg_quality))
    max_width = max(320, min(1920, max_width))

    # ---- Capture image ----
    content = _fetch_camera_snapshot_bytes()
    if not content:
        content = _demo_camera_frame("CAPTURED (FALLBACK)")

    content = _compress_jpeg(content, max_width=max_width, quality=jpg_quality)
    base64_data = base64.b64encode(content).decode()

    # ---- SIZE GUARD (INSIDE FUNCTION) ----
    size_kb = len(base64_data) * 0.75 / 1024
    if size_kb > 80:
        return jsonify({
            "status": "error",
            "message": f"Image too large for DB-only mode ({size_kb:.1f} KB)"
        }), 400

    item = {
        "id": uuid.uuid4().hex,
        "data": base64_data,
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "type": "jpeg"
    }

    # ---- Store locally for UI ----
    with image_lock:
        key = "vjr_images" if report_type == "vjr" else "vdr_images"
        captured_images[key].insert(0, item)
        if len(captured_images[key]) > MAX_CAPTURE_IMAGES:
            captured_images[key].pop()

    return jsonify({
        "status": "success",
        "report_type": report_type,
        "size_kb": round(size_kb, 2)
    })


@app.route("/camera/control", methods=["POST"])
@require_role("Operator")
def control_camera():
    """
    Always returns OK so buttons feel functional even without hardware.
    """
    data = request.json or {}
    action = data.get("action", "")
    value = data.get("value", 0)

    with camera_lock:
        if action == "pan_left":
            camera_controls["pan"] = max(-180, camera_controls["pan"] - 15)
        elif action == "pan_right":
            camera_controls["pan"] = min(180, camera_controls["pan"] + 15)
        elif action == "tilt_up":
            camera_controls["tilt"] = min(90, camera_controls["tilt"] + 15)
        elif action == "tilt_down":
            camera_controls["tilt"] = max(-90, camera_controls["tilt"] - 15)
        elif action == "zoom_in":
            camera_controls["zoom"] = min(10, camera_controls["zoom"] + 0.5)
        elif action == "zoom_out":
            camera_controls["zoom"] = max(1, camera_controls["zoom"] - 0.5)
        elif action == "zoom_set":
            try:
                camera_controls["zoom"] = max(1, min(10, float(value)))
            except Exception:
                pass
        elif action == "led_toggle":
            camera_controls["led_enabled"] = not camera_controls["led_enabled"]
        elif action == "led_brightness":
            try:
                camera_controls["led_brightness"] = max(0, min(100, int(value)))
            except Exception:
                pass
        elif action == "night_vision_toggle":
            camera_controls["night_vision"] = not camera_controls["night_vision"]

        control_copy = camera_controls.copy()

    return jsonify({"status": "ok", "message": f"Action executed: {action}", "controls": control_copy})

@app.route("/camera/status")
def get_camera_status():
    with camera_lock:
        st = camera_status.copy()
        st["controls"] = camera_controls.copy()
    with sync_lock:
        st["ptz_sync_enabled"] = ptz_sync_enabled
    return jsonify(st)

@app.route("/camera/reconnect", methods=["POST"])
@require_role("Operator")
def reconnect_camera():
    result = check_camera_connection()
    with camera_lock:
        status_copy = camera_status.copy()
    return jsonify({"reconnected": result, "status": status_copy})

@app.route("/camera/sync_cog", methods=["POST"])
@require_role("Operator")
def camera_sync_cog():
    global ptz_sync_enabled
    enabled = bool((request.json or {}).get("enabled", False))
    with sync_lock:
        ptz_sync_enabled = enabled
    return jsonify({"status": "ok", "enabled": ptz_sync_enabled})


def _extract_first_jpeg_from_mjpeg(resp, max_bytes=3_000_000, max_seconds=1.5):
    """Extract the first JPEG frame from an MJPEG stream response.

    Hard-stops after max_seconds to avoid UI feeling 'stuck' when the IP camera
    stalls or buffers.
    """
    buf = bytearray()
    start = -1
    t0 = time.time()
    for chunk in resp.iter_content(chunk_size=4096):
        if (time.time() - t0) > max_seconds:
            break
        if not chunk:
            continue
        buf.extend(chunk)
        if start < 0:
            s = buf.find(b"\xff\xd8")
            if s >= 0:
                start = s
        if start >= 0:
            e = buf.find(b"\xff\xd9", start)
            if e >= 0:
                return bytes(buf[start:e+2])
        if len(buf) > max_bytes:
            break
    return None


def _compress_jpeg(jpg_bytes: bytes, max_width=1280, quality=75):
    """Downscale and recompress JPEG for faster UI + smaller base64."""
    try:
        im = Image.open(BytesIO(jpg_bytes))
        im = im.convert("RGB")
        w, h = im.size
        if w > max_width:
            nh = int(h * (max_width / float(w)))
            im = im.resize((max_width, nh))
        out = BytesIO()
        im.save(out, format="JPEG", quality=quality, optimize=True)
        return out.getvalue()
    except Exception:
        return jpg_bytes

def _fetch_camera_snapshot_bytes():
    """Fast snapshot: get a single JPEG frame (works for MJPEG and JPEG endpoints)."""
    with camera_lock:
        mode = camera_status.get("mode", "demo")

    if mode == "demo":
        return _demo_camera_frame("CAPTURED (DEMO MODE)")

    try:
        auth = (CAMERA_USERNAME, CAMERA_PASSWORD) if CAMERA_USERNAME else None

        # Use streaming so we can cut after the first JPEG instead of downloading the full MJPEG feed
        resp = requests.get(
            IP_CAMERA_URL,
            auth=auth,
            stream=True,
            timeout=(CAMERA_CONNECTION_TIMEOUT, 2),
            headers={"User-Agent": "GDS-VMS/1.0"},
        )
        if resp.status_code != 200:
            return None

        ctype = (resp.headers.get("Content-Type") or "").lower()
        if "multipart" in ctype or "mjpeg" in ctype or "x-mixed-replace" in ctype:
            frame = _extract_first_jpeg_from_mjpeg(resp)
            if frame:
                return _compress_jpeg(frame)
            return None

        # Non-multipart: assume direct JPEG snapshot endpoint
        data = resp.content
        if data:
            return _compress_jpeg(data)
    except Exception:
        return None

    return None







@app.route("/camera/captured_images")
def get_captured_images_count():
    with image_lock:
        return jsonify({"vjr_images": len(captured_images["vjr_images"]), "vdr_images": len(captured_images["vdr_images"])})

@app.route("/camera/captured_images_full")
def get_captured_images_full():
    """
    Used by UI galleries (VJR & VDR) to display thumbnails.
    """
    with image_lock:
        return jsonify({
            "vjr": captured_images["vjr_images"],
            "vdr": captured_images["vdr_images"]
        })

@app.route("/camera/clear_captures", methods=["POST"])
@require_role("Operator")
def clear_captures():
    data = request.json or {}
    report_type = (data.get("report_type") or "").lower()
    if report_type not in ("vjr", "vdr"):
        return jsonify({"error": "Invalid report_type"}), 400

    key = "vjr_images" if report_type == "vjr" else "vdr_images"
    with image_lock:
        captured_images[key].clear()

    return jsonify({"status": "success", "cleared": report_type})

@app.route("/camera/delete_capture", methods=["POST"])
@require_role("Operator")
def delete_capture():
    """Delete one captured image.

    Old behavior used an array index (fragile when UI re-renders).
    New behavior supports stable deletion by capture_id.
    """
    data = request.json or {}
    report_type = (data.get("report_type") or "").lower()

    capture_id = data.get("capture_id") or data.get("id")
    idx = data.get("index", None)

    if report_type not in ("vjr", "vdr"):
        return jsonify({"error": "Invalid report_type"}), 400

    key = "vjr_images" if report_type == "vjr" else "vdr_images"

    with image_lock:
        arr = captured_images.get(key, [])

        # Preferred: delete by capture_id (stable)
        if capture_id:
            for i, it in enumerate(arr):
                if str(it.get("id")) == str(capture_id):
                    arr.pop(i)
                    return jsonify({"status": "success", "deleted_id": capture_id, "report_type": report_type})
            return jsonify({"error": "capture_id not found"}), 404

        # Backward-compatible: delete by index
        if idx is None:
            return jsonify({"error": "capture_id or index required"}), 400
        try:
            idx = int(idx)
        except Exception:
            return jsonify({"error": "index must be int"}), 400

        if idx < 0 or idx >= len(arr):
            return jsonify({"error": "index out of range"}), 400
        arr.pop(idx)

    return jsonify({"status": "success", "deleted_index": idx, "report_type": report_type})



@app.route("/nav_data")
def get_navigation_data():
    if LOCAL_DB_ENABLED:
        return jsonify(db_get_latest_marinelite())

    # fallback (old behavior)
    with nav_lock:
        return jsonify({
            "latitude": nav_current.get("latitude"),
            "longitude": nav_current.get("longitude"),
            "heading": nav_current.get("heading"),
            "timestamp": nav_current.get("timestamp")
        })



@app.route("/weather_data")
def get_weather_data():
    if LOCAL_DB_ENABLED:
        return jsonify(db_get_latest_weather())

    with weather_lock:
        return jsonify(weather_current)




@app.route("/export_nav_csv")
def export_nav_csv():
    """Browser download of the persistent navigation log (CSV)."""
    if not current_user():
        return redirect(url_for("login"))
    _ensure_nav_log_header()
    try:
        return send_file(
            NAV_LOG_FILE,
            as_attachment=True,
            download_name=f"NAV_Log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            mimetype="text/csv",
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ------------------------------------------------------------------------------
# VDR RECORDS + EXPORTS
# ------------------------------------------------------------------------------

@app.route("/vdr_records")
def get_vdr_records():
    with vdr_lock:
        return jsonify(vdr_records)

@app.route("/save_vdr", methods=["POST"])
@require_role("Operator")
def save_vdr_record():
    data = request.json or {}
    with vdr_lock:
        record = {k: data.get(k, "") for k in [
            "date","vessel","imo","mmsi","callsign","client","location","country",
            "activity","weather","sea_state","crew_count","deck_crew","engine_crew",
            "officers","fuel_consumption","main_engine_hours","generator_hours",
            "distance_traveled","incidents","maintenance_work","remarks"
        ]}
        record["id"] = (vdr_records[0]["id"] + 1) if vdr_records else 1
        record["timestamp"] = datetime.now().isoformat()
        vdr_records.insert(0, record)
    return jsonify({"status": "ok", "total": len(vdr_records)})

@app.route("/clear_vdr", methods=["POST"])
@require_role("Captain")
def clear_vdr_records():
    with vdr_lock:
        vdr_records.clear()
    return jsonify({"status": "ok"})

@app.route("/export_vdr_excel", methods=["POST"])
@require_role("Operator")
def export_vdr_excel():
    if not openpyxl:
        return jsonify({"error": "openpyxl not installed"}), 400

    records = (request.json or {}).get("records", [])
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Daily Report"

    headers = [
        "ID","Date","Vessel","IMO","MMSI","Call Sign","Client","Location","Country",
        "Activity","Weather","Sea State","Crew","Deck Crew","Engine Crew","Officers",
        "Fuel","M/E Hours","Gen Hours","Distance","Incidents","Maintenance","Remarks"
    ]
    ws.append(headers)

    for cell in ws[1]:
        cell.fill = PatternFill(start_color="CC0000", end_color="CC0000", fill_type="solid")
        cell.font = XLFont(color="FFFFFF", bold=True)

    for r in records:
        ws.append([r.get(k, "") for k in [
            "id","date","vessel","imo","mmsi","callsign","client","location","country",
            "activity","weather","sea_state","crew_count","deck_crew","engine_crew","officers",
            "fuel_consumption","main_engine_hours","generator_hours","distance_traveled",
            "incidents","maintenance_work","remarks"
        ]])

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return Response(
        out.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=VDR_Report.xlsx"}
    )

@app.route("/export_vdr_csv", methods=["POST"])
@require_role("Operator")
def export_vdr_csv():
    records = (request.json or {}).get("records", [])
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "ID","Date","Vessel","IMO","MMSI","Call Sign","Client","Location","Country",
        "Activity","Weather","Sea State","Crew","Deck Crew","Engine Crew","Officers",
        "Fuel","M/E Hours","Gen Hours","Distance","Incidents","Maintenance","Remarks"
    ])
    for r in records:
        writer.writerow([r.get(k, "") for k in [
            "id","date","vessel","imo","mmsi","callsign","client","location","country",
            "activity","weather","sea_state","crew_count","deck_crew","engine_crew","officers",
            "fuel_consumption","main_engine_hours","generator_hours","distance_traveled",
            "incidents","maintenance_work","remarks"
        ]])
    return Response(output.getvalue(), mimetype="text/csv",
                    headers={"Content-Disposition": "attachment; filename=VDR_Report.csv"})

def _logo_b64():
    if os.path.exists(LOGO_FILE):
        return base64.b64encode(open(LOGO_FILE, "rb").read()).decode()
    return FALLBACK_LOGO_PNG_BASE64

@app.route("/export_vdr_pdf", methods=["POST"])
@require_role("Operator")
def export_vdr_pdf():
    records = (request.json or {}).get("records", [])
    with image_lock:
        images = captured_images["vdr_images"][:18]

    # TRY REPORTLAB FIRST (Windows-friendly, pure Python)
    if REPORTLAB_AVAILABLE:
        try:
            buffer = BytesIO()
            doc = SimpleDocTemplate(
                buffer,
                pagesize=letter,
                rightMargin=0.5*inch,
                leftMargin=0.5*inch,
                topMargin=0.75*inch,
                bottomMargin=0.75*inch
            )
            
            story = []
            styles = getSampleStyleSheet()
            
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                textColor=colors.HexColor('#cc0000'),
                spaceAfter=10
            )
            
            story.append(Paragraph(
                "<b><font color='#cc0000'>Vessel Daily Report (VDR)</font></b><br/>"
                f"<font size=9>Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</font>",
                title_style
            ))
            story.append(Spacer(1, 0.2*inch))
            
            # Records Table
            if records:
                table_data = [['Date', 'Vessel', 'Activity', 'Location', 'Weather', 'Remarks']]
                for r in records[:50]:
                    table_data.append([
                        str(r.get('date', '')),
                        str(r.get('vessel', '')),
                        str(r.get('activity', '')),
                        str(r.get('location', '')),
                        str(r.get('weather', '')),
                        str(r.get('remarks', ''))[:40]
                    ])
                
                table = Table(table_data, colWidths=[0.8*inch, 0.9*inch, 0.8*inch, 0.9*inch, 0.9*inch, 1.2*inch])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a1a1a')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 9),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('FONTSIZE', (0, 1), (-1, -1), 7),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
                ]))
                story.append(table)
            
            story.append(Spacer(1, 0.2*inch))
            
            # Images
            if images:
                img_data = []
                for img in images[:18]:
                    try:
                        img_b64 = img.get('data', '')
                        if img_b64:
                            img_bytes = base64.b64decode(img_b64)
                            pil_img = Image.open(BytesIO(img_bytes))
                            pil_img.thumbnail((1.2*inch, 1.2*inch))
                            img_path = f'/tmp/vdr_img_{int(time.time()*1000)}.png'
                            pil_img.save(img_path)
                            img_data.append(RLImage(img_path, width=1.1*inch, height=1.1*inch))
                    except Exception:
                        pass
                
                if img_data:
                    img_grid = []
                    for i in range(0, len(img_data), 2):
                        if i+1 < len(img_data):
                            img_grid.append([img_data[i], img_data[i+1]])
                        else:
                            img_grid.append([img_data[i], ''])
                    
                    img_table = Table(img_grid)
                    img_table.setStyle(TableStyle([
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('PADDING', (5, 5), (-1, -1), 5),
                    ]))
                    story.append(img_table)
            
            story.append(Spacer(1, 0.1*inch))
            story.append(Paragraph(f"<font size=7>GDS Maritime System  {datetime.now().year}</font>", styles['Normal']))
            
            doc.build(story)
            pdf_bytes = buffer.getvalue()
            buffer.close()
            
            return Response(
                pdf_bytes,
                mimetype="application/pdf",
                headers={"Content-Disposition": "attachment; filename=VDR_Report.pdf"}
            )
        
        except Exception as e:
            print(f"ReportLab PDF generation failed: {e}, falling back to WeasyPrint")
    
    # FALLBACK TO WEASYPRINT
    if not WEASYPRINT_AVAILABLE:
        return jsonify({"error": "PDF generation failed. Install: pip install reportlab pillow"}), 400
    
    rows = ""
    for r in records[:50]:
        rows += f"""
        <tr>
          <td>{r.get('id','')}</td>
          <td>{r.get('date','')}</td>
          <td>{r.get('vessel','')}</td>
          <td>{r.get('activity','')}</td>
          <td>{r.get('location','')}</td>
          <td>{r.get('weather','')}</td>
          <td>{r.get('remarks','')}</td>
        </tr>
        """

    gallery = ""
    if images:
        gallery += "<h2>Captured Images (VDR)</h2><div class='grid'>"
        for img in images:
            gallery += f"""
            <div class="card">
              <img src="data:image/jpeg;base64,{img['data']}" />
              <div class="cap">{img['timestamp']}</div>
            </div>
            """
        gallery += "</div>"

    html = f"""
    <!doctype html>
    <html>
    <head>
      <meta charset="utf-8">
      <style>
        body {{ font-family: Arial, sans-serif; padding: 24px; color: #111; }}
        .header {{ display:flex; align-items:center; gap:14px; border-bottom: 4px solid #cc0000; padding-bottom: 14px; margin-bottom: 14px; }}
        .header img {{ height: 46px; }}
        h1 {{ margin:0; color:#cc0000; font-size: 22px; }}
        .sub {{ color:#555; font-size: 12px; margin-top: 4px; }}
        table {{ width:100%; border-collapse:collapse; font-size: 11px; margin-top: 12px; }}
        th, td {{ border:1px solid #ddd; padding: 8px; vertical-align: top; }}
        th {{ background:#111; color:#fff; }}
        tr:nth-child(even) {{ background:#fafafa; }}
        h2 {{ color:#cc0000; margin: 18px 0 10px; font-size: 14px; border-bottom: 2px solid #cc0000; padding-bottom: 6px; }}
        .grid {{ display:grid; grid-template-columns: repeat(2, 1fr); gap: 12px; }}
        .card {{ border:1px solid #ddd; border-radius: 8px; overflow:hidden; }}
        .card img {{ width:100%; height:auto; display:block; }}
        .cap {{ padding: 8px; font-size: 10px; color:#555; }}
        .foot {{ margin-top: 16px; border-top: 1px solid #ddd; padding-top: 10px; font-size: 10px; color:#666; text-align:center; }}
      </style>
    </head>
    <body>
      <div class="header">
        <img src="data:image/png;base64,{_logo_b64()}" />
        <div>
          <h1>Vessel Daily Report (VDR)</h1>
          <div class="sub">Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</div>
        </div>
      </div>

      <h2>Records Summary</h2>
      <table>
        <thead>
          <tr>
            <th>ID</th><th>Date</th><th>Vessel</th><th>Activity</th><th>Location</th><th>Weather</th><th>Remarks</th>
          </tr>
        </thead>
        <tbody>{rows if rows else '<tr><td colspan="7">No records</td></tr>'}</tbody>
      </table>

      {gallery}

      <div class="foot">Generated by GDS - Maritime Management System | &copy; {datetime.now().year}</div>
    </body>
    </html>
    """

    pdf = HTML(string=html).write_pdf()
    return Response(pdf, mimetype="application/pdf",
                    headers={"Content-Disposition": "attachment; filename=VDR_Report.pdf"})

# ------------------------------------------------------------------------------
# VJR PDF EXPORT (NEW)
# ------------------------------------------------------------------------------

@app.route("/export_vjr_pdf", methods=["POST"])
@require_role("Operator")
def export_vjr_pdf():
    vjr = (request.json or {}).get("vjr", {})
    navlog = (request.json or {}).get("nav", [])
    with image_lock:
        images = captured_images["vjr_images"][:18]

    # TRY REPORTLAB FIRST (Windows-friendly, pure Python)
    if REPORTLAB_AVAILABLE:
        try:
            buffer = BytesIO()
            doc = SimpleDocTemplate(
                buffer,
                pagesize=letter,
                rightMargin=0.5*inch,
                leftMargin=0.5*inch,
                topMargin=0.75*inch,
                bottomMargin=0.75*inch
            )
            
            story = []
            styles = getSampleStyleSheet()
            
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                textColor=colors.HexColor('#cc0000'),
                spaceAfter=10
            )
            
            story.append(Paragraph(
                "<b><font color='#cc0000'>Vessel Journey Report (VJR)</font></b><br/>"
                f"<font size=9>Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</font>",
                title_style
            ))
            story.append(Spacer(1, 0.2*inch))
            
            # Journey Details
            details_data = [
                ['Report Date:', vjr.get('date', '')],
                ['Vessel:', vjr.get('vessel', '')],
                ['IMO:', vjr.get('imo', '')],
                ['Master:', vjr.get('master', '')],
                ['Departure:', vjr.get('departure', '')],
                ['Arrival:', vjr.get('arrival', '')],
            ]
            
            details_table = Table(details_data, colWidths=[1.2*inch, 4*inch])
            details_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f0f0f0')),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('PADDING', (8, 6), (-1, -1), 6),
            ]))
            story.append(details_table)
            story.append(Spacer(1, 0.2*inch))
            
            # Images
            if images:
                img_data = []
                for img in images[:16]:
                    try:
                        img_b64 = img.get('data', '')
                        if img_b64:
                            img_bytes = base64.b64decode(img_b64)
                            pil_img = Image.open(BytesIO(img_bytes))
                            pil_img.thumbnail((1.2*inch, 1.2*inch))
                            img_path = f'/tmp/vjr_img_{int(time.time()*1000)}.png'
                            pil_img.save(img_path)
                            img_data.append(RLImage(img_path, width=1.1*inch, height=1.1*inch))
                    except Exception:
                        pass
                
                if img_data:
                    img_grid = []
                    for i in range(0, len(img_data), 2):
                        if i+1 < len(img_data):
                            img_grid.append([img_data[i], img_data[i+1]])
                        else:
                            img_grid.append([img_data[i], ''])
                    
                    img_table = Table(img_grid)
                    img_table.setStyle(TableStyle([
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('PADDING', (5, 5), (-1, -1), 5),
                    ]))
                    story.append(img_table)
            
            story.append(Spacer(1, 0.1*inch))
            story.append(Paragraph(f"<font size=7>GDS Maritime System {datetime.now().year}</font>", styles['Normal']))
            
            doc.build(story)
            pdf_bytes = buffer.getvalue()
            buffer.close()
            
            return Response(
                pdf_bytes,
                mimetype="application/pdf",
                headers={"Content-Disposition": "attachment; filename=VJR_Report.pdf"}
            )
        
        except Exception as e:
            print(f"ReportLab PDF generation failed: {e}, falling back to WeasyPrint")
    
    # FALLBACK TO WEASYPRINT
    if not WEASYPRINT_AVAILABLE:
        return jsonify({"error": "PDF generation failed. Install: pip install reportlab pillow"}), 400

    nav_rows = ""
    for i, n in enumerate(navlog[:40], start=1):
        nav_rows += f"""
        <tr>
          <td>{i}</td>
          <td>{n.get('date','')}</td>
          <td>{n.get('time','')}</td>
          <td>{n.get('latitude','')}</td>
          <td>{n.get('longitude','')}</td>
          <td>{n.get('speed','')}</td>
          <td>{n.get('cog','')}</td>
        </tr>
        """

    gallery = ""
    if images:
        gallery += "<h2>Captured Images (VJR)</h2><div class='grid'>"
        for img in images:
            gallery += f"""
            <div class="card">
              <img src="data:image/jpeg;base64,{img['data']}" />
              <div class="cap">{img['timestamp']}</div>
            </div>
            """
        gallery += "</div>"

    html = f"""
    <!doctype html>
    <html>
    <head>
      <meta charset="utf-8">
      <style>
        body {{ font-family: Arial, sans-serif; padding: 24px; color: #111; }}
        .header {{ display:flex; align-items:center; gap:14px; border-bottom: 4px solid #cc0000; padding-bottom: 14px; margin-bottom: 14px; }}
        .header img {{ height: 46px; }}
        h1 {{ margin:0; color:#cc0000; font-size: 22px; }}
        .sub {{ color:#555; font-size: 12px; margin-top: 4px; }}
        .meta {{ display:grid; grid-template-columns: 1fr 1fr; gap: 10px; margin: 12px 0 6px; }}
        .box {{ border:1px solid #ddd; border-radius: 8px; padding: 10px; font-size: 12px; }}
        .k {{ color:#666; font-size:11px; }}
        .v {{ font-weight:700; }}
        table {{ width:100%; border-collapse:collapse; font-size: 11px; margin-top: 12px; }}
        th, td {{ border:1px solid #ddd; padding: 8px; vertical-align: top; }}
        th {{ background:#111; color:#fff; }}
        tr:nth-child(even) {{ background:#fafafa; }}
        h2 {{ color:#cc0000; margin: 18px 0 10px; font-size: 14px; border-bottom: 2px solid #cc0000; padding-bottom: 6px; }}
        .grid {{ display:grid; grid-template-columns: repeat(2, 1fr); gap: 12px; }}
        .card {{ border:1px solid #ddd; border-radius: 8px; overflow:hidden; }}
        .card img {{ width:100%; height:auto; display:block; }}
        .cap {{ padding: 8px; font-size: 10px; color:#555; }}
        .foot {{ margin-top: 16px; border-top: 1px solid #ddd; padding-top: 10px; font-size: 10px; color:#666; text-align:center; }}
      </style>
    </head>
    <body>
      <div class="header">
        <img src="data:image/png;base64,{_logo_b64()}" />
        <div>
          <h1>Vessel Journey Report (VJR)</h1>
          <div class="sub">Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</div>
        </div>
      </div>

      <div class="meta">
        <div class="box"><div class="k">Report Date</div><div class="v">{vjr.get('date','')}</div></div>
        <div class="box"><div class="k">Vessel</div><div class="v">{vjr.get('vessel','')}</div></div>
        <div class="box"><div class="k">IMO</div><div class="v">{vjr.get('imo','')}</div></div>
        <div class="box"><div class="k">Master</div><div class="v">{vjr.get('master','')}</div></div>
        <div class="box"><div class="k">Departure</div><div class="v">{vjr.get('departure','')}</div></div>
        <div class="box"><div class="k">Arrival</div><div class="v">{vjr.get('arrival','')}</div></div>
      </div>

      <h2>Navigation Log</h2>
      <table>
        <thead>
          <tr>
            <th>#</th><th>Date</th><th>Time</th><th>Lat</th><th>Lon</th><th>Speed</th><th>COG</th>
          </tr>
        </thead>
        <tbody>{nav_rows if nav_rows else '<tr><td colspan="7">No navigation samples</td></tr>'}</tbody>
      </table>

      {gallery}

      <div class="foot">Generated by GDS - Maritime Management System | &copy; {datetime.now().year}</div>
    </body>
    </html>
    """

    pdf = HTML(string=html).write_pdf()
    return Response(pdf, mimetype="application/pdf",
                    headers={"Content-Disposition": "attachment; filename=VJR_Report.pdf"})

# ------------------------------------------------------------------------------
# UI
# ------------------------------------------------------------------------------

LOGIN_HTML = r"""
<!doctype html>
<html>
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>GDS Login</title>
  <style>
    :root{
      --bg:#070709; --panel:#0f0f13; --border:#262631;
      --red:#cc0000; --red2:#ff1f1f; --text:#ffffff; --muted:#a7a7b3;
    }
    body{
      background:
        radial-gradient(1200px 600px at 10% 10%, rgba(204,0,0,.14), transparent 60%),
        radial-gradient(900px 500px at 90% 20%, rgba(255,31,31,.10), transparent 55%),
        var(--bg);
      color:var(--text);font-family:Segoe UI,Arial,sans-serif;
      display:flex;min-height:100vh;align-items:center;justify-content:center;margin:0;
    }
    .card{
      width:380px;background:linear-gradient(180deg, rgba(255,255,255,.04), transparent 60%), var(--panel);
      border:1px solid var(--border);border-radius:14px;box-shadow:0 18px 50px rgba(0,0,0,.55);padding:22px;
    }
    .head{display:flex;align-items:center;gap:12px;margin-bottom:14px;}
    .head img{height:40px;filter: drop-shadow(0 10px 24px rgba(0,0,0,.5));}
    h1{font-size:15px;margin:0;letter-spacing:.3px;}
    .sub{font-size:12px;color:var(--muted);margin-top:3px;}
    label{font-size:10px;color:var(--muted);text-transform:uppercase;letter-spacing:.8px;font-weight:800;display:block;margin-top:12px;margin-bottom:6px;}
    input{width:100%;padding:11px 10px;background:rgba(0,0,0,.45);border:1px solid var(--border);border-radius:10px;color:var(--text);}
    input:focus{outline:none;border-color:rgba(255,31,31,.65);box-shadow:0 0 0 3px rgba(255,31,31,.12);}
    button{width:100%;margin-top:14px;padding:11px;border:0;border-radius:10px;background:linear-gradient(180deg, var(--red2), var(--red));
           color:#fff;font-weight:900;text-transform:uppercase;letter-spacing:.8px;cursor:pointer;}
    .err{margin-top:10px;background:rgba(255,31,31,.10);border:1px solid rgba(255,31,31,.6);color:#ff8a8a;padding:10px;border-radius:10px;font-size:12px;}
    .hint{margin-top:10px;font-size:11px;color:var(--muted);line-height:1.5;}
    code{background:rgba(255,255,255,.06);padding:2px 6px;border-radius:8px;border:1px solid rgba(255,255,255,.08);}
  </style>
</head>
<body>
  <div class="card">
    <div class="head">
      <img src="/assets/logo" alt="GDS">
      <div>
        <h1>GDS Vessel Management System</h1>
        <div class="sub">Black/Red Operational Console</div>
      </div>
    </div>
    <form method="POST">
      <label>Username</label>
      <input name="username" placeholder="captain / operator / viewer" autocomplete="username" required>
      <label>Password</label>
      <input type="password" name="password" autocomplete="current-password" required>
      <button type="submit">Login</button>
      {% if error %}<div class="err">{{ error }}</div>{% endif %}
      <div class="hint">
        Demo users:<br>
        <code>captain / captain123</code><br>
        <code>operator / operator123</code><br>
        <code>viewer / viewer123</code>
      </div>
    </form>
  </div>
</body>
</html>
"""

# NOTE: This HTML is your previous professional UI with added:
# - Gallery panels in VJR & VDR
# - VJR PDF button
# - Live gallery loader in JS

HTML_DASHBOARD = r'''<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>GDS - Vessel Management</title>
  <link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css" />
  <script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
  <style>
  /* ===== NAVIGATION FEED ===== */

.nav-grid{
  display:grid;
  grid-template-columns:1fr 1fr;
  gap:14px;
}

.nav-card{
  background:rgba(255,255,255,0.03);
  border:1px solid rgba(255,255,255,0.08);
  border-radius:10px;
  padding:12px;
  text-align:center;
}

.nav-card.wide{
  grid-column:span 2;
}

.nav-label{
  font-size:12px;
  color:#aaa;
  letter-spacing:0.5px;
}

.nav-value{
  font-size:22px;
  font-weight:600;
  color:#00ff99;
  margin-top:6px;
}

.nav-value.big{
  font-size:38px;
  font-weight:800;
  letter-spacing:.6px;

}

.nav-unit{
  font-size:11px;
  color:#888;
}

.nav-status{
  grid-column:span 2;
  display:flex;
  justify-content:space-between;
  align-items:center;
  padding:10px 14px;
  border-radius:10px;
  background:rgba(255,0,0,0.08);
  border:1px solid rgba(255,0,0,0.3);
}
.sys-health{
  display:flex;
  gap:8px;
}

.health-pill{
  padding:4px 10px;
  border-radius:20px;
  font-size:11px;
  font-weight:700;
  letter-spacing:.4px;
}

.health-pill.ok{
  background:rgba(0,208,132,.15);
  color:#00d084;
  border:1px solid rgba(0,208,132,.4);
}

.health-pill.warn{
  background:rgba(255,204,0,.15);
  color:#ffcc00;
  border:1px solid rgba(255,204,0,.4);
}

.health-pill.bad{
  background:rgba(255,68,102,.15);
  color:#ff4466;
  border:1px solid rgba(255,68,102,.4);
}

.status-bad{
  color:#ff4d4d;
  font-weight:700;
}

.status-good{
  color:#00ff99;
  font-weight:700;
}

.nav-time{
  grid-column:span 2;
  text-align:right;
  font-size:12px;
  color:#aaa;
}

    :root{
      --bg:#070709;
      --panel:#0f0f13;
      --panel2:#13131a;
      --border:#262631;
      --text:#ffffff;
      --muted:#a7a7b3;
      --red:#cc0000;
      --red2:#ff1f1f;
      --good:#00d084;
      --warn:#ffcc00;
      --bad:#ff4466;
      --shadow: 0 18px 55px rgba(0,0,0,.55);
    }
    *{ margin:0; padding:0; box-sizing:border-box; }
    html, body{ height:100%; background: var(--bg); color:var(--text); font-family: Segoe UI, Arial, sans-serif; }
    body{
      display:flex; flex-direction:column;
      background:
        radial-gradient(1200px 700px at 10% 10%, rgba(204,0,0,.14), transparent 60%),
        radial-gradient(900px 600px at 90% 20%, rgba(255,31,31,.10), transparent 55%),
        var(--bg);
    }
    .header{
      display:flex; align-items:center; justify-content:space-between; gap:12px;
      padding: 14px 16px;
      border-bottom: 1px solid rgba(255,255,255,.06);
      background: linear-gradient(180deg, rgba(255,255,255,.04), transparent 70%);
      backdrop-filter: blur(8px);
      flex-shrink:0;
    }
    .brand{ display:flex; align-items:center; gap:12px; min-width: 280px; }
    .brand img{
      height: 44px; width:auto; display:block;
      filter: drop-shadow(0 10px 26px rgba(0,0,0,.55));
    }
    .brand h1{ font-size: 15px; letter-spacing:.3px; margin:0; }
    .brand p{ font-size: 11px; color: var(--muted); margin-top:2px; }
    .hdr-right{ display:flex; align-items:center; gap:10px; flex-wrap:wrap; justify-content:flex-end; }
    .pill{
      display:flex; align-items:center; gap:8px;
      padding: 8px 10px;
      border: 1px solid rgba(255,255,255,.08);
      border-radius: 999px;
      background: rgba(0,0,0,.35);
      font-size: 11px;
      color: var(--text);
    }
    .pill b{ color: var(--red2); }
    .clock{
      font-family: ui-monospace, SFMono-Regular, Menlo, Monaco, Consolas, "Courier New", monospace;
      font-size: 12px;
      padding: 8px 10px;
      border-radius: 10px;
      border: 1px solid rgba(255,31,31,.30);
      background: rgba(255,31,31,.08);
    }
    .pill{
  padding:4px 10px;
  border-radius:999px;
  font-size:12px;
  font-weight:600;
  border:1px solid rgba(255,255,255,.35);
  color:#fff;
  background:rgba(0,0,0,.25);
  white-space:nowrap;
}

    .link{ color: #ffd1d1; text-decoration: none; font-weight: 800; font-size: 12px; }
    .link:hover{ text-decoration: underline; }

    .tabs{
      display:flex; gap:8px; padding: 10px 14px;
      border-bottom: 1px solid rgba(255,255,255,.06);
      background: rgba(0,0,0,.18);
      flex-shrink:0;
      overflow-x:auto;
    }
    .tab-btn{
      border: 1px solid rgba(255,255,255,.08);
      background: rgba(0,0,0,.25);
      color: var(--muted);
      padding: 10px 14px;
      border-radius: 999px;
      cursor:pointer;
      font-weight: 900;
      font-size: 11px;
      text-transform: uppercase;
      letter-spacing: .8px;
      white-space: nowrap;
      transition: all .18s ease;
    }
    .tab-btn:hover{ border-color: rgba(255,31,31,.30); color: #fff; }
    .tab-btn.active{
      background: linear-gradient(180deg, rgba(255,31,31,.25), rgba(204,0,0,.18));
      border-color: rgba(255,31,31,.55);
      color:#fff;
    }

    .content-wrapper{ flex:1; overflow:hidden; display:flex; }
    .tab-content{ display:none; width:100%; height:100%; }
    .tab-content.active{ display:block; }

    .grid2{
      display:grid;
      grid-template-columns: 1fr 1fr;
      height:100%;
      gap: 12px;
      padding: 12px;
      overflow:hidden;
    }
    .grid3{
      display:grid;
      grid-template-columns: 1fr 1fr 1fr;
      height:100%;
      gap: 12px;
      padding: 12px;
      overflow:hidden;
    }

    .panel{
      background: linear-gradient(180deg, rgba(255,255,255,.04), transparent 60%), var(--panel);
      border: 1px solid rgba(255,255,255,.08);
      border-radius: 14px;
      box-shadow: var(--shadow);
      overflow:hidden;
      display:flex;
      flex-direction:column;
      min-height: 0;
    }
/* ===============================
   CONNECTIVITY STATUS COLORS
   =============================== */

.status-ok {
  color: #2ecc71;   /* green */
  font-weight: 600;
}

.status-warn {
  color: #f39c12;   /* orange */
  font-weight: 600;
}

.status-bad {
  color: #e74c3c;   /* red */
  font-weight: 600;
}

.status-muted {
  color: #7f8c8d;   /* gray */
}

    .panel-header{
      padding: 12px 14px;
      display:flex;
      justify-content:space-between;
      align-items:center;
      gap:10px;
      border-bottom: 1px solid rgba(255,255,255,.06);
      background: rgba(0,0,0,.22);
      font-weight: 950;
      font-size: 11px;
      text-transform: uppercase;
      letter-spacing: .9px;
    }
.panel-head .sub{
  font-size:11px;
  color:#aaa;
  margin-left:8px;
}

    .panel-header .hint{ font-size: 11px; color: var(--muted); font-weight: 800; text-transform:none; letter-spacing:0; }
    .panel-content{ padding: 14px; overflow:auto; min-height:0; }

    .raw{
      background: rgba(0,0,0,.45);
      border: 1px solid rgba(255,255,255,.08);
      border-radius: 12px;
      padding: 12px;
      color: #b9ffdb;
      font-family: ui-monospace, SFMono-Regular, Menlo, Monaco, Consolas, "Courier New", monospace;
      font-size: 11px;
      word-break: break-all;
      margin-bottom: 12px;
    }

    .data-grid{
      display:grid;
      grid-template-columns: 1fr 1fr;
      gap: 10px;
      margin-bottom: 10px;
    }
    .card{
      background: rgba(0,0,0,.30);
      border: 1px solid rgba(255,255,255,.08);
      border-radius: 12px;
      padding: 12px;
    }
    .label{
      color: var(--muted);
      font-size: 10px;
      text-transform: uppercase;
      letter-spacing: .9px;
      font-weight: 900;
      margin-bottom: 6px;
    }
    .value{
      font-family: ui-monospace, SFMono-Regular, Menlo, Monaco, Consolas, "Courier New", monospace;
      font-weight: 950;
      color: var(--good);
      font-size: 16px;
      letter-spacing: .2px;
    }
    .value.red{ color: #ff8a8a; }

    .form-group{ margin-bottom: 12px; }
    .form-label{
      display:block;
      margin-bottom: 6px;
      font-size: 10px;
      font-weight: 900;
      text-transform: uppercase;
      letter-spacing: .9px;
      color: var(--muted);
    }
    input, select, textarea{
      width:100%;
      padding: 10px 10px;
      border-radius: 12px;
      border: 1px solid rgba(255,255,255,.10);
      background: rgba(0,0,0,.45);
      color: #fff;
      font-size: 12px;
      outline:none;
      transition: box-shadow .15s ease, border-color .15s ease;
    }
    input:focus, select:focus, textarea:focus{
      border-color: rgba(255,31,31,.55);
      box-shadow: 0 0 0 3px rgba(255,31,31,.14);
    }
    textarea{ min-height: 64px; resize: vertical; }

    .btn{
      width:100%;
      padding: 11px 12px;
      border: 0;
      border-radius: 12px;
      cursor:pointer;
      font-weight: 950;
      font-size: 11px;
      letter-spacing: .9px;
      text-transform: uppercase;
      background: linear-gradient(180deg, var(--red2), var(--red));
      color: #fff;
      transition: transform .12s ease, box-shadow .12s ease, filter .12s ease;
      box-shadow: 0 12px 28px rgba(204,0,0,.22);
    }
    .btn:hover{ transform: translateY(-1px); filter: brightness(1.02); }
    .btn:active{ transform: translateY(0); }
    .btn.secondary{
      background: rgba(255,255,255,.06);
      border: 1px solid rgba(255,255,255,.10);
      box-shadow:none;
    }
    .btn.secondary:hover{ border-color: rgba(255,31,31,.35); }
    .btn.danger{
      background: linear-gradient(180deg, #ff4466, #c2183a);
      box-shadow: 0 12px 28px rgba(255,68,102,.18);
    }
    .btn.disabled{ opacity:.45; pointer-events:none; }

    .row2{ display:grid; grid-template-columns: 1fr 1fr; gap: 10px; }
    .mapbox{
      width:100%;
      height: 360px;
      border-radius: 14px;
      border: 1px solid rgba(255,255,255,.10);
      overflow:hidden;
      background:#000;
    }

    .cam-wrap{
      position:relative;
      width:100%;
      height: 260px;
      border-radius: 14px;
      border: 1px solid rgba(255,255,255,.10);
      overflow:hidden;
      background:#000;
    }
    .cam-wrap img{
      width:100%;
      height:100%;
      object-fit: cover;
      display:block;
    }
    .badge{
      position:absolute;
      left:12px;
      top:12px;
      padding: 7px 10px;
      border-radius: 999px;
      background: rgba(0,0,0,.55);
      border: 1px solid rgba(255,255,255,.12);
      font-size: 11px;
      font-weight: 900;
      color: #fff;
    }
    .badge b{ color: var(--red2); }

    .joy{
      width: 160px; height:160px; border-radius:50%;
      margin: 0 auto;
      background: rgba(0,0,0,.45);
      border: 1px solid rgba(255,255,255,.12);
      position:relative;
      touch-action:none;
      box-shadow: inset 0 0 0 1px rgba(255,31,31,.14);
    }
    .stick{
      width: 48px; height:48px; border-radius: 50%;
      position:absolute; left:50%; top:50%;
      transform: translate(-50%,-50%);
      background: radial-gradient(circle at 30% 25%, rgba(255,255,255,.22), rgba(255,255,255,.06)),
                  linear-gradient(180deg, rgba(255,31,31,.95), rgba(204,0,0,.85));
      box-shadow: 0 16px 28px rgba(0,0,0,.55);
    }
    /* ===== PTZ PANEL WRAPPER (NEW) ===== */

.ptz-panel{
  margin-top: 14px;
  padding-top: 12px;
  border-top: 1px solid rgba(255,255,255,.08);
  text-align: center;
}

.ptz-title{
  font-size: 11px;
  font-weight: 900;
  letter-spacing: .9px;
  text-transform: uppercase;
  color: var(--muted);
  margin-bottom: 10px;
}

.ptz-zoom{
  margin-top: 10px;
  display: grid;
  grid-template-columns: 1fr 1fr;
  gap: 10px;
}

    table{ width:100%; border-collapse:collapse; font-size: 11px; }
    th{
      text-align:left;
      padding: 10px;
      background: rgba(0,0,0,.35);
      border-bottom: 2px solid rgba(255,31,31,.35);
      color: #ffd1d1;
      font-weight: 950;
    }
    td{
      padding: 9px 10px;
      border-bottom: 1px solid rgba(255,255,255,.06);
      color: #e9e9ee;
    }
    tr:hover td{ background: rgba(255,31,31,.06); }

    .note{
      padding: 12px;
      background: rgba(255,31,31,.07);
      border: 1px solid rgba(255,31,31,.25);
      border-radius: 14px;
      color: #ffd1d1;
      font-size: 12px;
      line-height: 1.55;
    }

    .notification{
      position: fixed; right: 16px; bottom: 16px;
      padding: 12px 14px;
      border-radius: 14px;
      background: rgba(0,0,0,.65);
      border: 1px solid rgba(255,255,255,.10);
      box-shadow: var(--shadow);
      z-index: 9999;
      font-weight: 900;
      letter-spacing: .2px;
      max-width: 360px;
    }
    .notification.ok{ border-color: rgba(0,208,132,.35); color: #b9ffdb; }
    .notification.err{ border-color: rgba(255,68,102,.45); color: #ffd1d1; }
    .notification.warn{ border-color: rgba(255,204,0,.45); color: #fff2b3; }

    /* NEW: image galleries */
    .image-grid{
      display:grid;
      grid-template-columns: repeat(auto-fill, minmax(140px, 1fr));
      gap: 10px;
    }
    .thumb{
      position:relative;
      border: 1px solid rgba(255,255,255,.10);
      border-radius: 12px;
      overflow:hidden;
      background: rgba(0,0,0,.35);
    }
    .thumb img{ width:100%; display:block; }

    .xbtn{
      position:absolute;
      top:6px; right:6px;
      width:28px; height:28px;
      border-radius:10px;
      border:1px solid rgba(255,255,255,.14);
      background: rgba(0,0,0,.55);
      color:#fff;
      font-weight:900;
      cursor:pointer;
      line-height:26px;
      text-align:center;
      padding:0;
    }
    .xbtn:hover{ border-color: rgba(255,31,31,.55); }
    .thumb .cap{
      padding: 8px;
      font-size: 10px;
      color: var(--muted);
      border-top: 1px solid rgba(255,255,255,.06);
    }

    @media (max-width: 1100px){
      .grid3{ grid-template-columns: 1fr 1fr; }
      .grid2{ grid-template-columns: 1fr; }
      .brand{ min-width: auto; }
      .mapbox{ height: 300px; }
      .cam-wrap{ height: 240px; }
    }
    .dot{
  width:10px;
  height:10px;
  border-radius:50%;
  display:inline-block;
  margin-right:6px;
}

.dot.red{ background:#ff4466; }
.dot.green{ background:#00d084; }
.dot.yellow{ background:#ffcc00; }

  </style>
</head>
<body>
  <div class="header">
    <div class="brand">
      <img src="/assets/logo" alt="GDS">
      <div>
        <h1>GDS Vessel Management System</h1>
        <p>Black/Red Operational Console</p>
      </div>
    </div>
    <div class="hdr-right">
    <div class="pill" id="netStatus">MODE: --</div>
      <div class="pill">User: <b>{{ user }}</b></div>
      <div class="pill">Role: <b>{{ role }}</b></div>
      <div class="clock" id="clock">--:--:--</div>
      <a class="link" href="/logout">Logout</a>
    </div>
  </div>
<div class="sys-health">
  <span class="health-pill ok" id="gnss_health">GNSS: NO FIX</span>
  <span class="health-pill warn" id="net_health">NET: OFFLINE</span>
  <span class="health-pill ok" id="db_health">DB: LOCAL</span>
</div>

  <div class="tabs">
    
    <button class="tab-btn active" onclick="openTab('navigation')">Navigation</button>

    <button class="tab-btn" onclick="openTab('connectivity')">Connectivity & Sync</button>

    <button class="tab-btn" onclick="openTab('weather')">Weather</button>
    <button class="tab-btn" onclick="openTab('camera')">Camera/PTZ</button>
    <button class="tab-btn" onclick="openTab('vjr')">VJR</button>
    <button class="tab-btn" onclick="openTab('vdr')">VDR</button>
    <button class="tab-btn" onclick="openTab('logs')">Logs</button>
    
  </div>
  <div class="tab-content active" id="tab-navigation">

  <div class="grid2">

    <!-- LEFT: NAVIGATION FEED -->
    <div class="panel">
  <div class="panel-header">
    NAVIGATION FEED
    <span class="hint">MarineLite GNSS (Realtime)</span>
  </div>

  <div class="panel-content nav-grid">

    <!-- POSITION -->
    <div class="nav-card">
      <div class="nav-label">Latitude</div>
      <div class="nav-value" id="nav_lat">--</div>
      <div class="nav-unit"> N</div>
    </div>

    <div class="nav-card">
      <div class="nav-label">Longitude</div>
      <div class="nav-value" id="nav_lon">--</div>
      <div class="nav-unit"> E</div>
    </div>

    <!-- HEADING -->
    <div class="nav-card wide">
      <div class="nav-label">Heading</div>
      <div class="nav-value big" id="nav_heading">--</div>
      <div class="nav-unit">deg</div>
    </div>

    <!-- STATUS -->
    <div class="nav-status">
  <span>GNSS STATUS</span>
  <span class="status-bad" id="gnss_fix">NO FIX</span>
</div>


    <!-- TIME -->
    <div class="nav-time">
      Last Update:
      <span id="nav_time">--</span>
    </div>

  </div>
</div>


    <!-- RIGHT: LIVE MAP -->
    <div class="panel">
      <div class="panel-header">
        LIVE MAP
        <span class="hint">Realtime Vessel Position</span>
      </div>

      <div class="panel-content">
        <div class="mapbox" id="mapNav"></div>

        <div class="note">
          Vessel position is updated from MarineLite GNSS via local database (offline-capable).
        </div>
      </div>
    </div>

  </div>
</div>

</div>

<div class="tab-content" id="tab-connectivity">

  <div class="grid3">

    <!-- ================= NETWORK STATUS ================= -->
    <div class="panel">
      <div class="panel-header">
        Network Status
        <span class="hint">Current connectivity</span>
      </div>
      <div class="panel-content">

        <table>
          <tr><th>Interface</th><th>Status</th><th>Details</th></tr>

          <tr>
  <td>Wi-Fi</td>
  <td>
    <span class="dot red" id="wifi_dot"></span>
    <span id="wifi_status">--</span>
  </td>
  <td id="wifi_info">--</td>
</tr>


          <tr>
  <td>SIM / LTE</td>
  <td>
    <span class="dot red" id="sim_dot"></span>
    <span id="sim_status">--</span>
  </td>
  <td id="sim_info">--</td>
</tr>


          <tr>
  <td>Internet</td>
  <td>
    <span class="dot red" id="net_dot"></span>
    <span id="net_status">--</span>
  </td>
  <td id="net_info">--</td>
</tr>


          <tr>
            <td>Active Path</td>
            <td colspan="2" id="active_path">--</td>
          </tr>

        </table>

      </div>
    </div>

    <!-- ================= CONNECTION MANAGEMENT ================= -->
    <div class="panel">
      <div class="panel-header">
        Connection Management
        <span class="hint">Policy & override</span>
      </div>
      <div class="panel-content">

        <div class="radio-group">
          <label>
            <input type="radio" name="conn_mode" value="auto" checked
                   onchange="setConnMode('auto')">
            Automatic (System decides)
          </label>

          <label>
            <input type="radio" name="conn_mode" value="manual"
                   onchange="setConnMode('manual')">
            Manual (Operator selects)
          </label>
        </div>

        <div class="row2" style="margin-top:12px;">
          <button class="btn secondary" id="btn_wifi"
                  onclick="requestManualPath('wifi')" disabled>
            Use Wi-Fi
          </button>

          <button class="btn secondary" id="btn_sim"
                  onclick="requestManualPath('sim')" disabled>
            Use SIM / LTE
          </button>
        </div>

        <div class="note" style="margin-top:10px;">
          Manual selection is only allowed when policy is set to Manual.
        </div>

      </div>
    </div>

    <!-- ================= OFFLINE SYNC ================= -->
    <div class="panel">
      <div class="panel-header">
        Offline Data Sync
      </div>
      <div class="panel-content">

        <div class="data-grid">
          <div class="card">
            <div class="label">Pending Records</div>
            <div class="value" id="sync_pending">--</div>
          </div>

          <div class="card">
            <div class="label">Last Sync</div>
            <div class="value" id="sync_last">--</div>
          </div>
        </div>

        <div style="margin-top:12px;">
          <button class="btn primary" onclick="requestSync()">
            Sync Offline Records
          </button>

          <button class="btn secondary" onclick="refreshConnectivity()">
            Refresh Status
          </button>
        </div>

      </div>
    </div>

  </div>
</div>

    </div>

    

  </div>
</div>


    <!-- ================= WEATHER TAB ================= -->
<div class="tab-content" id="tab-weather">
  <div class="grid2">

    <!-- -------- LEFT: WEATHER DATA -------- -->
    <div class="panel">
      <div class="panel-header">Weather Feed</div>
      <div class="panel-content">

        <div class="data-grid">

          <!-- Core Weather -->
          <div class="card">
            <div class="label">Wind Speed (m/s)</div>
            <div class="value" id="wspd">--</div>
          </div>

          <div class="card">
            <div class="label">Wind Direction </div>
            <div class="value" id="wdir">--</div>
          </div>

          <div class="card">
            <div class="label">Temperature (C)</div>
            <div class="value" id="temp">--</div>
          </div>

          <div class="card">
            <div class="label">Humidity (%RH)</div>
            <div class="value" id="hum">--</div>
          </div>

          <div class="card">
            <div class="label">Pressure (hPa)</div>
            <div class="value" id="pres">--</div>
          </div>

          <div class="card">
            <div class="label">Last Update</div>
            <div class="value red" id="wts">--</div>
          </div>

          <!-- Environmental (from Firebase) -->
          <div class="card">
            <div class="label">Noise (dB)</div>
            <div class="value" id="noise">--</div>
          </div>

          <div class="card">
            <div class="label">PM2.5 </div>
            <div class="value" id="pm25">--</div>
          </div>

          <div class="card">
            <div class="label">PM10 </div>
            <div class="value" id="pm10">--</div>
          </div>

          <div class="card">
            <div class="label">Rainfall (mm)</div>
            <div class="value" id="rain">--</div>
          </div>

        </div>

        <div style="margin-top:10px">
          <button class="btn secondary"
                  onclick="notify('Weather feed is Active (Local DB)', 'ok')">
            Status
          </button>
        </div>

      </div>
    </div>

    <!-- -------- RIGHT: WEATHER MAP -------- -->
    <div class="panel">
      <div class="panel-header">Map (Weather Context)</div>
      <div class="panel-content">

        <div class="mapbox" id="mapWeather"></div>

        <div style="height:10px"></div>

        <div class="note">
          Weather data is linked to vessel position for contextual awareness.
        </div>

      </div>
    </div>

  </div>
</div>
<!-- =============== END WEATHER TAB =============== -->


    <!-- CAMERA TAB -->
<div class="tab-content" id="tab-camera">
  <div class="grid3">

    <!-- ================= PTZ FEED ================= -->
    <div class="panel">
      <div class="panel-header">
        PTZ Feed <span class="hint">Mode: {{ camera_mode }}</span>
      </div>

      <div class="panel-content">

        <!-- CAMERA VIEW -->
        <div class="cam-wrap">
          <img id="camImg" src="/camera/stream" alt="Camera feed">
          <div class="badge">
            CAMERA: <b id="camMode">{{ camera_mode }}</b>
          </div>
        </div>
<!-- IMAGE ENCODING + FIREBASE OPTIONS -->
<div class="row2" style="margin-top:10px; align-items:center;">
  <label class="hint" style="display:flex; gap:8px; align-items:center;">
    <input type="checkbox" id="fbUpload">
    Upload to Firebase (disabled offline)

  </label>

  <div style="display:flex; gap:10px; justify-content:flex-end; flex-wrap:wrap;">
    <label class="hint">
      JPEG Quality
      <select id="jpgQuality">
        <option value="60">60</option>
        <option value="70">70</option>
        <option value="75" selected>75</option>
        <option value="85">85</option>
      </select>
    </label>

    <label class="hint">
      Max Width
      <select id="jpgWidth">
        <option value="640">640</option>
        <option value="960">960</option>
        <option value="1280" selected>1280</option>
      </select>
    </label>
  </div>
</div>

        <!-- CAPTURE BUTTONS -->
<div class="row2" style="margin-top:12px;">
  <button class="btn secondary" onclick="captureFromPTZ('vjr')">
    CAPTURE ? VJR
  </button>

  <button class="btn secondary" onclick="captureFromPTZ('vdr')">
    CAPTURE ? VDR
  </button>

  <!-- NEW: FIREBASE CAPTURE BUTTON -->
  <button class="btn primary" onclick="captureToFirebase()">
    CAPTURE ? FIREBASE
  </button>
</div>


        <!-- DELETE BUTTONS -->
        <div class="row2" style="margin-top:10px;">
          <button class="btn danger" onclick="deleteLastCapture('vjr')">
            DELETE LAST (VJR)
          </button>
          <button class="btn danger" onclick="deleteLastCapture('vdr')">
            DELETE LAST (VDR)
          </button>
        </div>

        <!-- CONNECTION CONTROLS -->
        <div style="height:10px"></div>
        <div class="row2">
          <button class="btn secondary" onclick="reconnectCamera()">
            Reconnect
          </button>
          <button class="btn secondary" onclick="toggleCamRefresh()">
            Refresh On/Off
          </button>
        </div>

        <div style="height:10px"></div>
        <div class="note" id="camNote">
          PTZ controls are demo-safe (will respond even without hardware).
        </div>

      </div>
    </div>

    <!-- ================= PTZ CONTROLS ================= -->
    <div class="panel">
      <div class="panel-header">PTZ Controls</div>

      <div class="panel-content">

        <!-- JOYSTICK -->
        <div class="joy" id="joy">
          <div class="stick" id="stick"></div>
        </div>

        <div style="height:10px"></div>

        <!-- ZOOM -->
        <div class="row2">
          <button class="btn secondary" onclick="controlCamera('zoom_in')">
            Zoom +
          </button>
          <button class="btn secondary" onclick="controlCamera('zoom_out')">
            Zoom -
          </button>
        </div>

        <div style="height:10px"></div>

        <!-- CAMERA MODES -->
        <div class="row2">
          <button class="btn secondary" onclick="controlCamera('led_toggle')">
            LED
          </button>
          <button class="btn secondary" onclick="controlCamera('night_vision_toggle')">
            Night
          </button>
        </div>

        <div style="height:10px"></div>

        <!-- PTZ SYNC -->
        <div class="row2">
          <button class="btn secondary" onclick="setPTZSync(true)">
            Sync COG
          </button>
          <button class="btn secondary" onclick="setPTZSync(false)">
            Unsync
          </button>
        </div>

      </div>
    </div>

    <!-- ================= CAMERA STATUS ================= -->
    <div class="panel">
      <div class="panel-header">Camera Status</div>

      <div class="panel-content">
        <div class="raw" id="camStatus">Loading...</div>
        <button class="btn secondary" onclick="updateCameraStatusUI()">
          Refresh Status
        </button>
      </div>
    </div>

  </div>
</div>


    <!-- VJR TAB -->
    <div class="tab-content" id="tab-vjr">
      <div class="grid2">
        <div class="panel">
          <div class="panel-header">VJR - Journey Details</div>
          <div class="panel-content">
            <div class="form-group">
              <label class="form-label">Report Date</label>
              <input id="vjr_date" placeholder="YYYY-MM-DD">
            </div>
            <div class="form-group">
              <label class="form-label">Vessel</label>
              <input id="vjr_vessel" placeholder="Vessel name">
            </div>
            <div class="row2">
              <div class="form-group">
                <label class="form-label">IMO</label>
                <input id="vjr_imo" placeholder="IMO">
              </div>
              <div class="form-group">
                <label class="form-label">Master</label>
                <input id="vjr_master" placeholder="Master name">
              </div>
            </div>
            <div class="row2">
              <div class="form-group">
                <label class="form-label">Departure</label>
                <input id="vjr_departure" placeholder="Port/Location">
              </div>
              <div class="form-group">
                <label class="form-label">Arrival</label>
                <input id="vjr_arrival" placeholder="Port/Location">
              </div>
            </div>

            <div class="row2">
              <button class="btn secondary" onclick="captureImage('vjr')">Capture Image (VJR)</button>
              <button class="btn secondary" onclick="exportVJRPDF()">Export VJR PDF</button>
            </div>
            <div style="height:10px"></div>
            <div class="row2">
              <button class="btn danger" onclick="clearCaptures('vjr')">Clear Captures (VJR)</button>
              <button class="btn secondary" onclick="notify('VJR will include navigation log + images.', 'ok')">Info</button>
            </div>
          </div>
        </div>

        <div class="panel">
          <div class="panel-header">VJR - Captured Images <span class="hint" id="vjrCount">0</span></div>
          <div class="panel-content">
            <div class="image-grid" id="vjrGallery"></div>
            <div style="height:10px"></div>
            <div class="note">Captured images will be embedded into the VJR PDF report.</div>
          </div>
        </div>
      </div>
    </div>

    <!-- VDR TAB -->
    <div class="tab-content" id="tab-vdr">
      <div class="grid2">
        <div class="panel">
          <div class="panel-header">VDR - Daily Report Entry</div>
          <div class="panel-content">
            <div class="form-group"><label class="form-label">Date</label><input id="vdr_date" placeholder="YYYY-MM-DD"></div>
            <div class="row2">
              <div class="form-group"><label class="form-label">Vessel</label><input id="vdr_vessel" placeholder="Vessel"></div>
              <div class="form-group"><label class="form-label">IMO</label><input id="vdr_imo" placeholder="IMO"></div>
            </div>
            <div class="row2">
              <div class="form-group"><label class="form-label">MMSI</label><input id="vdr_mmsi" placeholder="MMSI"></div>
              <div class="form-group"><label class="form-label">Call Sign</label><input id="vdr_callsign" placeholder="Call Sign"></div>
            </div>
            <div class="row2">
              <div class="form-group"><label class="form-label">Client</label><input id="vdr_client" placeholder="Client"></div>
              <div class="form-group"><label class="form-label">Location</label><input id="vdr_location" placeholder="Location"></div>
            </div>
            <div class="row2">
              <div class="form-group"><label class="form-label">Country</label><input id="vdr_country" placeholder="Country"></div>
              <div class="form-group"><label class="form-label">Activity</label><input id="vdr_activity" placeholder="Activity"></div>
            </div>
            <div class="row2">
              <div class="form-group"><label class="form-label">Weather</label><input id="vdr_weather" placeholder="Weather"></div>
              <div class="form-group"><label class="form-label">Sea State</label><input id="vdr_sea_state" placeholder="Sea State"></div>
            </div>
            <div class="row2">
              <div class="form-group"><label class="form-label">Crew Count</label><input id="vdr_crew_count" placeholder="Crew"></div>
              <div class="form-group"><label class="form-label">Fuel Consumption</label><input id="vdr_fuel_consumption" placeholder="Fuel"></div>
            </div>
            <div class="form-group"><label class="form-label">Remarks</label><textarea id="vdr_remarks" placeholder="Remarks"></textarea></div>

            <div class="row2">
              <button class="btn secondary" onclick="captureImage('vdr')">Capture Image (VDR)</button>
              <button class="btn secondary" onclick="saveVDRRecord()">Save VDR Record</button>
            </div>
            <div style="height:10px"></div>
            <div class="row2">
              <button class="btn secondary" onclick="exportVDRPDF()">Export VDR PDF</button>
              <button class="btn danger" onclick="clearCaptures('vdr')">Clear Captures (VDR)</button>
            </div>
          </div>
        </div>

        <div class="panel">
          <div class="panel-header">VDR - Records & Images <span class="hint" id="vdrCount">0</span></div>
          <div class="panel-content">
            <div class="row2">
              <button class="btn secondary" onclick="exportVDRCSV()">Export CSV</button>
              <button class="btn secondary" onclick="exportVDRExcel()">Export Excel</button>
            </div>
            <div style="height:10px"></div>
            <div class="row2">
              <button class="btn danger" onclick="deleteAllVDRRecords()">Delete All Records</button>
              <button class="btn secondary" onclick="loadVDRList()">Refresh List</button>
            </div>
            <div style="height:12px"></div>
            <div class="image-grid" id="vdrGallery"></div>
            <div style="height:12px"></div>
            <table>
              <thead><tr><th>ID</th><th>Date</th><th>Vessel</th><th>Activity</th><th>Location</th><th>Weather</th></tr></thead>
              <tbody id="vdrTable"><tr><td colspan="6">No records</td></tr></tbody>
            </table>
          </div>
        </div>
      </div>
    </div>

    <!-- LOGS TAB -->
    <div class="tab-content" id="tab-logs">
      <div class="grid2">
        <div class="panel">
          <div class="panel-header">System Logs</div>
          <div class="panel-content">
            <div class="note">
              Navigation data is auto-logged into <b>nav_log.csv</b>. You can download it from the NAV tab.
              VDR records are stored in memory (demo). Images are stored in memory (demo).
            </div>
          </div>
        </div>

        <div class="panel">
          <div class="panel-header">Developer Info</div>
          <div class="panel-content">
            <div class="raw">
              WeasyPrint available: {{ 'YES' if weasy else 'NO' }}
              \nReportLab available: {{ 'YES' if True else 'NO' }}
            </div>
          </div>
        </div>
      </div>
    </div>
  </div>

  <div id="toast" class="notification" style="display:none"></div>


  <script>
    // Role helpers
    const ROLE_ORDER = {"Viewer":1,"Operator":2,"Captain":3};
    let CURRENT_ROLE = "{{ role }}";

    function can(minRole){
      return (ROLE_ORDER[CURRENT_ROLE] || 0) >= (ROLE_ORDER[minRole] || 0);
    }
function updateNetStatus(){
  const el = document.getElementById("netStatus");
  if(!el) return;

  if(navigator.onLine){
    el.innerHTML = "MODE: <b>ONLINE</b>";
    el.style.borderColor = "rgba(0,208,132,.45)";
  }else{
    el.innerHTML = "MODE: <b>OFFLINE</b>";
    el.style.borderColor = "rgba(255,204,0,.45)";
  }
}

window.addEventListener("online", updateNetStatus);
window.addEventListener("offline", updateNetStatus);
updateNetStatus();

    function notify(msg, kind="ok"){
      const t = document.getElementById("toast");
      t.className = "notification " + kind;
      t.textContent = msg;
      t.style.display = "block";
      clearTimeout(window.__toastTimer);
      window.__toastTimer = setTimeout(()=>{ t.style.display = "none"; }, 2600);
    }

    // Clock
    function tickClock(){
      const now = new Date();
      const pad = (n)=> String(n).padStart(2,"0");
      document.getElementById("clock").textContent =
        pad(now.getHours()) + ":" + pad(now.getMinutes()) + ":" + pad(now.getSeconds());
    }
    setInterval(tickClock, 1000); tickClock();

    // Tabs
    function openTab(tab){
  document.querySelectorAll(".tab-btn").forEach(b=>b.classList.remove("active"));
  document.querySelectorAll(".tab-content").forEach(c=>c.classList.remove("active"));
  document.querySelector(`.tab-btn[onclick="openTab('${tab}')"]`).classList.add("active");
  document.getElementById("tab-"+tab).classList.add("active");

// ---- NEW: tab-specific actions ----
  if(tab === "connectivity"){
    refreshConnectivity();
    refreshSyncStatus();
  }
  
  // ?? FORCE WEATHER REFRESH WHEN WEATHER TAB OPENS
  if(tab === "weather"){
    fetchWeatherData();
  }

  // map fix
  setTimeout(()=>{
    try{ mapNav && mapNav.invalidateSize(); }catch(e){}
    try{ mapWeather && mapWeather.invalidateSize(); }catch(e){}
  }, 200);
}


    // Map - NAV (MarineLite Realtime Position Only)
let mapNav, navMarker;

mapNav = L.map('mapNav').setView([3.006633, 101.380133], 12);

L.tileLayer('https://{s}.tile.openstreetmap.org/{z}/{x}/{y}.png', {
  maxZoom: 19
}).addTo(mapNav);

navMarker = L.marker([3.006633, 101.380133]).addTo(mapNav);


    // Map - Weather
    let mapWeather, weatherMarker;
    mapWeather = L.map('mapWeather').setView([3.006633, 101.380133], 12);
    L.tileLayer('https://{s}.tile.openstreetmap.org/{z}/{x}/{y}.png', { maxZoom: 19 }).addTo(mapWeather);
    weatherMarker = L.marker([3.006633, 101.380133]).addTo(mapWeather);

    // NAV CSV download
    function downloadNavCSV(){
      window.location = "/export_nav_csv";
    }

    // Camera stream refresh toggle
    let camRefresh = true;
    function toggleCamRefresh(){
      camRefresh = !camRefresh;
      notify("Camera refresh " + (camRefresh ? "ON" : "OFF"), camRefresh ? "ok" : "warn");
    }
    

    // Ensure all UI actions work (async fixes)
    async function controlCamera(action, value=0){
      if (!can("Operator")) return notify("Permission denied (Operator required).", "err");
      try{
        const res = await fetch("/camera/control", {
          method:"POST",
          headers:{ "Content-Type":"application/json" },
          body: JSON.stringify({action, value})
        });
        const j = await res.json().catch(()=>({}));
        if (res.ok){
          notify(j.message || "Control OK", "ok");
          updateCameraStatusUI();
        } else {
          notify(j.error || "Control failed", "err");
        }
      } catch(e){
        notify("Control error: " + e, "err");
      }
    }

    async function reconnectCamera(){
      if (!can("Operator")) return notify("Permission denied (Operator required).", "err");
      try{
        const res = await fetch("/camera/reconnect", { method:"POST" });
        const j = await res.json().catch(()=>({}));
        if (res.ok){
          notify("Reconnect: " + (j.reconnected ? "OK" : "FALLBACK"), j.reconnected ? "ok" : "warn");
          updateCameraStatusUI();
        } else {
          notify("Reconnect failed", "err");
        }
      } catch(e){
        notify("Reconnect error: " + e, "err");
      }
    }

    async function setPTZSync(enabled){
      if (!can("Operator")) return notify("Permission denied (Operator required).", "err");
      try{
        const res = await fetch("/camera/sync_cog", {
          method:"POST",
          headers:{ "Content-Type":"application/json" },
          body: JSON.stringify({enabled})
        });
        const j = await res.json().catch(()=>({}));
        if (res.ok){
          notify("PTZ Sync: " + (j.enabled ? "ON" : "OFF"), "ok");
          updateCameraStatusUI();
        } else {
          notify("Sync failed", "err");
        }
      } catch(e){
        notify("Sync error: " + e, "err");
      }
    }

    async function captureImage(type){
      if (!can("Operator")) return notify("Permission denied (Operator required).", "err");
      try{
        const res = await fetch("/camera/capture", {
          method:"POST",
          headers:{ "Content-Type":"application/json" },
          body: JSON.stringify({report_type:type})
        });
        const j = await res.json().catch(()=>({}));
        if (res.ok && j.status === "success"){
          notify("Captured image for " + type.toUpperCase(), "ok");
          updateCaptureCounts();
          loadCapturedImages();
        } else {
          notify("Capture failed: " + (j.error || j.message || res.status), "err");
        }
      } catch(e){
        notify("Capture error: " + e, "err");
      }
    }
function captureFromPTZ(reportType) {
// Firebase upload is disabled in offline-first architecture
const fbUpload = false;
  const jpgQuality = Number(document.getElementById("jpgQuality")?.value || 75);
  const jpgWidth = Number(document.getElementById("jpgWidth")?.value || 1280);

  fetch("/camera/capture", {
    method: "POST",
    headers: { "Content-Type": "application/json" },
    body: JSON.stringify({
      report_type: reportType,
      upload_firebase: fbUpload,
      jpg_quality: jpgQuality,
      max_width: jpgWidth
    })
  })
  .then(r => r.json())
  .then(d => {
    console.log("Capture result:", d);

    if (d.firebase_uploaded) {
      console.log("Uploaded to Firebase:", d.firebase_url);
    }
  })
  .catch(err => console.error("Capture error:", err));

  // force live feed refresh (avoid frozen frame)
  const img = document.getElementById("camImg");
  if (img) {
    img.src = "/camera/stream?t=" + Date.now();
  }
}
/* ===== ADD THIS RIGHT BELOW ===== */

function captureToFirebase() {
  fetch("/camera/capture", {
    method: "POST",
    headers: { "Content-Type": "application/json" },
    body: JSON.stringify({
      report_type: "firebase",
      upload_firebase: true,
      include_pixel_data: true,
      jpg_quality: 70,
      max_width: 1280,
      pixel_quality: 40
    })
  })
  .then(r => r.json())
  .then(d => {
    console.log("Firebase capture:", d);
    if (d.firebase_uploaded) {
      alert("? Image sent to Firebase successfully");
      console.log("Firebase URL:", d.firebase_url);
    } else {
      alert("? Capture done, but Firebase upload failed");
    }
  })
  .catch(err => {
    console.error("Firebase capture error:", err);
    alert("? Firebase capture error (see console)");
  });

  const img = document.getElementById("camImg");
  if (img) {
    img.src = "/camera/stream?t=" + Date.now();
  }
}


async function deleteLastCapture(reportType) {
  if (!can("Operator")) return notify("Permission denied (Operator required).", "err");

  try {
    const res = await fetch("/camera/captured_images_full");
    const data = await res.json().catch(() => ({}));

    const list = (reportType === "vjr") ? (data.vjr || []) : (data.vdr || []);

    if (!list.length) {
      return notify("No images to delete for " + reportType.toUpperCase(), "warn");
    }

    const payload = { report_type: reportType };

    if (list[0] && list[0].id){
      payload.capture_id = list[0].id;
    } else {
      payload.index = 0;
    }

    const del = await fetch("/camera/delete_capture", {
      method: "POST",
      headers: { "Content-Type": "application/json" },
      body: JSON.stringify(payload)
    });

    const j = await del.json().catch(() => ({}));

    if (del.ok && j.status === "success") {
      notify("Deleted last image (" + reportType.toUpperCase() + ")", "ok");
      updateCaptureCounts();
      loadCapturedImages();
    } else {
      notify("Delete failed", "err");
    }
  } catch (e) {
    notify("Delete error: " + e, "err");
  }
}


    async function updateCaptureCounts(){
      try{
        const res = await fetch("/camera/captured_images");
        const j = await res.json().catch(()=>({}));
        document.getElementById("vjrCount").textContent = j.vjr_images || 0;
        document.getElementById("vdrCount").textContent = j.vdr_images || 0;
      }catch(e){}
    }

    async function deleteCapture(type, captureId, index){
      if (!can("Operator")) return notify("Permission denied (Operator required).", "err");
      try{
        const payload = { report_type: type };

        // Prefer stable delete by capture_id. Fall back to index for old items.
        if (captureId && String(captureId).trim().length > 0){
          payload.capture_id = captureId;
          console.log("Deleting", type, "capture_id", captureId);
        } else {
          payload.index = index;
          console.log("Deleting", type, "index", index);
        }

        const res = await fetch("/camera/delete_capture", {
          method:"POST",
          headers:{ "Content-Type":"application/json" },
          body: JSON.stringify(payload)
        });

        const j = await res.json().catch(()=>({}));
        console.log("Delete response:", res.ok, j);

        if (res.ok && j.status === "success"){
          notify("Deleted image", "ok");
          updateCaptureCounts();
          loadCapturedImages();
        } else {
          notify("Delete failed: " + (j.error || j.message || res.status), "err");
        }
      } catch(e){
        notify("Delete error: " + e, "err");
      }
    }

    async function clearCaptures(type){
      if (!can("Operator")) return notify("Permission denied (Operator required).", "err");
      if (!confirm("Clear all captured images for " + type.toUpperCase() + "?")) return;
      try{
        const res = await fetch("/camera/clear_captures", {
          method:"POST",
          headers:{ "Content-Type":"application/json" },
          body: JSON.stringify({report_type:type})
        });
        const j = await res.json().catch(()=>({}));
        console.log("Clear response:", res.ok, j);
        if (res.ok && j.status === "success"){
          notify("Cleared captures for " + type.toUpperCase(), "ok");
          updateCaptureCounts();
          loadCapturedImages();
        } else {
          notify("Clear failed: " + (j.error || j.message || res.status), "err");
        }
      } catch(e){
        notify("Clear error: " + e, "err");
      }
    }
    async function loadCapturedImages(){
      try{
        const res = await fetch("/camera/captured_images_full");
        const j = await res.json().catch(()=>({}));
        const vjr = j.vjr || [];
        const vdr = j.vdr || [];

        const vjrG = document.getElementById("vjrGallery");
        const vdrG = document.getElementById("vdrGallery");
        vjrG.innerHTML = "";
        vdrG.innerHTML = "";

        vjr.forEach((img, idx)=>{
          const d = document.createElement("div");
          d.className = "thumb";
          d.innerHTML = `
            <img src="data:image/jpeg;base64,${img.data}" />
<button class="xbtn" title="Delete" onclick="deleteCapture('vjr', '${img.id || ''}', ${idx})">&times;</button>
            <div class="cap">${img.timestamp || ""}</div>
          `;
          vjrG.appendChild(d);
        });

        vdr.forEach((img, idx)=>{
          const d = document.createElement("div");
          d.className = "thumb";
          d.innerHTML = `
            <img src="data:image/jpeg;base64,${img.data}" />
<button class="xbtn" title="Delete" onclick="deleteCapture('vdr', '${img.id || ''}', ${idx})">&times;</button>
            <div class="cap">${img.timestamp || ""}</div>
          `;
          vdrG.appendChild(d);
        });

      }catch(e){
        console.warn("loadCapturedImages error", e);
      }
    }

    async function updateCameraStatusUI(){
      try{
        const res = await fetch("/camera/status");
        const j = await res.json().catch(()=>({}));
        document.getElementById("camMode").textContent = (j.mode || "demo").toUpperCase();
        document.getElementById("camStatus").textContent = JSON.stringify(j, null, 2);
        const note = document.getElementById("camNote");
        if (j.mode === "ip"){
          note.textContent = "Connected to IP camera. Snapshot capture uses first JPEG frame extraction.";
        } else {
          note.textContent = "Demo mode (no live camera). Controls remain functional.";
        }
      }catch(e){
        document.getElementById("camStatus").textContent = "Status error: " + e;
      }
    }

    // NAV + WEATHER polling
    async function fetchNavigationData(){
  try{
    const res = await fetch("/nav_data");
    const n = await res.json().catch(()=>({}));

    if(!n.latitude || !n.longitude) return;

    // Update NAV text fields
    document.getElementById("nav_lat").textContent =
  Math.abs(Number(n.latitude)).toFixed(6) + (Number(n.latitude) >= 0 ? " N" : " S");

document.getElementById("nav_lon").textContent =
  Math.abs(Number(n.longitude)).toFixed(6) + (Number(n.longitude) >= 0 ? " E" : " W");

    const hdg = Number(n.heading);
document.getElementById("nav_heading").textContent =
  Number.isFinite(hdg) ? `${hdg.toFixed(0)}` : "N/A";


    if(n.timestamp){
      const t = new Date(n.timestamp * 1000);
      document.getElementById("nav_time").textContent = t.toLocaleTimeString();
    }
// GNSS status indicator (no UI change)
const fix = document.getElementById("nav_fix");
if(!n.timestamp){
  fix.textContent = "NO DATA";
}else{
  const age = Math.floor(Date.now()/1000 - Number(n.timestamp));
  if(age <= 5) fix.textContent = "LIVE";
  else if(age <= 30) fix.textContent = "STALE";
  else fix.textContent = "NO FIX";
}

    // Update map marker
    navMarker.setLatLng([n.latitude, n.longitude]);
    mapNav.panTo([n.latitude, n.longitude], { animate: true });
// Sync Weather map position with NAV (GNSS)
weatherMarker.setLatLng([n.latitude, n.longitude]);
mapWeather.panTo([n.latitude, n.longitude], { animate: true });


  }catch(e){
    console.error("NAV fetch failed", e);
  }
}


    async function fetchWeatherData(){
  try{
    const res = await fetch("/weather_data");
    const j = await res.json();

    document.getElementById("wspd").innerText  = j.wind_speed ?? "--";
    document.getElementById("wdir").innerText  = j.wind_direction ?? "--";
    document.getElementById("temp").innerText  = j.temperature ?? "--";
    document.getElementById("hum").innerText   = j.humidity ?? "--";
    document.getElementById("pres").innerText  = j.pressure ?? "--";
    document.getElementById("noise").innerText = j.noise ?? "--";
    document.getElementById("pm25").innerText  = j.pm25 ?? "--";
    document.getElementById("pm10").innerText  = j.pm10 ?? "--";
    document.getElementById("rain").innerText  = j.rainfall ?? "--";

    if(j.timestamp){
      document.getElementById("wts").innerText =
        new Date(j.timestamp * 1000).toLocaleString();
    }else{
      document.getElementById("wts").innerText = "--";
    }

  }catch(e){
    console.error("Weather fetch failed", e);
  }
}


    // Export VJR PDF
    async function exportVJRPDF(){
      if (!can("Operator")) return notify("Permission denied (Operator required).", "err");

      const vjr = {
        date: document.getElementById("vjr_date").value,
        vessel: document.getElementById("vjr_vessel").value,
        imo: document.getElementById("vjr_imo").value,
        master: document.getElementById("vjr_master").value,
        departure: document.getElementById("vjr_departure").value,
        arrival: document.getElementById("vjr_arrival").value
      };

      // take nav samples from the live endpoint (most recent)
      let nav = [];
      try{
        const resNav = await fetch("/nav_data");
        const jNav = await resNav.json().catch(()=>({}));
        nav = (jNav.history || []).slice(0, 40);
      }catch(e){}

      try{
        const res = await fetch("/export_vjr_pdf", {
          method:"POST",
          headers:{ "Content-Type":"application/json" },
          body: JSON.stringify({vjr, nav})
        });
        if (!res.ok){
          const j = await res.json().catch(()=>({}));
          return notify("PDF failed: " + (j.error || res.status), "err");
        }
        const blob = await res.blob();
        const url = URL.createObjectURL(blob);
        const a = document.createElement("a");
        a.href = url;
        a.download = "VJR_Report.pdf";
        document.body.appendChild(a);
        a.click();
        a.remove();
        URL.revokeObjectURL(url);
        notify("VJR PDF downloaded.", "ok");
      }catch(e){
        notify("PDF error: " + e, "err");
      }
    }

    // VDR record operations
    async function saveVDRRecord(){
      if (!can("Operator")) return notify("Permission denied (Operator required).", "err");

      const record = {
        date: document.getElementById("vdr_date").value,
        vessel: document.getElementById("vdr_vessel").value,
        imo: document.getElementById("vdr_imo").value,
        mmsi: document.getElementById("vdr_mmsi").value,
        callsign: document.getElementById("vdr_callsign").value,
        client: document.getElementById("vdr_client").value,
        location: document.getElementById("vdr_location").value,
        country: document.getElementById("vdr_country").value,
        activity: document.getElementById("vdr_activity").value,
        weather: document.getElementById("vdr_weather").value,
        sea_state: document.getElementById("vdr_sea_state").value,
        crew_count: document.getElementById("vdr_crew_count").value,
        fuel_consumption: document.getElementById("vdr_fuel_consumption").value,
        remarks: document.getElementById("vdr_remarks").value
      };

      try{
        const res = await fetch("/save_vdr", {
          method:"POST",
          headers:{ "Content-Type":"application/json" },
          body: JSON.stringify(record)
        });
        const j = await res.json().catch(()=>({}));
        if (res.ok){
          notify("VDR record saved. Total: " + (j.total || 0), "ok");
          loadVDRList();
        }else{
          notify("Save failed: " + (j.error || res.status), "err");
        }
      }catch(e){
        notify("Save error: " + e, "err");
      }
    }

    async function loadVDRList(){
      try{
        const res = await fetch("/vdr_records");
        const rows = await res.json().catch(()=>[]);
        const tbody = document.getElementById("vdrTable");
        if (!rows.length){
          tbody.innerHTML = `<tr><td colspan="6">No records</td></tr>`;
          return;
        }
        tbody.innerHTML = rows.slice(0, 30).map(r => `
          <tr>
            <td>${r.id || ""}</td>
            <td>${r.date || ""}</td>
            <td>${r.vessel || ""}</td>
            <td>${r.activity || ""}</td>
            <td>${r.location || ""}</td>
            <td>${r.weather || ""}</td>
          </tr>
        `).join("");
      }catch(e){}
    }

    async function exportVDRCSV(){
      if (!can("Operator")) return notify("Permission denied (Operator required).", "err");
      try{
        const resList = await fetch("/vdr_records");
        const records = await resList.json().catch(()=>[]);
        const res = await fetch("/export_vdr_csv", {
          method:"POST",
          headers:{ "Content-Type":"application/json" },
          body: JSON.stringify({records})
        });
        const blob = await res.blob();
        const url = URL.createObjectURL(blob);
        const a = document.createElement("a");
        a.href = url;
        a.download = "VDR_Report.csv";
        document.body.appendChild(a);
        a.click();
        a.remove();
        URL.revokeObjectURL(url);
        notify("CSV downloaded.", "ok");
      }catch(e){
        notify("CSV error: " + e, "err");
      }
    }

    async function exportVDRExcel(){
      if (!can("Operator")) return notify("Permission denied (Operator required).", "err");
      try{
        const resList = await fetch("/vdr_records");
        const records = await resList.json().catch(()=>[]);
        const res = await fetch("/export_vdr_excel", {
          method:"POST",
          headers:{ "Content-Type":"application/json" },
          body: JSON.stringify({records})
        });
        if (!res.ok){
          const j = await res.json().catch(()=>({}));
          return notify("Excel failed: " + (j.error || res.status), "err");
        }
        const blob = await res.blob();
        const url = URL.createObjectURL(blob);
        const a = document.createElement("a");
        a.href = url;
        a.download = "VDR_Report.xlsx";
        document.body.appendChild(a);
        a.click();
        a.remove();
        URL.revokeObjectURL(url);
        notify("Excel downloaded.", "ok");
      }catch(e){
        notify("Excel error: " + e, "err");
      }
    }

    async function exportVDRPDF(){
      if (!can("Operator")) return notify("Permission denied (Operator required).", "err");
      try{
        const resList = await fetch("/vdr_records");
        const records = await resList.json().catch(()=>[]);
        const res = await fetch("/export_vdr_pdf", {
          method:"POST",
          headers:{ "Content-Type":"application/json" },
          body: JSON.stringify({records})
        });
        if (!res.ok){
          const j = await res.json().catch(()=>({}));
          return notify("PDF failed: " + (j.error || res.status), "err");
        }
        const blob = await res.blob();
        const url = URL.createObjectURL(blob);
        const a = document.createElement("a");
        a.href = url;
        a.download = "VDR_Report.pdf";
        document.body.appendChild(a);
        a.click();
        a.remove();
        URL.revokeObjectURL(url);
        notify("VDR PDF downloaded.", "ok");
      }catch(e){
        notify("PDF error: " + e, "err");
      }
    }

    async function deleteAllVDRRecords(){
      if (!can("Captain")) return notify("Permission denied (Captain required).", "err");
      if (!confirm("Delete ALL VDR records?")) return;
      try{
        const res = await fetch("/clear_vdr", { method:"POST" });
        const j = await res.json().catch(()=>({}));
        if (res.ok){
          notify("All VDR records deleted.", "ok");
          loadVDRList();
        } else {
          notify("Delete failed: " + (j.error || res.status), "err");
        }
      }catch(e){
        notify("Delete error: " + e, "err");
      }
    }

    // Joystick
    const joy = document.getElementById("joy");
    const stick = document.getElementById("stick");
    let dragging = false;

    function setStick(dx, dy){
      const r = 60;
      const len = Math.min(r, Math.hypot(dx, dy));
      const ang = Math.atan2(dy, dx);
      const x = len * Math.cos(ang);
      const y = len * Math.sin(ang);
      stick.style.transform = `translate(${x}px, ${y}px)`;
      return {x, y};
    }

    async function up(dx, dy){
      // map stick to PTZ actions
      const ax = Math.abs(dx), ay = Math.abs(dy);
      if (ax < 10 && ay < 10) return;
      if (ax > ay){
        await controlCamera(dx > 0 ? "pan_right" : "pan_left");
      }else{
        await controlCamera(dy > 0 ? "tilt_down" : "tilt_up");
      }
    }

    joy.addEventListener("pointerdown", (e)=>{
      dragging = true;
      joy.setPointerCapture(e.pointerId);
    });
    joy.addEventListener("pointerup", ()=>{
      dragging = false;
      stick.style.transform = "translate(-50%,-50%)";
    });
    joy.addEventListener("pointermove", (e)=>{
      if (!dragging) return;
      const rect = joy.getBoundingClientRect();
      const dx = e.clientX - (rect.left + rect.width/2);
      const dy = e.clientY - (rect.top + rect.height/2);
      setStick(dx, dy);
      up(dx, dy);
    });

    // Init
    updateCameraStatusUI();
    updateCaptureCounts();
    loadCapturedImages();
    loadVDRList();

    setInterval(fetchNavigationData, 1000);
    setInterval(fetchWeatherData, 2000);
    fetchNavigationData(); fetchWeatherData();
  /* ================= PTZ JOYSTICK LOGIC ================= */

let joyActive = false;
let joyCenter = { x: 0, y: 0 };

document.addEventListener("DOMContentLoaded", () => {
  const joy = document.getElementById("joy");
  const stick = document.getElementById("stick");

  if (!joy || !stick) return;

  // Start dragging
  joy.addEventListener("pointerdown", (e) => {
    joyActive = true;
    const r = joy.getBoundingClientRect();
    joyCenter = {
      x: r.left + r.width / 2,
      y: r.top + r.height / 2
    };
    joy.setPointerCapture(e.pointerId);
  });

  // Drag move
  document.addEventListener("pointermove", (e) => {
    if (!joyActive) return;

    let dx = e.clientX - joyCenter.x;
    let dy = e.clientY - joyCenter.y;

    const max = 45; // joystick radius
    dx = Math.max(-max, Math.min(max, dx));
    dy = Math.max(-max, Math.min(max, dy));

    stick.style.transform = `translate(${dx}px, ${dy}px)`;

    // Send PTZ commands
    if (Math.abs(dx) > 12) {
      controlCamera(dx > 0 ? "pan_right" : "pan_left");
    }
    if (Math.abs(dy) > 12) {
      controlCamera(dy > 0 ? "tilt_down" : "tilt_up");
    }
  });

  // Release joystick
  document.addEventListener("pointerup", () => {
    if (!joyActive) return;
    joyActive = false;
    stick.style.transform = "translate(-50%, -50%)";
  });
});
document.addEventListener("DOMContentLoaded", () => {
    fetchWeatherData();
    setInterval(fetchWeatherData, 2000);
  });
  let currentConnMode = "auto";

/* ---- Refresh connectivity status ---- */
function refreshConnectivity(){
  fetch("/api/connectivity/status")
    .then(r => r.json())
    .then(d => {
      // Wi-Fi
      document.getElementById("wifi_status").innerText =
        d.wifi.connected ? "Connected" : "Disconnected";
        document.getElementById("wifi_dot").className =
  "dot " + (d.wifi.connected ? "green" : "red");

      document.getElementById("wifi_info").innerText =
        d.wifi.ssid || "--";

      // SIM
      document.getElementById("sim_status").innerText =
        d.sim.connected ? "Connected" : "Disconnected";
        document.getElementById("sim_dot").className =
  "dot " + (d.sim.connected ? "green" : "red");

      document.getElementById("sim_info").innerText =
        d.sim.operator || "--";

      // Internet
      document.getElementById("net_status").innerText =
        d.internet ? "Available" : "Offline";
        document.getElementById("net_dot").className =
  "dot " + (d.internet ? "green" : "red");

      document.getElementById("net_info").innerText =
        d.internet ? "Online" : "No route";

      // Active path
      document.getElementById("active_path").innerText =
        d.active_path || "None";

      // Policy
      currentConnMode = d.policy?.mode || "auto";
      document.querySelector(
        `input[name="conn_mode"][value="${currentConnMode}"]`
      ).checked = true;

      updateManualButtons();
    })
    .catch(()=>{
      notify("Failed to read connectivity status", "error");
    });
}


function refreshSyncStatus(){
  fetch("/api/sync/status")
    .then(r => r.json())
    .then(d => {
      document.getElementById("sync_pending").innerText =
        d.pending_records;
      document.getElementById("sync_last").innerText =
        d.last_sync || "--";
    })
    .catch(()=>{});
}

function requestSync(){
  fetch("/api/sync/start", {method:"POST"})
    .then(r => r.json())
    .then(d => notify(d.message, "info"))
    .catch(()=>notify("Sync request failed","error"));
}


/* ---- Auto / Manual toggle ---- */
function setConnMode(mode){
  currentConnMode = mode;
  updateManualButtons();

  fetch("/api/connectivity/policy", {
    method: "POST",
    headers: {"Content-Type":"application/json"},
    body: JSON.stringify({ mode: mode })
  }).catch(()=>{});
}

/* ---- Enable / disable buttons ---- */
function updateManualButtons(){
  const manual = (currentConnMode === "manual");
  document.getElementById("btn_wifi").disabled = !manual;
  document.getElementById("btn_sim").disabled = !manual;
}

/* ---- Manual Wi-Fi / SIM request ---- */
function requestManualPath(path){
  fetch("/api/connectivity/policy", {
    method: "POST",
    headers: {"Content-Type":"application/json"},
    body: JSON.stringify({
      mode: "manual",
      preferred: path
    })
  })
  .then(()=>{
    notify(`Manual path requested: ${path.toUpperCase()}`, "info");
  })
  .catch(()=>{
    notify("Request failed", "error");
  });
}
  </script>
</body>
</html>
'''
# ======================================================
# Firebase initialization & background listeners
# ======================================================

# Initialize Firebase ONCE before starting any listeners
""" if not init_firebase():
    print("? Firebase failed to initialize at startup")
else:
    print("? Firebase initialized at startup")

# Start listeners AFTER Firebase init
threading.Thread(
    target=firebase_weather_listener,
    daemon=True
).start()

threading.Thread(
    target=firebase_marinelite_listener,
    daemon=True
).start() """
""" def firebase_poll_loop():
    global nav_current, weather_current

    print("?? Firebase polling loop started")

    while True:
        try:
            nav = db.reference(
                "gsd_vessel_sensor_data/marinelite/current"
            ).get()

            if isinstance(nav, dict):
                with nav_lock:
                    nav_current["latitude"]  = nav.get("latitude")
                    nav_current["longitude"] = nav.get("longitude")
                    nav_current["heading"]   = nav.get("heading")
                    nav_current["timestamp"] = nav.get("timestamp")

            weather = db.reference(
                "gsd_vessel_sensor_data/weather/current"
            ).get()

            if isinstance(weather, dict):
                with weather_lock:
                    weather_current.update({
                        "wind_speed": weather.get("wind_speed"),
                        "wind_direction": weather.get("wind_dir"),
                        "temperature": weather.get("temperature"),
                        "humidity": weather.get("humidity"),
                        "pressure": weather.get("pressure"),
                        "noise": weather.get("noise"),
                        "pm25": weather.get("pm25"),
                        "pm10": weather.get("pm10"),
                        "rainfall": weather.get("rainfall"),
                        "timestamp": weather.get("timestamp"),
                    })

            print("? Firebase polled")

        except Exception as e:
            print("? Firebase poll error:", e)

        time.sleep(1)  # 1 Hz is perfect """

# ------------------------------------------------------------------------------
# MAIN
# ------------------------------------------------------------------------------

if __name__ == "__main__":
    init_camera()
    app.run(host="0.0.0.0", port=5000, debug=False)



